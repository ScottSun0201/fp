#!/usr/bin/env python3
"""Generate the unified supplier statement Excel template."""
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
OUTPUT = ROOT / "templates" / "对账单统一模板_v1.xlsx"


def style_cell(cell, fill=None, bold=False, color="111827"):
    cell.font = Font(name="Arial", size=10, bold=bold, color=color)
    cell.alignment = Alignment(vertical="center", wrap_text=True)
    if fill:
        cell.fill = fill


def apply_border(ws, cell_range):
    thin = Side(style="thin", color="D1D5DB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws[cell_range]:
        for cell in row:
            cell.border = border


def build_template():
    wb = Workbook()
    ws = wb.active
    ws.title = "对账单"
    guide = wb.create_sheet("填写说明")

    title_fill = PatternFill("solid", fgColor="1F4E78")
    section_fill = PatternFill("solid", fgColor="D9EAF7")
    required_fill = PatternFill("solid", fgColor="FFF2CC")
    header_fill = PatternFill("solid", fgColor="E2F0D9")
    locked_fill = PatternFill("solid", fgColor="F3F4F6")
    required_font_color = "C00000"

    # Main sheet
    ws.merge_cells("A1:N1")
    ws["A1"] = "统一对账单模板 v1"
    ws["A1"].font = Font(name="Arial", size=16, bold=True, color="FFFFFF")
    ws["A1"].fill = title_fill
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    base_rows = [
        ("A3", "* 供应商名称", "B3", ""),
        ("C3", "* 对账月份", "D3", "2026-06"),
        ("E3", "* 客户名称", "F3", "安徽骊威科技集团有限公司"),
        ("G3", "制表日期", "H3", ""),
        ("I3", "结算天数", "J3", 30),
        ("A4", "供应商税号", "B4", ""),
        ("C4", "联系人", "D4", ""),
        ("E4", "联系电话", "F4", ""),
        ("G4", "备注", "H4", ""),
    ]
    for label_cell, label, value_cell, hint in base_rows:
        ws[label_cell] = label
        ws[label_cell].fill = section_fill
        ws[label_cell].font = Font(
            name="Arial",
            size=10,
            bold=True,
            color=required_font_color if label.startswith("*") else "111827",
        )
        ws[value_cell] = hint
        ws[value_cell].fill = required_fill if "*" in label else locked_fill
        ws[value_cell].alignment = Alignment(vertical="center", wrap_text=True)

    ws["H3"] = '=TEXT(TODAY(),"yyyy-mm-dd")'

    ws.merge_cells("A7:N7")
    ws["A7"] = "明细区：从第 9 行开始填写。黄色表头为必填；金额列可用公式，也可直接填写。"
    ws["A7"].fill = section_fill
    ws["A7"].font = Font(name="Arial", size=10, bold=True)

    headers = [
        "序号",
        "客户订单号",
        "* 客户物料编码",
        "供应商物料编码",
        "送货单号",
        "* 交货日期",
        "* 商品名称",
        "规格型号",
        "* 单位",
        "* 数量",
        "* 单价",
        "* 金额",
        "税率",
        "备注",
    ]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=8, column=col, value=header)
        required = header.startswith("*")
        cell.fill = required_fill if required else header_fill
        cell.font = Font(name="Arial", size=10, bold=True, color=required_font_color if required else "111827")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row in range(9, 209):
        ws.cell(row=row, column=1, value=f'=IF(C{row}<>"",ROW()-8,"")')
        ws.cell(row=row, column=9, value="PCS")
        ws.cell(row=row, column=12, value=f'=IF(AND(J{row}<>"",K{row}<>""),ROUND(J{row}*K{row},2),"")')
        ws.cell(row=row, column=13, value="13%")

    summary_row = 210
    ws.cell(row=summary_row, column=9, value="合计")
    ws.cell(row=summary_row, column=10, value=f"=SUM(J9:J{summary_row - 1})")
    ws.cell(row=summary_row, column=12, value=f"=SUM(L9:L{summary_row - 1})")
    for col in range(1, 15):
        cell = ws.cell(row=summary_row, column=col)
        cell.fill = section_fill
        cell.font = Font(name="Arial", size=10, bold=True)

    widths = [8, 18, 18, 18, 18, 14, 22, 20, 10, 12, 14, 14, 10, 24]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    ws.freeze_panes = "A9"
    ws.auto_filter.ref = "A8:N209"
    apply_border(ws, "A3:N210")

    # Validations
    month_validation = DataValidation(type="custom", formula1='=AND(LEN(D3)=7,MID(D3,5,1)="-",ISNUMBER(VALUE(LEFT(D3,4))),ISNUMBER(VALUE(RIGHT(D3,2))))')
    month_validation.error = "对账月份请填写 YYYY-MM，例如 2026-06"
    ws.add_data_validation(month_validation)
    month_validation.add(ws["D3"])

    unit_validation = DataValidation(type="list", formula1='"PCS,KPCS,千个,个,只,套,米,KG"')
    ws.add_data_validation(unit_validation)
    unit_validation.add("I9:I209")

    # Guide sheet
    guide_rows = [
        ("必填项", "填写要求"),
        ("供应商名称", "填写贵司完整公司名称。"),
        ("对账月份", "格式 YYYY-MM，例如 2026-06。"),
        ("客户名称", "默认安徽骊威科技集团有限公司，不要改成供应商名称。"),
        ("客户物料编码", "填写我方料号。"),
        ("交货日期", "格式 YYYY-MM-DD。"),
        ("商品名称", "填写物料或商品名称。"),
        ("单位", "默认 PCS；如 KPCS/千个请直接选择对应单位。"),
        ("数量、单价、金额", "按含税口径填写，金额保留 2 位。"),
    ]
    for row_idx, row in enumerate(guide_rows, start=1):
        for col_idx, value in enumerate(row, start=1):
            cell = guide.cell(row=row_idx, column=col_idx, value=value)
            style_cell(cell, fill=header_fill if row_idx == 1 else None, bold=row_idx == 1)
    guide.column_dimensions["A"].width = 18
    guide.column_dimensions["B"].width = 80
    apply_border(guide, f"A1:B{len(guide_rows)}")

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT)
    return OUTPUT


if __name__ == "__main__":
    print(build_template())
