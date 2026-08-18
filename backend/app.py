# -*- coding: utf-8 -*-
"""
FP进销存财务系统 - Flask主应用
REQ-039: 端口统一8090
REQ-042: 集成全部模块 (config/models/invoice_parser/statement_parser/matching_engine/export_utils)
"""
import os
import io
import json
import logging
import csv
import hashlib
import re
import secrets
import shutil
import tempfile
import time
from datetime import datetime
from pathlib import Path
from flask import Flask, jsonify, redirect, request, send_file, send_from_directory, session
from flask_cors import CORS
from werkzeug.utils import secure_filename
try:
    from dotenv import load_dotenv
except ImportError:
    load_dotenv = None

if load_dotenv:
    load_dotenv()

from config import (
    SECRET_KEY, UPLOAD_DIR, MAX_CONTENT_LENGTH,
    SESSION_LIFETIME_HOURS, DB_ENGINE,
    ERP_MYSQL_CONFIG, QWEN_OCR_MODEL,
)
from models import init_db, get_db, audit_log, rows_to_list, dict_from_row
from invoice_parser import parse_invoice_pdf
from statement_parser import (
    parse_statement_image, parse_statement_pdf, parse_statement_xls,
    parse_statement_xlsx, recognize_excel_supplier_locally,
    recognize_supplier_locally,
)
from qwen_ocr import recognize_supplier as recognize_supplier_with_qwen
from matching_engine import match_invoice_statement
from export_utils import (
    export_invoices_csv, export_invoices_excel,
    export_statements_csv, export_statements_excel,
    export_match_results_csv,
)
import delivery_compare
import feedback_engine

# ─── 日志 ───
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s [%(levelname)s] %(name)s: %(message)s'
)
logger = logging.getLogger('fp-app')


class StatementIdentityConflict(ValueError):
    def __init__(self, existing_id):
        super().__init__("该合作商本月对账单已存在，但文件内容发生变化，请对原记录使用“重新比对”")
        self.existing_id = existing_id

# ─── Flask 初始化 ───
app = Flask(__name__)
app.config['SECRET_KEY'] = SECRET_KEY
app.config['MAX_CONTENT_LENGTH'] = MAX_CONTENT_LENGTH
CORS(app)
PENDING_STATEMENT_IMPORTS = {}
PENDING_STATEMENT_DIR = UPLOAD_DIR / 'pending_statements'
PENDING_STATEMENT_DIR.mkdir(exist_ok=True)
PROGRESS_DIR = UPLOAD_DIR / 'progress'
PROGRESS_DIR.mkdir(exist_ok=True)
STATEMENT_PARSE_CACHE_DIR = UPLOAD_DIR / 'statement_parse_cache'
STATEMENT_PARSE_CACHE_DIR.mkdir(exist_ok=True)
DELIVERY_COMPARE_DIR = Path('/Users/liweitas001/Downloads/快递对比')
DELIVERY_OUTPUT_DIR = UPLOAD_DIR / 'delivery_results'
DELIVERY_OUTPUT_DIR.mkdir(exist_ok=True)
STATEMENT_RECORD_DIR = UPLOAD_DIR / 'statement_records'
STATEMENT_RECORD_DIR.mkdir(exist_ok=True)


class PersistentProgress(dict):
    """Keep task progress available across browser changes and service restarts."""
    max_running_age = 15 * 60

    @staticmethod
    def _path(task_id):
        if not re.fullmatch(r'[A-Za-z0-9_-]{8,120}', str(task_id or '')):
            return None
        return PROGRESS_DIR / f'{task_id}.json'

    def __setitem__(self, task_id, value):
        payload = dict(value or {})
        payload['updated_at'] = time.time()
        super().__setitem__(task_id, payload)
        path = self._path(task_id)
        if path:
            temp = path.with_suffix('.tmp')
            temp.write_text(json.dumps(payload, ensure_ascii=False), encoding='utf-8')
            temp.replace(path)

    def get(self, task_id, default=None):
        value = super().get(task_id)
        if value is None:
            path = self._path(task_id)
            if path and path.is_file():
                try:
                    value = json.loads(path.read_text(encoding='utf-8'))
                    super().__setitem__(task_id, value)
                except (OSError, ValueError, TypeError):
                    value = None
        if value and value.get('status') == 'running':
            updated_at = float(value.get('updated_at') or 0)
            if updated_at and time.time() - updated_at > self.max_running_age:
                value = dict(value, status='error', percent=100,
                             message='任务已超时或服务曾重启，请重新提交')
                self[task_id] = value
        return value if value is not None else default


PROGRESS = PersistentProgress()


def _pending_statement_path(token):
    if not re.fullmatch(r'[A-Za-z0-9_-]{20,80}', str(token or '')):
        return None
    return PENDING_STATEMENT_DIR / f'{token}.json'


def _save_pending_statement(token, value):
    PENDING_STATEMENT_IMPORTS[token] = value
    path = _pending_statement_path(token)
    if path:
        temp_path = path.with_suffix('.tmp')
        temp_path.write_text(
            json.dumps(value, ensure_ascii=False), encoding='utf-8'
        )
        temp_path.replace(path)


def _load_pending_statement(token):
    pending = PENDING_STATEMENT_IMPORTS.get(token)
    if pending:
        return pending
    path = _pending_statement_path(token)
    if not path or not path.is_file():
        return None
    try:
        pending = json.loads(path.read_text(encoding='utf-8'))
        if time.time() - float(pending.get('created_at') or 0) > 86400:
            path.unlink(missing_ok=True)
            return None
        PENDING_STATEMENT_IMPORTS[token] = pending
        return pending
    except Exception:
        logger.exception("读取待确认对账单状态失败 token=%s", token)
        return None


def _delete_pending_statement(token):
    PENDING_STATEMENT_IMPORTS.pop(token, None)
    path = _pending_statement_path(token)
    if path:
        path.unlink(missing_ok=True)

# ─── 启动时初始化数据库 ───
init_db()

# ─── 前端静态文件 ───
FRONTEND_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', 'frontend')

@app.route('/')
def index():
    """LWFP 首页"""
    return send_from_directory(FRONTEND_DIR, 'dashboard.html')

# ─── 通用 HTML 静态文件路由 ───
@app.route('/<path:filename>')
def serve_static(filename):
    """服务前端静态 HTML 文件"""
    if filename.endswith('.html'):
        return send_from_directory(FRONTEND_DIR, filename)
    return send_from_directory(FRONTEND_DIR, filename)


@app.route('/login', methods=['GET', 'POST'])
def login_page():
    """测试环境登录页，兼容 107 表单登录。"""
    if request.method == 'GET':
        return send_from_directory(FRONTEND_DIR, 'login.html')
    username = request.form.get('username', '').strip()
    password = request.form.get('password', '')
    import bcrypt
    with get_db() as conn:
        user = conn.execute(
            "SELECT * FROM sys_user WHERE username=? AND is_active=1", (username,)
        ).fetchone()
        if user and bcrypt.checkpw(password.encode('utf-8'), user['password_hash'].encode('utf-8')):
            conn.execute("UPDATE sys_user SET login_attempts=0, locked_until=NULL WHERE id=?", (user['id'],))
            audit_log(conn, user['id'], username, 'LOGIN', 'user', user['id'], ip=_get_client_ip())
            session['username'] = username
            session['role'] = user['role']
            session['user_id'] = user['id']
            return redirect('/')
    return send_from_directory(FRONTEND_DIR, 'login.html'), 401


@app.route('/management')
def management_page():
    """LWFP 对账管理页"""
    response = send_from_directory(FRONTEND_DIR, 'management.html')
    response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
    response.headers['Pragma'] = 'no-cache'
    return response


@app.route('/supplier-summary')
def supplier_summary_page():
    """供应商统计页"""
    return send_from_directory(FRONTEND_DIR, 'supplier_summary.html')


@app.route('/delivery')
def delivery_page():
    """快递对账页"""
    if not _can_use_delivery():
        return redirect('/login')
    return send_from_directory(FRONTEND_DIR, 'delivery.html')


@app.route('/warnings')
def warnings_page():
    """超时预警页"""
    return send_from_directory(FRONTEND_DIR, 'warnings.html')


@app.route('/supplier-progress')
def supplier_progress_page():
    """供应商进度看板页"""
    return send_from_directory(FRONTEND_DIR, 'supplier_progress.html')


@app.route('/logout')
def logout_page():
    session.clear()
    return redirect('/login')


# ================================================================
#  工具函数
# ================================================================

def _get_client_ip():
    """获取客户端IP"""
    return request.headers.get('X-Forwarded-For', request.remote_addr)


def _can_use_delivery():
    return session.get('username') in ('admin', '快递对账') or session.get('role') == 'admin'


def _paginate_query(conn, sql, count_sql, page=1, size=20, params=None):
    """通用分页查询"""
    offset = (page - 1) * size
    base_params = list(params or [])
    total = conn.execute(count_sql, base_params).fetchone()[0]
    rows = conn.execute(
        f"{sql} LIMIT ? OFFSET ?", base_params + [size, offset]
    ).fetchall()
    return rows_to_list(rows), total, page


def _status_from_statement(row):
    row = dict(row)
    status = row.get('overall_status') or ''
    if status:
        return status
    if row.get('status') in ('confirmed', 'archived'):
        return 'COMPLETED'
    return 'WAITING_INVOICE'


def _history_row(row):
    row = dict(row)
    total = row.get('total_invoice_amount') or row.get('current_payment') or 0
    key = row.get('reconciliation_key') or row.get('statement_key') or ''
    statement_no = row.get('statement_no') or f"ST{row.get('id')}"
    overall_status = _status_from_statement(row)
    invoice_status = row.get('invoice_status') or ('PASS' if overall_status == 'COMPLETED' else 'NOT_UPLOADED')
    return {
        "id": row.get('id'),
        "statement_no": statement_no,
        "reconciliation_key": key,
        "supplier": row.get('supplier_name') or '',
        "supplier_code": row.get('supplier_code') or '',
        "statement_total": f"{float(total or 0):.2f}",
        "erp_purchase_total": f"{float(row.get('erp_purchase_total') or total or 0):.2f}",
        "invoice_status": invoice_status,
        "overall_status": overall_status,
        "invoice_date": row.get('invoice_date') or '',
        "payment_date": row.get('payment_date') or '',
        "usage_remark": row.get('usage_remark') or '',
        "payment_log": row.get('payment_log') or '',
        "reconciliation_log": row.get('reconciliation_log') or '',
        "invoice_log": row.get('invoice_log') or '',
        "statement_date": row.get('statement_date') or '',
        "statement_period": row.get('statement_period') or '',
        "original_filename": row.get('original_filename') or row.get('source_file') or '',
        "created_at": str(row.get('created_at') or ''),
    }


def _amount_from_text(text):
    text = str(text or '')
    for pattern in (r'(?:金额|付款金额|发票金额)[：:\s]*[¥￥]?\s*([0-9][0-9,]*(?:\.\d+)?)', r'[¥￥]\s*([0-9][0-9,]*(?:\.\d+)?)'):
        m = re.search(pattern, text)
        if m:
            try:
                return float(m.group(1).replace(',', ''))
            except ValueError:
                return 0.0
    return 0.0


def _money_status(total, amount, kind):
    total = float(total or 0)
    amount = float(amount or 0)
    if amount <= 0.005:
        return 'UNPAID' if kind == 'payment' else 'UNINVOICED'
    if amount + 0.005 < total:
        return 'PARTIAL_PAID' if kind == 'payment' else 'UNDER_INVOICED'
    if amount - total > 0.005:
        return 'OVER_PAID' if kind == 'payment' else 'OVER_INVOICED'
    return 'PAID' if kind == 'payment' else 'INVOICED'


def _enrich_history_rows(rows):
    history_rows = [_history_row(row) for row in rows]
    ids = [row["id"] for row in history_rows]
    if not ids:
        return history_rows
    placeholders = ",".join(["?"] * len(ids))
    with get_db() as conn:
        records = conn.execute(f"""
            SELECT statement_id, title, text_content, record_date, file_name, created_at, amount
            FROM stm_statement_record
            WHERE statement_id IN ({placeholders})
            ORDER BY created_at DESC, id DESC
        """, ids).fetchall()
    grouped = {
        stmt_id: {
            "payment_log": "", "reconciliation_log": "", "invoice_log": "",
            "payment_amount": 0.0, "invoice_amount": 0.0,
            "payment_time": "", "invoice_time": "",
        }
        for stmt_id in ids
    }
    for record in records:
        stmt_id = record['statement_id']
        title = str(record.get('title') or '')
        text = " ".join(str(record.get(k) or '') for k in ('title', 'text_content', 'record_date', 'file_name')).strip()
        amount = float(record.get('amount') or 0) or _amount_from_text(record.get('text_content'))
        if '付款' in title or '支付' in title:
            grouped[stmt_id]["payment_amount"] += amount
            grouped[stmt_id]["payment_time"] = grouped[stmt_id]["payment_time"] or str(record.get('record_date') or '')
            grouped[stmt_id]["payment_log"] = grouped[stmt_id]["payment_log"] or text[:80]
        elif '开票' in title or '发票' in title:
            grouped[stmt_id]["invoice_amount"] += amount
            grouped[stmt_id]["invoice_time"] = grouped[stmt_id]["invoice_time"] or str(record.get('record_date') or '')
            grouped[stmt_id]["invoice_log"] = grouped[stmt_id]["invoice_log"] or text[:80]
        elif '对账' in title:
            grouped[stmt_id]["reconciliation_log"] = grouped[stmt_id]["reconciliation_log"] or text[:80]
    for row in history_rows:
        agg = grouped.get(row["id"], {})
        statement_total = float(row.get('statement_total') or 0)
        row.update(agg)
        row["payment_amount"] = f"{float(row.get('payment_amount') or 0):.2f}"
        row["invoice_amount"] = f"{float(row.get('invoice_amount') or 0):.2f}"
        row["payment_status"] = _money_status(statement_total, row["payment_amount"], 'payment')
        row["invoice_amount_status"] = _money_status(statement_total, row["invoice_amount"], 'invoice')
        row["statement_time"] = row.get("statement_date") or row.get("statement_period") or row.get("created_at", "")[:10]
        row["payment_date"] = row.get("payment_time") or row.get("payment_date") or ""
        row["invoice_date"] = row.get("invoice_time") or row.get("invoice_date") or ""
    return history_rows


def _history_summary(rows):
    summary = {
        "statement_count": len(rows),
        "statement_amount": 0.0,
        "payment_amount": 0.0,
        "invoice_amount": 0.0,
        "unpaid_count": 0,
        "partial_paid_count": 0,
        "paid_count": 0,
        "over_paid_count": 0,
        "uninvoiced_count": 0,
        "under_invoiced_count": 0,
        "invoiced_count": 0,
        "over_invoiced_count": 0,
    }
    for row in rows:
        summary["statement_amount"] += float(row.get("statement_total") or 0)
        summary["payment_amount"] += float(row.get("payment_amount") or 0)
        summary["invoice_amount"] += float(row.get("invoice_amount") or 0)
        payment_status = row.get("payment_status")
        invoice_status = row.get("invoice_amount_status")
        if payment_status == "PAID":
            summary["paid_count"] += 1
        elif payment_status == "PARTIAL_PAID":
            summary["partial_paid_count"] += 1
        elif payment_status == "OVER_PAID":
            summary["over_paid_count"] += 1
        else:
            summary["unpaid_count"] += 1
        if invoice_status == "INVOICED":
            summary["invoiced_count"] += 1
        elif invoice_status == "UNDER_INVOICED":
            summary["under_invoiced_count"] += 1
        elif invoice_status == "OVER_INVOICED":
            summary["over_invoiced_count"] += 1
        else:
            summary["uninvoiced_count"] += 1
    for key in ("statement_amount", "payment_amount", "invoice_amount"):
        summary[key] = f"{summary[key]:.2f}"
    return summary


def _merge_compare_line_checks(rows):
    """Merge same material/price lines for operator review; keep raw recognition rows."""
    groups = {}
    order = []
    sum_fields = (
        "statement_quantity", "statement_amount", "current_quantity",
        "current_amount", "historical_quantity", "historical_amount",
        "cumulative_quantity", "cumulative_amount",
    )
    for row in rows:
        key = (
            str(row.get("material_code") or "").strip().upper(),
            _norm_money(row.get("statement_unit_price")),
        )
        if not key[0]:
            key = ("__ITEM__", str(row.get("statement_item_id")))
        if key not in groups:
            merged = dict(row)
            merged["statement_item_ids"] = [row.get("statement_item_id")]
            merged["_orders"] = [row.get("purchase_order_id")] if row.get("purchase_order_id") else []
            merged["_dates"] = [row.get("delivery_date")] if row.get("delivery_date") else []
            merged["merged_count"] = 1
            groups[key] = merged
            order.append(key)
            continue
        merged = groups[key]
        merged["statement_item_ids"].append(row.get("statement_item_id"))
        merged["merged_count"] += 1
        for field in sum_fields:
            merged[field] = float(merged.get(field) or 0) + float(row.get(field) or 0)
        for source, target in (("purchase_order_id", "_orders"), ("delivery_date", "_dates")):
            value = row.get(source)
            if value and value not in merged[target]:
                merged[target].append(value)
        merged["manual_approved"] = bool(merged.get("manual_approved")) and bool(row.get("manual_approved"))
        for field in ("quantity_status", "amount_status"):
            if row.get(field) != "PASS":
                merged[field] = row.get(field)
        if row.get("allocation_status") == "OVER":
            merged["allocation_status"] = "OVER"
        if row.get("issue_text") and row.get("issue_text") not in str(merged.get("issue_text") or ""):
            merged["issue_text"] = "；".join(filter(None, [merged.get("issue_text"), row.get("issue_text")]))
    result = []
    for key in order:
        merged = groups[key]
        merged["purchase_order_id"] = "；".join(merged.pop("_orders"))
        merged["delivery_date"] = "；".join(merged.pop("_dates"))
        result.append(merged)
    return result


def _search_value(value, exact=False):
    value = str(value or '').strip()
    return value if exact else f"%{value}%"


def _search_operator(exact=False):
    return "=" if exact else "LIKE"


_supplier_code_cache = {}


def _find_supplier_candidates(supplier_name):
    """从已确认的数据库记录查询合作商；禁止用税号或HTTP结果冒充编码。"""
    supplier_name = str(supplier_name or '').strip()
    if not supplier_name:
        return []
    compact = re.sub(r'\s+', '', supplier_name)
    candidates = {}
    with get_db() as conn:
        rows = conn.execute("""
            SELECT DISTINCT supplier_name, supplier_code
            FROM stm_statement
            WHERE COALESCE(supplier_code, '') <> ''
              AND (supplier_name LIKE ? OR supplier_code LIKE ?)
            ORDER BY supplier_name
            LIMIT 20
        """, (f"%{supplier_name}%", f"%{supplier_name}%")).fetchall()
    for row in rows:
        name = str(row['supplier_name'] or '').strip()
        code = str(row['supplier_code'] or '').strip()
        if name and code:
            candidates[code] = {"name": name, "code": code, "source": "history"}

    if ERP_MYSQL_CONFIG.get("host") and ERP_MYSQL_CONFIG.get("user"):
        import pymysql
        erp_conn = pymysql.connect(
            **ERP_MYSQL_CONFIG,
            connect_timeout=10,
            read_timeout=15,
            cursorclass=pymysql.cursors.DictCursor,
        )
        try:
            with erp_conn.cursor() as cursor:
                cursor.execute("""
                    SELECT partners_code AS code, name
                    FROM partners
                    WHERE deleted_at IS NULL
                      AND status=1
                      AND type='GYS'
                      AND name LIKE %s
                    ORDER BY name
                    LIMIT 20
                """, (f"%{supplier_name}%",))
                for row in cursor.fetchall():
                    name = str(row.get("name") or "").strip()
                    code = str(row.get("code") or "").strip()
                    if name and code:
                        candidates[code] = {
                            "name": name,
                            "code": code,
                            "source": "erp_database",
                        }
        finally:
            erp_conn.close()
    return list(candidates.values())


def _resolve_supplier_code(supplier_name):
    supplier_name = str(supplier_name or "").strip()
    if not supplier_name:
        return ""
    if supplier_name in _supplier_code_cache:
        return _supplier_code_cache[supplier_name]

    # 只接受真正的合作商编码字段；sys_enterprise.tax_id 是税号，不能使用。
    code = _lookup_supplier_code_from_db(supplier_name)
    _supplier_code_cache[supplier_name] = code
    return code


def _lookup_supplier_code_from_db(supplier_name):
    """从ERP只读库的供应商主数据表查询权威合作商编码。"""
    if not ERP_MYSQL_CONFIG.get("host") or not ERP_MYSQL_CONFIG.get("user"):
        return ""
    try:
        import pymysql
        conn = pymysql.connect(
            **ERP_MYSQL_CONFIG,
            connect_timeout=10,
            read_timeout=15,
            cursorclass=pymysql.cursors.DictCursor,
        )
        try:
            with conn.cursor() as cursor:
                cursor.execute("""
                    SELECT partners_code
                    FROM partners
                    WHERE deleted_at IS NULL
                      AND status=1
                      AND type='GYS'
                      AND name=%s
                    LIMIT 1
                """, (supplier_name,))
                row = cursor.fetchone()
                return str((row or {}).get("partners_code") or "").strip()
        finally:
            conn.close()
    except Exception as exc:
        logger.warning("ERP数据库查询合作商编码失败: %s", exc)
    return ""


def _save_statement_upload(file_storage):
    original_filename = Path(file_storage.filename or "对账单").name
    ext = Path(original_filename).suffix.lower()
    image_exts = ('.jpg', '.jpeg', '.png', '.webp')
    if ext not in ('.pdf', '.xls', '.xlsx', *image_exts):
        raise ValueError("仅支持 PDF、图片或 Excel")
    filename = secure_filename(original_filename) or f"statement{ext}"
    ts = datetime.now().strftime('%Y%m%d%H%M%S')
    stored_name = f"stm_{ts}_{filename}"
    filepath = str(UPLOAD_DIR / stored_name)
    file_storage.save(filepath)
    return {
        "ext": ext, "filename": original_filename, "stored_name": stored_name,
        "filepath": filepath,
    }


def _detect_supplier_from_saved_file(saved):
    """First stage: supplier only. Vision reads one page with a tiny token budget."""
    ext = saved["ext"]
    filepath = saved["filepath"]
    if ext in ('.xlsx', '.xls'):
        data = (
            parse_statement_xlsx(filepath)
            if ext == '.xlsx' else parse_statement_xls(filepath)
        )
        supplier = str(data.get("supplier_name") or "").strip()
        if not supplier and ext == '.xlsx':
            supplier = recognize_excel_supplier_locally(filepath)
        return supplier
    if ext == '.pdf':
        try:
            import pdfplumber
            with pdfplumber.open(filepath) as pdf:
                text = (pdf.pages[0].extract_text() or "") if pdf.pages else ""
            match = re.search(
                r'(?:供应商|供货商|供货单位|协力厂商)[：:\\s]*'
                r'([^\\n]+?(?:有限责任公司|有限公司|公司))',
                text,
            )
            if match and '骊威' not in match.group(1):
                return match.group(1).strip()
        except Exception:
            pass
    local_detected = recognize_supplier_locally(filepath)
    if local_detected:
        return local_detected
    try:
        detected = recognize_supplier_with_qwen(filepath)
        if detected:
            return detected
    except Exception as exc:
        logger.warning("千问合作商识别失败，切换本地OCR: %s", exc)
    return recognize_supplier_locally(filepath)


def _parse_saved_statement(saved, progress_callback=None):
    ext = saved["ext"]
    filename = saved["filename"]
    stored_name = saved["stored_name"]
    filepath = saved["filepath"]
    image_exts = ('.jpg', '.jpeg', '.png', '.webp')
    if ext == '.xlsx':
        data = parse_statement_xlsx(filepath)
    elif ext == '.xls':
        data = parse_statement_xls(filepath)
    elif ext in image_exts:
        data = parse_statement_image(filepath)
    else:
        data = parse_statement_pdf(filepath, progress_callback=progress_callback)
    if not data.get('items'):
        detail = '; '.join(data.get('errors') or []) or '未识别到对账单明细'
        raise ValueError(f"对账单识别失败：{detail}")
    supplier_name = (data.get('supplier_name') or '').strip()
    supplier_code = (data.get('supplier_code') or '').strip() or _resolve_supplier_code(supplier_name)
    if supplier_code:
        data['supplier_code'] = supplier_code
    statement_month = (data.get('statement_month') or '').strip()
    statement_key = data.get('statement_key') or (f"{supplier_code}_{statement_month}" if supplier_code and statement_month else '')
    statement_no = data.get('statement_no') or statement_key or Path(filename).stem
    total_amount = data.get('total_invoice_amount') or data.get('current_payment') or 0
    return data, {
        "stored_name": stored_name,
        "filepath": filepath,
        "supplier_code": supplier_code,
        "statement_month": statement_month,
        "statement_key": statement_key,
        "statement_no": statement_no,
        "statement_fingerprint": _statement_fingerprint(supplier_code, data.get('supplier_name') or '', data.get('items', [])),
        "total_amount": total_amount,
        "original_filename": filename,
    }


def _parse_saved_statement_cached(saved, progress_callback=None):
    """Reuse a confirmed file's full OCR result without spending Qwen tokens again."""
    digest = hashlib.sha256()
    with open(saved["filepath"], "rb") as source:
        for chunk in iter(lambda: source.read(1024 * 1024), b""):
            digest.update(chunk)
    model_key = re.sub(r'[^A-Za-z0-9_.-]+', '_', QWEN_OCR_MODEL or 'local')
    cache_path = STATEMENT_PARSE_CACHE_DIR / (
        f"v5_{model_key}_{digest.hexdigest()}.json"
    )
    if cache_path.exists():
        try:
            cached = json.loads(cache_path.read_text(encoding="utf-8"))
            data = cached["data"]
            meta = cached["meta"]
            meta.update({
                "stored_name": saved["stored_name"],
                "filepath": saved["filepath"],
                "original_filename": saved["filename"],
            })
            logger.info("复用对账单完整识别缓存 file=%s", saved["filename"])
            return data, meta
        except (OSError, ValueError, KeyError, TypeError):
            pass
    data, meta = _parse_saved_statement(saved, progress_callback=progress_callback)
    try:
        temp_path = cache_path.with_suffix(".tmp")
        temp_path.write_text(
            json.dumps({"data": data, "meta": meta}, ensure_ascii=False, default=str),
            encoding="utf-8",
        )
        temp_path.replace(cache_path)
    except OSError as exc:
        logger.warning("保存对账单识别缓存失败 file=%s: %s", saved["filename"], exc)
    return data, meta

def _norm_key_part(value):
    return str(value or '').strip().upper()


def _norm_money(value):
    try:
        return f"{float(value or 0):.6f}"
    except (TypeError, ValueError):
        return "0.000000"


def _statement_line_key(supplier_code, supplier_name, item):
    raw = "|".join([
        _norm_key_part(supplier_code or supplier_name),
        _norm_key_part(item.get('customer_order_no')),
        _norm_key_part(item.get('customer_material_code')),
        _norm_money(item.get('unit_price_incl_tax')),
    ])
    return hashlib.sha1(raw.encode('utf-8')).hexdigest()


def _statement_fingerprint(supplier_code, supplier_name, items):
    parts = []
    for item in items or []:
        parts.append("|".join([
            _statement_line_key(supplier_code, supplier_name, item),
            _norm_money(item.get('quantity')),
            _norm_money(item.get('amount_incl_tax')),
        ]))
    raw = "\n".join(sorted(parts))
    return hashlib.sha1(raw.encode('utf-8')).hexdigest() if raw else ''


def _recognition_row_issues(item):
    issues = []
    qty = float(item.get('quantity') or 0)
    price = float(item.get('unit_price_incl_tax') or 0)
    amount = float(item.get('amount_incl_tax') or 0)
    if qty <= 0:
        issues.append("数量缺失")
    if price <= 0:
        issues.append("含税单价缺失")
    if amount <= 0:
        issues.append("含税金额缺失")
    if qty > 0 and price > 0 and amount > 0:
        tolerance = max(0.05, abs(amount) * 0.005)
        if abs(qty * price - amount) > tolerance:
            issues.append("数量×单价与金额不一致")
    order_no = str(item.get('customer_order_no') or '')
    material_code = str(item.get('customer_material_code') or '')
    if re.fullmatch(r'LW[A-Z0-9]{6,}', order_no, re.I):
        issues.append("采购单号疑似物料编码")
    if re.match(r'^(?:AHLW[-_/]|\d{8}[-_.])', material_code, re.I):
        issues.append("物料编码疑似采购单号")
    return issues


def _open_erp_read_connection():
    import pymysql
    return pymysql.connect(
        **ERP_MYSQL_CONFIG,
        connect_timeout=10,
        read_timeout=15,
        cursorclass=pymysql.cursors.DictCursor,
    )


def _erp_statement_line(
    supplier_code, material_code, purchase_order_id='', delivery_date='',
    erp_conn=None,
):
    """Read the closest authoritative ERP purchase detail for one statement line."""
    if not supplier_code or not material_code or not ERP_MYSQL_CONFIG.get("host"):
        return {}
    target_date = ''
    order_date = re.search(r'(\d{8})', str(purchase_order_id or ''))
    if order_date:
        try:
            target_date = datetime.strptime(order_date.group(1), '%Y%m%d').strftime('%Y-%m-%d')
        except ValueError:
            pass
    target_date = target_date or str(delivery_date or '')[:10]
    owns_connection = erp_conn is None
    try:
        if owns_connection:
            erp_conn = _open_erp_read_connection()
        try:
            with erp_conn.cursor() as cursor:
                cursor.execute("""
                    SELECT
                        d.id AS detail_id,
                        d.created_at AS order_date,
                        d.purchase_quantity,
                        d.received_quantity,
                        d.arrived_quantity,
                        d.unit_price,
                        d.amount,
                        o.name AS erp_order_name,
                        COALESCE((
                            SELECT SUM(a.arrival_quantity)
                            FROM purchase_arrival_records a
                            WHERE a.purchase_order_detail_id=d.id
                              AND a.deleted_at IS NULL
                              AND a.arrival_type='NORMAL'
                        ), 0) AS arrival_record_quantity,
                        (
                            SELECT MAX(a.created_at)
                            FROM purchase_arrival_records a
                            WHERE a.purchase_order_detail_id=d.id
                              AND a.deleted_at IS NULL
                              AND a.arrival_type='NORMAL'
                        ) AS arrival_date
                    FROM purchase_orders_details d
                    LEFT JOIN purchase_orders o
                      ON o.id=CAST(d.purchase_order_id AS UNSIGNED)
                     AND o.deleted_at IS NULL
                    WHERE d.deleted_at IS NULL
                      AND d.partners_id=%s
                      AND d.nameid=%s
                    ORDER BY
                      CASE WHEN o.name=%s THEN 0 ELSE 1 END,
                      CASE WHEN %s<>'' THEN ABS(TIMESTAMPDIFF(
                          DAY, DATE(d.created_at), DATE(%s)
                      )) ELSE 999999 END,
                      d.id DESC
                    LIMIT 1
                """, (
                    supplier_code, material_code, purchase_order_id,
                    target_date, target_date,
                ))
                row = cursor.fetchone()
                if not row:
                    return {}
                arrived = float(
                    row.get('arrival_record_quantity')
                    or row.get('arrived_quantity')
                    or row.get('received_quantity')
                    or 0
                )
                price = float(row.get('unit_price') or 0)
                return {
                    "erp_order_name": str(row.get('erp_order_name') or ''),
                    "erp_order_date": str(row.get('order_date') or '')[:10],
                    "erp_arrival_date": str(row.get('arrival_date') or '')[:10],
                    "erp_purchase_quantity": float(row.get('purchase_quantity') or 0),
                    "erp_arrival_quantity": arrived,
                    "erp_unit_price": price,
                    "erp_amount": round(arrived * price, 2),
                }
        finally:
            if owns_connection and erp_conn:
                erp_conn.close()
    except Exception as exc:
        logger.warning(
            "ERP采购明细查询失败 supplier=%s material=%s: %s",
            supplier_code, material_code, exc,
        )
        return {}


def _sync_statement_allocations(conn, stmt_id):
    stmt = conn.execute("SELECT * FROM stm_statement WHERE id=?", (stmt_id,)).fetchone()
    if not stmt:
        return
    stmt = dict(stmt)
    conn.execute("DELETE FROM stm_statement_allocation WHERE statement_id=?", (stmt_id,))
    # 同步清理该对账单的差异工单（保留跨月带入的工单）
    conn.execute(
        "DELETE FROM stm_diff_ticket WHERE statement_id=? AND is_carried_forward=0",
        (stmt_id,),
    )
    items = conn.execute(
        "SELECT * FROM stm_statement_item WHERE statement_id=? ORDER BY seq", (stmt_id,)
    ).fetchall()
    supplier_code = stmt.get('supplier_code') or ''
    supplier_name = stmt.get('supplier_name') or ''
    erp_total = 0.0
    shared_erp_conn = None
    if supplier_code and ERP_MYSQL_CONFIG.get("host"):
        try:
            shared_erp_conn = _open_erp_read_connection()
        except Exception as exc:
            logger.warning("ERP批量读取连接失败 supplier=%s: %s", supplier_code, exc)
    for row in items:
        item = dict(row)
        erp = _erp_statement_line(
            supplier_code,
            item.get('customer_material_code') or '',
            item.get('customer_order_no') or '',
            item.get('delivery_date') or '',
            erp_conn=shared_erp_conn,
        )
        line_key = _statement_line_key(supplier_code, supplier_name, item)
        hist = conn.execute("""
            SELECT
                COALESCE(SUM(a.current_quantity), 0) AS qty,
                COALESCE(SUM(a.current_amount), 0) AS amount
            FROM stm_statement_allocation
            a JOIN stm_statement s ON s.id = a.statement_id
            WHERE a.line_key=?
              AND a.statement_id<>?
              AND (s.created_at < ? OR (s.created_at = ? AND s.id < ?))
        """, (line_key, stmt_id, stmt.get('created_at') or '', stmt.get('created_at') or '', stmt_id)).fetchone()
        current_qty = float(item.get('quantity') or 0)
        current_amount = float(item.get('amount_incl_tax') or 0)
        hist = dict(hist)
        historical_qty = float(hist.get('qty') or 0)
        historical_amount = float(hist.get('amount') or 0)
        cumulative_qty = historical_qty + current_qty
        cumulative_amount = historical_amount + current_amount
        erp_quantity = float(erp.get("erp_arrival_quantity") or 0)
        erp_amount = float(erp.get("erp_amount") or 0)
        erp_total += erp_amount
        remaining_quantity = erp_quantity - cumulative_qty
        remaining_amount = erp_amount - cumulative_amount
        if not erp:
            allocation_status = 'INFO'
            issue_text = "ERP未找到对应采购到库明细"
        elif abs(
            float(item.get('unit_price_incl_tax') or 0)
            - float(erp.get('erp_unit_price') or 0)
        ) > 0.000001:
            allocation_status = 'OVER'
            issue_text = (
                f"对账单价 {float(item.get('unit_price_incl_tax') or 0):g} "
                f"与ERP单价 {float(erp.get('erp_unit_price') or 0):g} 不一致"
            )
        elif cumulative_qty > erp_quantity + 0.000001:
            allocation_status = 'OVER'
            issue_text = f"累计对账数量超过ERP到库数量 {cumulative_qty - erp_quantity:g}"
        elif cumulative_amount > erp_amount + 0.05:
            allocation_status = 'OVER'
            issue_text = f"累计对账金额超过ERP可对金额 {cumulative_amount - erp_amount:.2f}"
        else:
            allocation_status = 'PASS'
            issue_text = "ERP到库数量、单价和金额校验通过"
        conn.execute("""
            INSERT INTO stm_statement_allocation (
                statement_id, statement_item_id, supplier_code, supplier_name,
                line_key, purchase_order_id, material_code, delivery_date, unit_price,
                current_quantity, current_amount,
                historical_quantity, historical_amount,
                cumulative_quantity, cumulative_amount,
                erp_quantity, erp_amount, remaining_quantity, remaining_amount,
                allocation_status, issue_text
            ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        """, (
            stmt_id,
            item.get('id'),
            supplier_code,
            supplier_name,
            line_key,
            item.get('customer_order_no') or '',
            item.get('customer_material_code') or '',
            item.get('delivery_date') or '',
            item.get('unit_price_incl_tax') or 0,
            current_qty,
            current_amount,
            historical_qty,
            historical_amount,
            cumulative_qty,
            cumulative_amount,
            erp_quantity,
            erp_amount,
            remaining_quantity,
            remaining_amount,
            allocation_status,
            issue_text,
        ))
    conn.execute(
        "UPDATE stm_statement SET erp_purchase_total=? WHERE id=?",
        (round(erp_total, 2), stmt_id),
    )
    if shared_erp_conn:
        shared_erp_conn.close()


def _rebuild_all_statement_allocations(conn):
    rows = conn.execute("SELECT id FROM stm_statement ORDER BY created_at ASC, id ASC").fetchall()
    for row in rows:
        _sync_statement_allocations(conn, row["id"])


def _insert_statement(conn, data, meta):
    if not str(meta.get("supplier_code") or "").strip():
        raise ValueError("合作商编码为空，请先选择有效的ERP合作商")
    if not str(meta.get("statement_month") or "").strip():
        raise ValueError("账期为空，请确认文件中包含明确的对账月份")
    if not data.get("items"):
        raise ValueError("未识别到有效对账明细，不允许保存空对账单")
    identity_row = conn.execute("""
        SELECT * FROM stm_statement
        WHERE supplier_code=? AND statement_period=?
        ORDER BY id DESC LIMIT 1
    """, (meta["supplier_code"], meta["statement_month"])).fetchone()
    if identity_row:
        if (identity_row.get("statement_fingerprint") or "") == meta["statement_fingerprint"]:
            return identity_row["id"], True
        raise StatementIdentityConflict(identity_row["id"])

    existing = None
    if meta["statement_month"] and meta["statement_fingerprint"] and (meta["supplier_code"] or data.get('supplier_name')):
        existing = conn.execute("""
            SELECT * FROM stm_statement
            WHERE COALESCE(NULLIF(supplier_code, ''), supplier_name, '')=?
              AND statement_period=?
              AND statement_fingerprint=?
            LIMIT 1
        """, (meta["supplier_code"] or data.get('supplier_name') or '', meta["statement_month"], meta["statement_fingerprint"])).fetchone()
    if existing:
        return existing["id"], True

    has_issue = bool(data.get('errors')) or not meta["supplier_code"] or not data.get('items')
    overall_status = 'ERP_FAILED' if has_issue else 'ERP_PENDING'
    cursor = conn.execute("""
        INSERT INTO stm_statement (
            statement_period, statement_date,
            customer_name, customer_tax_id,
            supplier_code, statement_key,
            statement_no, reconciliation_key,
            supplier_name, supplier_tax_id,
            settlement_days,
            opening_balance, current_payment, closing_balance,
            delivered_unpaid, total_invoice_amount, total_quantity,
            balance_status, source_file, pdf_path,
            invoice_status, overall_status, erp_purchase_total, original_filename,
            statement_fingerprint
        ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
    """, (
        meta["statement_month"],
        data.get('statement_date', ''),
        (data.get('customer_name') or '骊威')[:512],
        data.get('customer_tax_id', ''),
        meta["supplier_code"],
        meta["statement_key"],
        meta["statement_no"],
        meta["statement_key"],
        (data.get('supplier_name') or '')[:512],
        data.get('supplier_tax_id', ''),
        data.get('settlement_days', 30),
        data.get('opening_balance', 0),
        data.get('current_payment') or 0,
        data.get('closing_balance', 0),
        data.get('delivered_unpaid', 0),
        meta["total_amount"],
        data.get('total_quantity', 0),
        'balanced' if data.get('balance_check', True) else 'unbalanced',
        meta["stored_name"],
        meta["filepath"],
        'NOT_UPLOADED',
        overall_status,
        0,
        meta["original_filename"],
        meta["statement_fingerprint"],
    ))
    stmt_id = cursor.lastrowid
    _replace_statement_items(conn, stmt_id, data)
    conn.execute(
        "UPDATE stm_statement SET recognition_metadata=? WHERE id=?",
        (json.dumps({
            "column_mapping": data.get("column_mapping") or {},
            "recognition_errors": data.get("errors") or [],
        }, ensure_ascii=False), stmt_id),
    )
    audit_log(conn, None, 'system', 'CREATE', 'statement', stmt_id,
              new_values={"statement_no": meta["statement_no"]}, ip=_get_client_ip())
    return stmt_id, False


def _replace_statement_items(conn, stmt_id, data):
    conn.execute("DELETE FROM stm_statement_item WHERE statement_id=?", (stmt_id,))
    for idx, item in enumerate(data.get('items', []), start=1):
        conn.execute("""
            INSERT INTO stm_statement_item (
                statement_id, seq, customer_order_no,
                customer_material_code, delivery_no, delivery_date,
                product_name, specification, quantity, unit,
                unit_price_incl_tax, amount_incl_tax
            ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?)
        """, (
            stmt_id, idx,
            item.get('customer_order_no', ''),
            item.get('customer_material_code', ''),
            item.get('delivery_no', ''),
            item.get('delivery_date', ''),
            item.get('product_name', ''),
            item.get('specification', ''),
            item.get('quantity', 0),
            item.get('unit', 'PCS'),
            item.get('unit_price_incl_tax', 0),
            item.get('amount_incl_tax', 0),
        ))
    _sync_statement_allocations(conn, stmt_id)


def _recompare_statement(conn, stmt_id, data, meta, usage_remark=''):
    row = conn.execute("SELECT * FROM stm_statement WHERE id=?", (stmt_id,)).fetchone()
    if not row:
        raise ValueError("原对账单不存在")
    identity_row = conn.execute("""
        SELECT id FROM stm_statement
        WHERE supplier_code=? AND statement_period=? AND id<>?
        LIMIT 1
    """, (meta["supplier_code"], meta["statement_month"], stmt_id)).fetchone()
    if identity_row:
        raise StatementIdentityConflict(identity_row["id"])
    has_issue = bool(data.get('errors')) or not data.get('items')
    overall_status = 'ERP_FAILED' if has_issue else 'ERP_PENDING'
    conn.execute("""
        UPDATE stm_statement
        SET statement_period=?, statement_date=?,
            customer_name=?, customer_tax_id=?,
            supplier_code=?, statement_key=?,
            statement_no=?, reconciliation_key=?,
            supplier_name=?, supplier_tax_id=?,
            settlement_days=?,
            opening_balance=?, current_payment=?, closing_balance=?,
            delivered_unpaid=?, total_invoice_amount=?, total_quantity=?,
            balance_status=?, source_file=?, pdf_path=?,
            invoice_status=?, overall_status=?, erp_purchase_total=?, original_filename=?,
            usage_remark=?, statement_fingerprint=?
        WHERE id=?
    """, (
        meta["statement_month"],
        data.get('statement_date', ''),
        (data.get('customer_name') or '骊威')[:512],
        data.get('customer_tax_id', ''),
        meta["supplier_code"],
        meta["statement_key"],
        meta["statement_no"],
        meta["statement_key"],
        (data.get('supplier_name') or '')[:512],
        data.get('supplier_tax_id', ''),
        data.get('settlement_days', 30),
        data.get('opening_balance', 0),
        data.get('current_payment') or 0,
        data.get('closing_balance', 0),
        data.get('delivered_unpaid', 0),
        meta["total_amount"],
        data.get('total_quantity', 0),
        'balanced' if data.get('balance_check', True) else 'unbalanced',
        meta["stored_name"],
        meta["filepath"],
        'NOT_UPLOADED',
        overall_status,
        0,
        meta["original_filename"],
        usage_remark or data.get('usage_remark', '') or row.get('usage_remark', ''),
        meta["statement_fingerprint"],
        stmt_id,
    ))
    _replace_statement_items(conn, stmt_id, data)
    conn.execute(
        "UPDATE stm_statement SET recognition_metadata=? WHERE id=?",
        (json.dumps({
            "column_mapping": data.get("column_mapping") or {},
            "recognition_errors": data.get("errors") or [],
        }, ensure_ascii=False), stmt_id),
    )
    audit_log(conn, None, 'system', 'UPDATE', 'statement', stmt_id,
              new_values={"statement_no": meta["statement_no"], "action": "recompare"}, ip=_get_client_ip())
    return stmt_id


def _quote_mysql_name(name):
    return '`' + str(name).replace('`', '``') + '`'


def _export_table_csv(table_name, output_path):
    with get_db() as conn:
        columns = [
            row['COLUMN_NAME']
            for row in conn.execute("""
                SELECT COLUMN_NAME
                FROM information_schema.COLUMNS
                WHERE TABLE_SCHEMA = DATABASE() AND TABLE_NAME = ?
                ORDER BY ORDINAL_POSITION
            """, (table_name,)).fetchall()
        ]
        if not columns:
            raise ValueError(f"数据库表不存在：{table_name}")
        rows = conn.execute(f"SELECT * FROM {_quote_mysql_name(table_name)}").fetchall()
    with open(output_path, 'w', newline='', encoding='utf-8-sig') as file:
        writer = csv.DictWriter(file, fieldnames=columns)
        writer.writeheader()
        for row in rows:
            data = dict(row)
            writer.writerow({col: data.get(col, '') for col in columns})


def _save_delivery_upload(work_dir, field_name, default_name):
    uploaded = request.files.get(field_name)
    target = work_dir / default_name
    if uploaded and uploaded.filename:
        uploaded.save(target)
        return target
    default_path = DELIVERY_COMPARE_DIR / default_name
    if not default_path.exists():
        raise ValueError(f"缺少默认文件：{default_name}")
    shutil.copy(default_path, target)
    return target


def _month_range(month):
    if not month:
        return '', ''
    start = datetime.strptime(month + '-01', '%Y-%m-%d')
    if start.month == 12:
        end = start.replace(year=start.year + 1, month=1)
    else:
        end = start.replace(month=start.month + 1)
    return start, end


def _previous_month():
    today = datetime.now()
    first_day = today.replace(day=1)
    if first_day.month == 1:
        previous = first_day.replace(year=first_day.year - 1, month=12)
    else:
        previous = first_day.replace(month=first_day.month - 1)
    return previous.strftime('%Y-%m')


def _format_delivery_cell(value):
    if value is None:
        return ''
    if isinstance(value, datetime):
        return value.strftime('%Y-%m-%d')
    return str(value).strip()


def _normalize_delivery_month(value):
    text = _format_delivery_cell(value)
    if not text:
        return ''
    for fmt in ('%Y-%m', '%Y-%m-%d', '%Y/%m', '%Y/%m/%d'):
        try:
            return datetime.strptime(text[:10], fmt).strftime('%Y-%m')
        except ValueError:
            pass
    return text[:7] if len(text) >= 7 else text


def _month_from_delivery_date(value):
    if isinstance(value, datetime):
        return value.strftime('%Y-%m')
    text = _format_delivery_cell(value)
    if not text:
        return ''
    for fmt in ('%Y-%m-%d', '%Y/%m/%d', '%Y-%m', '%Y/%m', '%d/%m/%Y'):
        try:
            return datetime.strptime(text[:10], fmt).strftime('%Y-%m')
        except ValueError:
            pass
    return text[:7] if len(text) >= 7 and text[4:5] in ('-', '/') else ''


def _parse_delivery_datetime(value):
    if isinstance(value, datetime):
        return value
    value = (value or '').strip()
    for fmt in ("%d/%m/%Y %H:%M:%S.%f", "%d/%m/%Y %H:%M:%S", "%d/%m/%Y %H:%M", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
        try:
            return datetime.strptime(value, fmt)
        except ValueError:
            pass
    return None


def _row_in_month(row, date_column, start, end):
    if not start or not end:
        return True
    date_columns = [date_column]
    for fallback in ('transaction_time', 'created_at'):
        if fallback not in date_columns:
            date_columns.append(fallback)
    for column in date_columns:
        dt = _parse_delivery_datetime(row.get(column, ''))
        if dt:
            return start <= dt < end
    return False


def _load_table_rows(table_name):
    with get_db() as conn:
        return [dict(row) for row in conn.execute(f"SELECT * FROM {_quote_mysql_name(table_name)}").fetchall()]


def _file_sha256(path):
    digest = hashlib.sha256()
    with open(path, 'rb') as file:
        for chunk in iter(lambda: file.read(1024 * 1024), b''):
            digest.update(chunk)
    return digest.hexdigest()


def _display_filename(filename):
    return Path(filename or '').name or '快递对账单.xlsx'


def _storage_filename(filename, fallback_prefix='delivery_statement'):
    display = _display_filename(filename)
    suffix = Path(display).suffix or '.xlsx'
    safe = secure_filename(display)
    if not safe or not Path(safe).suffix:
        safe = f"{fallback_prefix}_{datetime.now().strftime('%Y%m%d%H%M%S')}{suffix}"
    return safe


def _delivery_history_count(unique_key='', month='', filename=''):
    if unique_key:
        with get_db() as conn:
            return conn.execute("""
                SELECT COUNT(*) AS cnt
                FROM delivery_reconciliation_run
                WHERE unique_key=?
            """, (unique_key,)).fetchone()[0]
    if not month or not filename:
        return 0
    with get_db() as conn:
        return conn.execute("""
            SELECT COUNT(*) AS cnt
            FROM delivery_reconciliation_run
            WHERE statement_month=? AND original_filename=?
        """, (month, filename)).fetchone()[0]


def _record_delivery_run(meta, filename, file_hash, counts, result_path):
    with get_db() as conn:
        cursor = conn.execute("""
            INSERT INTO delivery_reconciliation_run (
                unique_key, courier_company, fill_date,
                statement_month, original_filename, file_hash,
                statement_count, matched_count, only_statement_count, only_system_count,
                result_path, created_by
            ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?)
        """, (
            meta.get('唯一标识', ''),
            meta.get('快递公司', ''),
            meta.get('填写日期', ''),
            meta.get('对账月份', ''),
            filename,
            file_hash,
            counts.get('statement_count', 0),
            counts.get('matched_count', 0),
            counts.get('only_statement_count', 0),
            counts.get('only_system_count', 0),
            str(result_path),
            session.get('username', ''),
        ))
        return cursor.lastrowid


def _excel_sheet_name(name):
    safe = re.sub(r"[\[\]\:\*\?\/\\]", "_", str(name or "Sheet")).strip()
    return (safe or "Sheet")[:31]


def _read_delivery_sheet(ws):
    meta = {'唯一标识': '', '快递公司': '', '对账月份': '', '填写日期': '', '批次名称': '', '备注': ''}
    header_row = 1
    first_row = [str(cell.value or '').strip().replace('* ', '').replace('*', '') for cell in ws[1]]
    if {'快递公司', '对账月份', '填写日期'}.issubset(set(first_row)):
        meta_values = [cell.value for cell in ws[2]]
        meta = {
            key: _format_delivery_cell(meta_values[first_row.index(key)]) if key in first_row and first_row.index(key) < len(meta_values) else ''
            for key in meta
        }
        for idx, row in enumerate(ws.iter_rows(min_row=3, values_only=True), start=3):
            labels = [str(value or '').strip().replace('* ', '').replace('*', '') for value in row]
            if '快递单号' in labels:
                header_row = idx
                break
        else:
            raise ValueError("模板缺少明细表头：快递单号")
    headers = [str(cell.value or '').strip().replace('* ', '').replace('*', '') for cell in ws[header_row]]
    if '快递单号' not in set(headers):
        raise ValueError("模板表头必须包含：快递单号")
    if not meta.get('快递公司') and ws.title != '快递对账':
        meta['快递公司'] = ws.title
    records = []
    for row_number, values in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
        data = {headers[idx]: values[idx] if idx < len(values) else '' for idx in range(len(headers))}
        tracking = delivery_compare.normalize_tracking(data.get('快递单号', ''))
        if not tracking:
            continue
        remark = str(data.get('备注') or '')
        if '示例行' in remark:
            continue
        records.append({
            '_row': row_number,
            '_tracking': tracking,
            '快递公司': str(data.get('快递公司') or meta.get('快递公司') or '').strip(),
            '快递单号': tracking,
            '账单日期': data.get('账单日期') or '',
            '运费': data.get('运费') or '',
            '重量': data.get('重量') or '',
            '备注': remark,
        })
    first_bill_month = ''
    for record in records:
        first_bill_month = _month_from_delivery_date(record.get('账单日期'))
        if first_bill_month:
            break
    meta['对账月份'] = _normalize_delivery_month(meta.get('对账月份')) or first_bill_month or _previous_month()
    meta['填写日期'] = _format_delivery_cell(meta.get('填写日期')) or datetime.now().strftime('%Y-%m-%d')
    meta['批次名称'] = _format_delivery_cell(meta.get('批次名称')) or '月度账单'
    if not meta.get('快递公司') and records:
        meta['快递公司'] = records[0].get('快递公司', '')
    if not meta.get('唯一标识'):
        parts = [meta.get('快递公司') or '快递', meta.get('对账月份') or _previous_month(), meta.get('批次名称') or '月度账单']
        meta['唯一标识'] = '-'.join(parts)
    return meta, records


def _read_delivery_template(path):
    from openpyxl import load_workbook
    wb = load_workbook(path, data_only=True)
    batches = []
    errors = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        try:
            meta, records = _read_delivery_sheet(ws)
        except ValueError as exc:
            errors.append(f"{sheet_name}: {exc}")
            continue
        if records:
            batches.append((meta, records))
    if not batches and errors:
        raise ValueError("；".join(errors))
    return batches


def _delivery_result_rows(statement_records, delivery_rows, return_rows, month, date_column):
    start, end = _month_range(month)
    delivery_by_no = {}
    for row in delivery_rows:
        tracking = delivery_compare.normalize_tracking(row.get('logistics_number', ''))
        if tracking:
            delivery_by_no.setdefault(tracking, []).append(row)
    return_by_no = {}
    for row in return_rows:
        for col in ('logistics_number', 'return_tracking_number'):
            tracking = delivery_compare.normalize_tracking(row.get(col, ''))
            if tracking:
                return_by_no.setdefault(tracking, []).append(row)

    matched, only_statement = [], []
    statement_numbers = set()
    for record in statement_records:
        tracking = record['_tracking']
        statement_numbers.add(tracking)
        system_matches = [r for r in delivery_by_no.get(tracking, []) if _row_in_month(r, date_column, start, end)]
        return_matches = return_by_no.get(tracking, [])
        row = {
            '物流单号': tracking,
            '对账月份': month,
            '快递Excel行号': record.get('_row', ''),
            '快递公司': record.get('快递公司', ''),
            '账单日期': record.get('账单日期', ''),
            '运费': record.get('运费', ''),
            '重量': record.get('重量', ''),
            '备注': record.get('备注', ''),
            '系统匹配数': len(system_matches),
            '退换货匹配数': len(return_matches),
            '匹配状态': '匹配成功' if system_matches or return_matches else '仅快递账单',
            '系统来源': 'delivery' if system_matches else ('returnandexchangestop' if return_matches else ''),
        }
        if system_matches:
            first = system_matches[0]
            for col in ('id', 'created_at', 'platform', 'shop', 'order_number', 'logistics_company', 'receiver_name', 'receiver_mobile', 'order_status', 'amount'):
                row['系统_' + col] = first.get(col, '')
        if return_matches:
            first_return = return_matches[0]
            for col in ('id', 'created_at', 'platform', 'shop', 'order_number', 'logistics_company', 'return_tracking_number', 'return_logistics_company', 'reason'):
                row['退换货_' + col] = first_return.get(col, '')
        (matched if system_matches or return_matches else only_statement).append(row)

    only_system = []
    for row in delivery_rows:
        if not _row_in_month(row, date_column, start, end):
            continue
        tracking = delivery_compare.normalize_tracking(row.get('logistics_number', ''))
        if not tracking or tracking in statement_numbers:
            continue
        only_system.append({
            '物流单号': tracking,
            '对账月份': month,
            '匹配状态': '仅系统',
            '系统_id': row.get('id', ''),
            '系统_created_at': row.get('created_at', ''),
            '系统_platform': row.get('platform', ''),
            '系统_shop': row.get('shop', ''),
            '系统_order_number': row.get('order_number', ''),
            '系统_logistics_company': row.get('logistics_company', ''),
            '系统_amount': row.get('amount', ''),
        })
    summary = [
        ['指标', '数值'],
        ['对账月份', month],
        ['快递账单唯一单号数', len(statement_numbers)],
        ['匹配成功记录数', len(matched)],
        ['仅快递账单记录数', len(only_statement)],
        ['仅系统记录数', len(only_system)],
    ]
    return summary, matched, only_statement, only_system


# ================================================================
#  用户 API
# ================================================================

@app.route("/api/users", methods=["GET"])
def get_users():
    """获取所有活跃用户"""
    with get_db() as conn:
        rows = conn.execute(
            "SELECT id, username, real_name, role, is_active, created_at FROM sys_user WHERE is_active=1"
        ).fetchall()
    return jsonify({"code": 0, "data": rows_to_list(rows)})


@app.route("/api/users/<int:user_id>", methods=["GET"])
def get_user(user_id):
    """获取单个用户详情"""
    with get_db() as conn:
        row = conn.execute(
            "SELECT id, username, real_name, role, is_active, created_at FROM sys_user WHERE id=?",
            (user_id,)
        ).fetchone()
    if row:
        return jsonify({"code": 0, "data": dict_from_row(row)})
    return jsonify({"code": 404, "message": "用户不存在"}), 404


@app.route("/api/login", methods=["POST"])
def login():
    """用户登录 — bcrypt密码验证"""
    data = request.get_json()
    if not data:
        return jsonify({"code": 400, "message": "请求体为空"}), 400

    username = data.get("username", "").strip()
    password = data.get("password", "")

    if not username or not password:
        return jsonify({"code": 401, "message": "用户名或密码不能为空"}), 401

    import bcrypt
    with get_db() as conn:
        user = conn.execute(
            "SELECT * FROM sys_user WHERE username=? AND is_active=1", (username,)
        ).fetchone()

        if user and bcrypt.checkpw(password.encode('utf-8'), user['password_hash'].encode('utf-8')):
            # 登录成功，重置失败次数
            conn.execute(
                "UPDATE sys_user SET login_attempts=0, locked_until=NULL WHERE id=?",
                (user['id'],)
            )
            audit_log(conn, user['id'], username, 'LOGIN', 'user', user['id'],
                      ip=_get_client_ip())

            # 生成简易 token（生产环境应使用 JWT）
            import hashlib
            token_raw = f"{user['id']}:{username}:{datetime.now().isoformat()}"
            token = hashlib.sha256(token_raw.encode()).hexdigest()
            session['username'] = username
            session['role'] = user['role']
            session['user_id'] = user['id']

            return jsonify({
                "code": 0,
                "message": "登录成功",
                "data": {
                    "token": token,
                    "username": username,
                    "real_name": user['real_name'],
                    "role": user['role'],
                    "user_id": user['id'],
                }
            })
        else:
            # 登录失败，累加失败次数
            if user:
                conn.execute(
                    "UPDATE sys_user SET login_attempts = login_attempts + 1 WHERE id=?",
                    (user['id'],)
                )
            return jsonify({"code": 401, "message": "用户名或密码错误"}), 401


# ================================================================
#  发票 API
# ================================================================

@app.route("/api/invoices/upload", methods=["POST"])
def upload_invoice():
    """上传并解析发票PDF"""
    if 'file' not in request.files:
        return jsonify({"code": 400, "message": "未上传文件"}), 400

    file = request.files['file']
    if not file.filename.lower().endswith('.pdf'):
        return jsonify({"code": 400, "message": "仅支持PDF文件"}), 400

    filename = secure_filename(file.filename)
    # 防止重名：加时间戳
    ts = datetime.now().strftime('%Y%m%d%H%M%S')
    filename = f"inv_{ts}_{filename}"
    filepath = str(UPLOAD_DIR / filename)
    file.save(filepath)

    try:
        data = parse_invoice_pdf(filepath)

        if data.get('errors') and not data.get('invoice_number'):
            return jsonify({
                "code": 500,
                "message": f"发票解析失败: {'; '.join(data['errors'])}",
                "data": data
            }), 500

        with get_db() as conn:
            cursor = conn.execute("""
                INSERT INTO inv_invoice (
                    invoice_number, invoice_date, invoice_type,
                    buyer_name, buyer_tax_id, seller_name, seller_tax_id,
                    total_amount_excl, total_tax, total_amount_incl,
                    amount_capital, source, pdf_path
                ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)
            """, (
                data['invoice_number'], data['invoice_date'], data['invoice_type'],
                data['buyer_name'], data['buyer_tax_id'],
                data['seller_name'], data['seller_tax_id'],
                data['total_amount_excl'], data['total_tax'], data['total_amount_incl'],
                data['amount_capital'], 'ocr', filepath
            ))
            invoice_id = cursor.lastrowid

            # 写入明细行
            for idx, item in enumerate(data.get('items', []), start=1):
                conn.execute("""
                    INSERT INTO inv_invoice_item (
                        invoice_id, line_number, category_prefix, material_name,
                        specification, unit, quantity, unit_price_excl,
                        amount_excl, tax_rate, tax_amount
                    ) VALUES (?,?,?,?,?,?,?,?,?,?,?)
                """, (
                    invoice_id, idx,
                    item.get('category_prefix', ''),
                    item.get('material_name', ''),
                    item.get('specification', ''),
                    item.get('unit', 'PCS'),
                    item.get('quantity', 0),
                    item.get('unit_price_excl', 0),
                    item.get('amount_excl', 0),
                    item.get('tax_rate', 13.0),
                    item.get('tax_amount', 0),
                ))

            audit_log(conn, None, 'system', 'CREATE', 'invoice', invoice_id,
                      new_values={"invoice_number": data['invoice_number']},
                      ip=_get_client_ip())

        data['id'] = invoice_id
        return jsonify({"code": 0, "message": "发票解析成功", "data": data})

    except Exception as e:
        logger.exception("发票上传处理异常")
        return jsonify({"code": 500, "message": f"服务器错误: {str(e)}"}), 500


@app.route("/api/invoices", methods=["GET"])
def list_invoices():
    """分页查询发票列表"""
    page = request.args.get('page', 1, type=int)
    size = request.args.get('size', 20, type=int)
    status = request.args.get('status', '')
    keyword = request.args.get('keyword', '')

    where_clauses = []
    params = []

    if status:
        where_clauses.append("status = ?")
        params.append(status)
    if keyword:
        where_clauses.append(
            "(invoice_number LIKE ? OR buyer_name LIKE ? OR seller_name LIKE ?)"
        )
        kw = f"%{keyword}%"
        params.extend([kw, kw, kw])

    where_sql = f"WHERE {' AND '.join(where_clauses)}" if where_clauses else ""

    with get_db() as conn:
        data, total, page = _paginate_query(
            conn,
            f"SELECT * FROM inv_invoice {where_sql} ORDER BY id DESC",
            f"SELECT count(*) FROM inv_invoice {where_sql}",
            page, size, params
        )

    return jsonify({"code": 0, "data": data, "total": total, "page": page, "size": size})


@app.route("/api/invoices/<int:invoice_id>", methods=["GET"])
def get_invoice(invoice_id):
    """获取发票详情（含明细行）"""
    with get_db() as conn:
        inv = conn.execute("SELECT * FROM inv_invoice WHERE id=?", (invoice_id,)).fetchone()
        if not inv:
            return jsonify({"code": 404, "message": "发票不存在"}), 404

        items = conn.execute(
            "SELECT * FROM inv_invoice_item WHERE invoice_id=? ORDER BY line_number",
            (invoice_id,)
        ).fetchall()

    result = dict_from_row(inv)
    result['items'] = rows_to_list(items)
    return jsonify({"code": 0, "data": result})


@app.route("/api/invoices/<int:invoice_id>", methods=["DELETE"])
def delete_invoice(invoice_id):
    """删除发票（级联删除明细）"""
    with get_db() as conn:
        inv = conn.execute("SELECT id, invoice_number FROM inv_invoice WHERE id=?", (invoice_id,)).fetchone()
        if not inv:
            return jsonify({"code": 404, "message": "发票不存在"}), 404

        conn.execute("DELETE FROM inv_invoice WHERE id=?", (invoice_id,))
        audit_log(conn, None, 'system', 'DELETE', 'invoice', invoice_id,
                  old_values={"invoice_number": inv['invoice_number']},
                  ip=_get_client_ip())

    return jsonify({"code": 0, "message": "删除成功"})


# ================================================================
#  对账单 API
# ================================================================

@app.route("/api/statements/upload", methods=["POST"])
def upload_statement():
    """上传并解析对账单 PDF/Excel"""
    if 'file' not in request.files:
        return jsonify({"code": 400, "message": "未上传文件"}), 400

    file = request.files['file']
    ext = Path(file.filename).suffix.lower()
    if ext not in ('.pdf', '.xlsx'):
        return jsonify({"code": 400, "message": "仅支持PDF或XLSX文件"}), 400

    filename = secure_filename(file.filename)
    ts = datetime.now().strftime('%Y%m%d%H%M%S')
    filename = f"stm_{ts}_{filename}"
    filepath = str(UPLOAD_DIR / filename)
    file.save(filepath)

    try:
        data = parse_statement_xlsx(filepath) if ext == '.xlsx' else parse_statement_pdf(filepath)
        supplier_code = data.get('supplier_code', '').strip()
        statement_month = data.get('statement_month', '').strip()
        if not supplier_code:
            return jsonify({
                "code": 422,
                "message": "未获取到有效合作商编码，请通过对账管理页选择ERP合作商",
            }), 422
        if not statement_month:
            return jsonify({
                "code": 422,
                "message": "未识别到账期，不允许保存空对账单",
            }), 422
        if not data.get("items"):
            return jsonify({
                "code": 422,
                "message": "未识别到有效对账明细，不允许保存空对账单",
            }), 422
        if supplier_code and statement_month:
            data['statement_key'] = data.get('statement_key') or f"{supplier_code}_{statement_month}"

        with get_db() as conn:
            if not supplier_code and statement_month and data.get('supplier_name'):
                existing = conn.execute("""
                    SELECT id FROM stm_statement
                    WHERE COALESCE(supplier_code, '') = ''
                      AND supplier_name = ?
                      AND statement_period = ?
                    LIMIT 1
                """, (data['supplier_name'], statement_month)).fetchone()
                if existing:
                    return jsonify({"code": 409, "message": "同一供应商和对账月份的对账单已存在"}), 409

            cursor = conn.execute("""
                INSERT INTO stm_statement (
                    statement_period, statement_date,
                    customer_name, customer_tax_id,
                    supplier_code, statement_key,
                    supplier_name, supplier_tax_id,
                    settlement_days,
                    opening_balance, current_payment, closing_balance,
                    delivered_unpaid, total_invoice_amount, total_quantity,
                    balance_status, source_file, pdf_path
                ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
            """, (
                statement_month,
                data.get('statement_date', ''),
                data.get('customer_name', '')[:512], data['customer_tax_id'],
                supplier_code, data.get('statement_key', ''),
                data.get('supplier_name', '')[:512], data['supplier_tax_id'],
                data.get('settlement_days', 30),
                data['opening_balance'], data['current_payment'],
                data['closing_balance'], data['delivered_unpaid'],
                data['total_invoice_amount'], data['total_quantity'],
                'balanced' if data.get('balance_check', True) else 'unbalanced',
                filename, filepath
            ))
            stmt_id = cursor.lastrowid

            # 写入明细行
            for idx, item in enumerate(data.get('items', []), start=1):
                conn.execute("""
                    INSERT INTO stm_statement_item (
                        statement_id, seq, customer_order_no,
                        customer_material_code, delivery_no, delivery_date,
                        product_name, specification, quantity, unit,
                        unit_price_incl_tax, amount_incl_tax
                    ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?)
                """, (
                    stmt_id, idx,
                    item.get('customer_order_no', ''),
                    item.get('customer_material_code', ''),
                    item.get('delivery_no', ''),
                    item.get('delivery_date', ''),
                    item.get('product_name', ''),
                    item.get('specification', ''),
                    item.get('quantity', 0),
                    item.get('unit', 'PCS'),
                    item.get('unit_price_incl_tax', 0),
                    item.get('amount_incl_tax', 0),
                ))

            audit_log(conn, None, 'system', 'CREATE', 'statement', stmt_id,
                      new_values={"period": data.get('statement_month', '')},
                      ip=_get_client_ip())

        data['id'] = stmt_id
        return jsonify({"code": 0, "message": "对账单解析成功", "data": data})

    except Exception as e:
        logger.exception("对账单上传处理异常")
        if 'UNIQUE constraint failed: stm_statement.supplier_code, stm_statement.statement_period' in str(e):
            return jsonify({"code": 409, "message": "同一供应商编码和对账月份的对账单已存在"}), 409
        return jsonify({"code": 500, "message": f"服务器错误: {str(e)}"}), 500


@app.route("/api/statements", methods=["GET"])
def list_statements():
    """分页查询对账单列表"""
    page = request.args.get('page', 1, type=int)
    size = request.args.get('size', 20, type=int)
    status = request.args.get('status', '')
    customer = request.args.get('customer', '')

    where_clauses = []
    params = []

    if status:
        where_clauses.append("status = ?")
        params.append(status)
    if customer:
        where_clauses.append("customer_name LIKE ?")
        params.append(f"%{customer}%")

    where_sql = f"WHERE {' AND '.join(where_clauses)}" if where_clauses else ""

    with get_db() as conn:
        data, total, page = _paginate_query(
            conn,
            f"SELECT * FROM stm_statement {where_sql} ORDER BY id DESC",
            f"SELECT count(*) FROM stm_statement {where_sql}",
            page, size, params
        )

    return jsonify({"code": 0, "data": data, "total": total, "page": page, "size": size})


@app.route("/api/statements", methods=["POST"])
def create_statement_from_management():
    """107 对账管理页上传入口。"""
    task_id = request.form.get('task_id') or ''
    recompare_from = request.form.get('recompare_from', '').strip()
    import_token = request.form.get('import_token', '').strip()
    selected_supplier_code = request.form.get('selected_supplier_code', '').strip()
    selected_supplier_name = request.form.get('selected_supplier_name', '').strip()
    if task_id:
        PROGRESS[task_id] = {"status": "running", "percent": 10, "step": 1, "total": 8, "message": "文件已接收"}
    file = request.files.get('statement_pdf') or request.files.get('file')
    if not file and not import_token:
        return jsonify({"error": "未上传文件"}), 400
    try:
        pending = _load_pending_statement(import_token) if import_token else None
        if import_token and not pending:
            return jsonify({"error": "识别结果已过期，请重新上传对账单"}), 400
        if pending:
            # The operator may run another fuzzy search in the confirmation
            # dialog. Validate the final selection against current ERP master
            # data instead of limiting it to the first OCR candidate list.
            selected = next((
                item for item in _find_supplier_candidates(selected_supplier_name)
                if item.get("code") == selected_supplier_code
                and item.get("name") == selected_supplier_name
            ), None)
            if not selected:
                return jsonify({
                    "error": "请选择有效且带编码的已有合作商",
                    "supplier_selection_invalid": True,
                }), 400
            if task_id:
                PROGRESS[task_id] = {
                    "status": "running", "percent": 40, "step": 3, "total": 8,
                    "message": "合作商已确认，正在提取对账单明细",
                }
            def report_ocr_progress(done, total, items):
                if not task_id or not total:
                    return
                percent = 40 + round(22 * done / total)
                detail = f"正在识别第 {done}/{total} 个页面"
                if items is not None:
                    detail += f"，本页提取 {items} 条"
                PROGRESS[task_id] = {
                    "status": "running", "percent": min(percent, 62),
                    "step": 3, "total": 8, "message": detail,
                }
            data, meta = _parse_saved_statement_cached(
                pending["saved"], progress_callback=report_ocr_progress
            )
            if task_id:
                PROGRESS[task_id] = {
                    "status": "running", "percent": 65, "step": 5, "total": 8,
                    "message": "明细提取完成，正在校验账期和金额",
                }
            data["supplier_name"] = selected["name"]
            data["supplier_code"] = selected["code"]
            meta["supplier_code"] = selected["code"]
            if not meta.get("statement_month"):
                return jsonify({"error": "未识别到账期，无法生成对账单唯一标识"}), 400
            meta["statement_key"] = f"{selected['code']}_{meta['statement_month']}"
            meta["statement_no"] = meta["statement_key"]
            meta["statement_fingerprint"] = _statement_fingerprint(
                selected["code"], selected["name"], data.get("items", [])
            )
        else:
            if task_id:
                PROGRESS[task_id] = {
                    "status": "running", "percent": 20, "step": 2, "total": 8,
                    "message": "正在识别合作商",
                }
            saved = _save_statement_upload(file)
            detected_supplier = _detect_supplier_from_saved_file(saved)
            if task_id:
                PROGRESS[task_id] = {
                    "status": "running", "percent": 20, "step": 2, "total": 8,
                    "message": "合作商名称已识别，正在查询 ERP 主数据",
                }
            candidates = _find_supplier_candidates(detected_supplier)
            token = secrets.token_urlsafe(24)
            _save_pending_statement(token, {
                "saved": saved, "candidates": candidates,
                "created_at": time.time(),
            })
            if task_id:
                PROGRESS[task_id] = {
                    "status": "waiting", "percent": 70, "step": 6, "total": 8,
                    "message": "请选择已有合作商",
                }
            return jsonify({
                "needs_supplier": True,
                "import_token": token,
                "detected_supplier": detected_supplier,
                "statement_month": "",
                "original_filename": saved["filename"],
                "preview_url": f"/api/statements/pending/{token}/preview",
                "candidates": candidates,
                "error": (
                    "请选择匹配的已有合作商"
                    if candidates else
                    "未查询到带编码的已有合作商，不允许继续处理"
                ),
            }), 422
        if task_id:
            PROGRESS[task_id] = {"status": "running", "percent": 70, "step": 6, "total": 8, "message": "正在保存结果"}
        with get_db() as conn:
            usage_remark = request.form.get('usage_remark', '').strip() or data.get('usage_remark', '')
            if recompare_from:
                stmt_id = _recompare_statement(conn, int(recompare_from), data, meta, usage_remark)
                duplicate = False
            else:
                stmt_id, duplicate = _insert_statement(conn, data, meta)
            if not duplicate:
                conn.execute("""
                    UPDATE stm_statement
                    SET usage_remark=?
                    WHERE id=?
                """, (
                    usage_remark,
                    stmt_id,
                ))
            row = conn.execute("SELECT * FROM stm_statement WHERE id=?", (stmt_id,)).fetchone()
        if duplicate:
            if task_id:
                PROGRESS[task_id] = {"status": "done", "percent": 100, "step": 8, "total": 8, "message": "该对账单已存在"}
            return jsonify({"duplicate": True, "existing": _history_row(row)}), 409
        if task_id:
            PROGRESS[task_id] = {"status": "done", "percent": 100, "step": 8, "total": 8, "message": "处理完成"}
        if import_token:
            _delete_pending_statement(import_token)
        return jsonify({"id": stmt_id, "summary": _history_row(row)})
    except StatementIdentityConflict as exc:
        with get_db() as conn:
            row = conn.execute(
                "SELECT * FROM stm_statement WHERE id=?", (exc.existing_id,)
            ).fetchone()
        if task_id:
            PROGRESS[task_id] = {
                "status": "error", "percent": 100, "step": 8, "total": 8,
                "message": str(exc),
            }
        return jsonify({
            "identity_conflict": True,
            "error": str(exc),
            "existing": _history_row(row) if row else {"id": exc.existing_id},
        }), 409
    except ValueError as exc:
        if task_id:
            PROGRESS[task_id] = {
                "status": "error", "percent": 100, "step": 8, "total": 8,
                "message": str(exc),
            }
        return jsonify({"error": str(exc)}), 400
    except Exception as exc:
        logger.exception("107 管理页上传处理失败")
        if task_id:
            PROGRESS[task_id] = {"status": "error", "percent": 100, "step": 8, "total": 8, "message": str(exc)}
        return jsonify({"error": str(exc)}), 500


@app.route("/api/statements/pending/<token>/preview")
def preview_pending_statement(token):
    pending = _load_pending_statement(token)
    if not pending:
        return jsonify({"error": "临时文件已过期"}), 404
    saved = pending.get("saved") or {}
    path = saved.get("filepath")
    if not path or not Path(path).is_file():
        return jsonify({"error": "临时文件不存在"}), 404
    return send_file(
        path,
        as_attachment=False,
        download_name=saved.get("filename") or Path(path).name,
    )


@app.route("/api/erp/suppliers")
def search_erp_suppliers():
    keyword = request.args.get("q", "").strip()
    if not keyword:
        return jsonify({"rows": []})
    try:
        return jsonify({"rows": _find_supplier_candidates(keyword)})
    except Exception:
        logger.exception("ERP 合作商模糊查询失败")
        return jsonify({"error": "ERP 合作商查询失败"}), 500


@app.route("/api/statements/<int:stmt_id>", methods=["GET"])
def get_statement(stmt_id):
    """获取对账单详情（含明细行）"""
    with get_db() as conn:
        stmt = conn.execute("SELECT * FROM stm_statement WHERE id=?", (stmt_id,)).fetchone()
        if not stmt:
            return jsonify({"code": 404, "message": "对账单不存在"}), 404

        items = conn.execute(
            "SELECT * FROM stm_statement_item WHERE statement_id=? ORDER BY seq",
            (stmt_id,)
        ).fetchall()

    result = dict_from_row(stmt)
    result['items'] = rows_to_list(items)
    return jsonify({"code": 0, "data": result})


@app.route("/api/statements/<int:stmt_id>", methods=["DELETE"])
def delete_statement(stmt_id):
    """删除对账单"""
    with get_db() as conn:
        stmt = conn.execute("SELECT id, statement_period FROM stm_statement WHERE id=?", (stmt_id,)).fetchone()
        if not stmt:
            return jsonify({"code": 404, "message": "对账单不存在"}), 404

        conn.execute("DELETE FROM stm_statement WHERE id=?", (stmt_id,))
        audit_log(conn, None, 'system', 'DELETE', 'statement', stmt_id,
                  old_values={"period": stmt['statement_period']},
                  ip=_get_client_ip())

    return jsonify({"code": 0, "message": "删除成功"})


@app.route("/api/history", methods=["GET"])
def history_list():
    """107 对账管理页列表接口。"""
    page = request.args.get('page', 1, type=int)
    size = request.args.get('page_size', request.args.get('size', 20), type=int)
    status = request.args.get('status', '')
    search_mode = request.args.get('search_mode', 'fuzzy')
    exact = search_mode == 'exact'
    keyword = request.args.get('keyword', '').strip()
    created_at = request.args.get('created_at', '').strip()
    statement_no = request.args.get('statement_no', '')
    reconciliation_key = request.args.get('reconciliation_key', '')
    supplier = request.args.get('supplier', '')
    supplier_code = request.args.get('supplier_code', '').strip()
    usage_remark = request.args.get('usage_remark', '').strip()
    payment_keyword = request.args.get('payment_keyword', '').strip()
    reconciliation_log = request.args.get('reconciliation_log', '').strip()
    invoice_log = request.args.get('invoice_log', '').strip()
    payment_date = request.args.get('payment_date', '').strip()
    statement_start = request.args.get('statement_start', '').strip()
    statement_end = request.args.get('statement_end', '').strip()
    payment_start = request.args.get('payment_start', '').strip()
    payment_end = request.args.get('payment_end', '').strip()
    invoice_start = request.args.get('invoice_start', '').strip()
    invoice_end = request.args.get('invoice_end', '').strip()
    amount_min = request.args.get('amount_min', '').strip()
    amount_max = request.args.get('amount_max', '').strip()
    payment_amount_min = request.args.get('payment_amount_min', '').strip()
    payment_amount_max = request.args.get('payment_amount_max', '').strip()
    invoice_amount_min = request.args.get('invoice_amount_min', '').strip()
    invoice_amount_max = request.args.get('invoice_amount_max', '').strip()
    payment_statuses = [x for x in request.args.getlist('payment_status') if x]
    invoice_statuses = [x for x in request.args.getlist('invoice_amount_status') if x]
    start_date = request.args.get('start_date', '')
    end_date = request.args.get('end_date', '')
    where, params = [], []
    if status:
        where.append("(overall_status=? OR status=?)")
        params.extend([status, status])
    if keyword:
        like = f"%{keyword}%"
        where.append("""(
            statement_no LIKE ? OR reconciliation_key LIKE ? OR statement_key LIKE ?
            OR supplier_name LIKE ? OR original_filename LIKE ? OR source_file LIKE ?
            OR invoice_date LIKE ? OR payment_date LIKE ? OR usage_remark LIKE ?
            OR created_at LIKE ? OR statement_period LIKE ?
            OR EXISTS (
                SELECT 1 FROM stm_statement_record r
                WHERE r.statement_id=stm_statement.id
                  AND (r.title LIKE ? OR r.text_content LIKE ? OR r.record_date LIKE ? OR r.file_name LIKE ?)
            )
        )""")
        params.extend([like] * 15)
    if statement_no:
        where.append(f"statement_no {_search_operator(exact)} ?")
        params.append(_search_value(statement_no, exact))
    if reconciliation_key:
        op = _search_operator(exact)
        where.append(f"(reconciliation_key {op} ? OR statement_key {op} ?)")
        params.extend([_search_value(reconciliation_key, exact)] * 2)
    if supplier:
        where.append(f"supplier_name {_search_operator(exact)} ?")
        params.append(_search_value(supplier, exact))
    if supplier_code:
        where.append(f"supplier_code {_search_operator(exact)} ?")
        params.append(_search_value(supplier_code, exact))
    if created_at:
        where.append(f"created_at {_search_operator(exact)} ?")
        params.append(_search_value(created_at, exact))
    if usage_remark:
        where.append(f"usage_remark {_search_operator(exact)} ?")
        params.append(_search_value(usage_remark, exact))
    if payment_date:
        where.append(f"payment_date {_search_operator(exact)} ?")
        params.append(_search_value(payment_date, exact))
    if statement_start:
        where.append("statement_date >= ?")
        params.append(statement_start)
    if statement_end:
        where.append("statement_date <= ?")
        params.append(statement_end)
    if amount_min:
        where.append("total_invoice_amount >= ?")
        params.append(float(amount_min))
    if amount_max:
        where.append("total_invoice_amount <= ?")
        params.append(float(amount_max))
    record_filters = [
        (payment_keyword, ['付款', '支付']),
        (reconciliation_log, ['对账']),
        (invoice_log, ['开票', '发票']),
    ]
    for value, category_words in record_filters:
        if value:
            op = _search_operator(exact)
            category_parts = []
            category_params = []
            for word in category_words:
                category_parts.extend(["r.title LIKE ?", "r.text_content LIKE ?"])
                category_params.extend([f"%{word}%", f"%{word}%"])
            category_sql = "(" + " OR ".join(category_parts) + ")"
            where.append(f"""EXISTS (
                SELECT 1 FROM stm_statement_record r
                WHERE r.statement_id=stm_statement.id
                  AND {category_sql}
                  AND (r.title {op} ? OR r.text_content {op} ? OR r.record_date {op} ? OR r.file_name {op} ?)
            )""")
            params.extend(category_params)
            params.extend([_search_value(value, exact)] * 4)
    if start_date:
        where.append("created_at >= ?")
        params.append(start_date)
    if end_date:
        where.append("created_at <= ?")
        params.append(end_date + " 23:59:59")
    where_sql = "WHERE " + " AND ".join(where) if where else ""
    post_filters = any([payment_start, payment_end, invoice_start, invoice_end, payment_amount_min,
                        payment_amount_max, invoice_amount_min, invoice_amount_max, payment_statuses, invoice_statuses])
    def keep(row):
        if payment_start and (row.get('payment_date') or '') < payment_start:
            return False
        if payment_end and (row.get('payment_date') or '') > payment_end:
            return False
        if invoice_start and (row.get('invoice_date') or '') < invoice_start:
            return False
        if invoice_end and (row.get('invoice_date') or '') > invoice_end:
            return False
        if payment_amount_min and float(row.get('payment_amount') or 0) < float(payment_amount_min):
            return False
        if payment_amount_max and float(row.get('payment_amount') or 0) > float(payment_amount_max):
            return False
        if invoice_amount_min and float(row.get('invoice_amount') or 0) < float(invoice_amount_min):
            return False
        if invoice_amount_max and float(row.get('invoice_amount') or 0) > float(invoice_amount_max):
            return False
        if payment_statuses and row.get('payment_status') not in payment_statuses:
            return False
        if invoice_statuses and row.get('invoice_amount_status') not in invoice_statuses:
            return False
        return True
    with get_db() as conn:
        summary_rows = conn.execute(f"SELECT * FROM stm_statement {where_sql} ORDER BY created_at DESC, id DESC", params).fetchall()
        summary_rows = _enrich_history_rows(summary_rows)
        if post_filters:
            summary_rows = [row for row in summary_rows if keep(row)]
        summary = _history_summary(summary_rows)
        if post_filters:
            history_rows = summary_rows
            total = len(history_rows)
            start = (page - 1) * size
            history_rows = history_rows[start:start + size]
        else:
            rows, total, _ = _paginate_query(
                conn,
                f"SELECT * FROM stm_statement {where_sql} ORDER BY created_at DESC, id DESC",
                f"SELECT count(*) FROM stm_statement {where_sql}",
                page, size, params
            )
            history_rows = _enrich_history_rows(rows)
    return jsonify({"rows": history_rows, "total": total, "page": page, "summary": summary})


@app.route("/api/history/<int:stmt_id>", methods=["GET"])
def history_detail(stmt_id):
    with get_db() as conn:
        stmt = conn.execute("SELECT * FROM stm_statement WHERE id=?", (stmt_id,)).fetchone()
        if not stmt:
            return jsonify({"error": "对账记录不存在"}), 404
        items = conn.execute(
            "SELECT * FROM stm_statement_item WHERE statement_id=? ORDER BY seq", (stmt_id,)
        ).fetchall()
        records = conn.execute("""
            SELECT id, title, record_type, record_date, text_content,
                   amount, file_name, created_by, created_at
            FROM stm_statement_record
            WHERE statement_id=?
            ORDER BY created_at DESC, id DESC
        """, (stmt_id,)).fetchall()
        allocations = conn.execute("""
            SELECT *
            FROM stm_statement_allocation
            WHERE statement_id=?
        """, (stmt_id,)).fetchall()
        overrides = conn.execute("""
            SELECT * FROM stm_statement_line_override
            WHERE statement_id=?
        """, (stmt_id,)).fetchall()
        if items and not allocations:
            _sync_statement_allocations(conn, stmt_id)
            allocations = conn.execute("""
                SELECT *
                FROM stm_statement_allocation
                WHERE statement_id=?
            """, (stmt_id,)).fetchall()
    summary = _enrich_history_rows([stmt])[0]
    try:
        recognition_metadata = json.loads(stmt.get("recognition_metadata") or "{}")
    except (TypeError, ValueError):
        recognition_metadata = {}
    allocation_by_item = {row["statement_item_id"]: dict(row) for row in allocations}
    override_by_item = {row["statement_item_id"]: dict(row) for row in overrides}
    statement_lines = []
    line_checks = []
    shared_erp_conn = None
    if stmt.get('supplier_code') and ERP_MYSQL_CONFIG.get("host"):
        try:
            shared_erp_conn = _open_erp_read_connection()
        except Exception as exc:
            logger.warning("ERP详情批量读取连接失败 statement=%s: %s", stmt_id, exc)
    for item in items:
        item = dict(item)
        alloc = allocation_by_item.get(item.get('id'), {})
        override = override_by_item.get(item.get('id'), {})
        manual_approved = override.get("override_type") == "MANUAL_PASS"
        # An operator's decision is final. ERP is only an aid, so a manually
        # approved line must not be queried or revalidated against ERP again.
        erp = {} if manual_approved else _erp_statement_line(
            stmt.get('supplier_code') or '',
            item.get('customer_material_code') or '',
            item.get('customer_order_no') or '',
            item.get('delivery_date') or '',
            erp_conn=shared_erp_conn,
        )
        amount = item.get('amount_incl_tax') or 0
        qty = item.get('quantity') or 0
        erp_qty = erp.get("erp_arrival_quantity")
        erp_price = erp.get("erp_unit_price")
        erp_amount = erp.get("erp_amount")
        cumulative_qty = float(alloc.get("cumulative_quantity", qty) or 0)
        cumulative_amount = float(alloc.get("cumulative_amount", amount) or 0)
        quantity_status = (
            "PENDING" if erp_qty is None
            else ("PASS" if cumulative_qty <= float(erp_qty) + 0.000001 else "FAIL")
        )
        price_matches = (
            erp_price is not None
            and abs(float(item.get('unit_price_incl_tax') or 0) - float(erp_price)) <= 0.000001
        )
        amount_status = (
            "PENDING" if erp_amount is None
            else (
                "PASS"
                if price_matches and cumulative_amount <= float(erp_amount) + 0.05
                else "FAIL"
            )
        )
        if manual_approved:
            quantity_status = "PASS"
            amount_status = "PASS"
        statement_lines.append({
            "item_id": item.get('id'),
            "customer_order_no": item.get('customer_order_no') or '',
            "customer_material_no": item.get('customer_material_code') or '',
            "delivery_no": item.get('delivery_no') or '',
            "delivery_date": item.get('delivery_date') or '',
            "product_name": item.get('product_name') or '',
            "supplier_spec": item.get('specification') or '',
            "quantity": qty,
            "unit": item.get('unit') or '',
            "tax_inclusive_unit_price": item.get('unit_price_incl_tax') or 0,
            "tax_inclusive_amount": amount,
            "row_status": "PASS" if not _recognition_row_issues(item) else "FAIL",
            "row_issues": _recognition_row_issues(item),
        })
        line_checks.append({
            "statement_item_id": item.get('id'),
            "material_code": item.get('customer_material_code') or '',
            "purchase_order_id": item.get('customer_order_no') or '',
            "delivery_date": item.get('delivery_date') or '',
            "erp_order_dates": erp.get("erp_order_date") or '',
            "erp_arrival_dates": erp.get("erp_arrival_date") or '',
            "statement_quantity": qty,
            "erp_purchase_quantity": erp.get("erp_purchase_quantity"),
            "arrival_record_quantity": erp_qty,
            "statement_unit_price": item.get('unit_price_incl_tax') or 0,
            "erp_unit_price": erp_price,
            "statement_amount": amount,
            "erp_amount": erp_amount,
            "historical_quantity": alloc.get("historical_quantity", 0),
            "historical_amount": alloc.get("historical_amount", 0),
            "current_quantity": alloc.get("current_quantity", qty),
            "current_amount": alloc.get("current_amount", amount),
            "cumulative_quantity": cumulative_qty,
            "cumulative_amount": cumulative_amount,
            "remaining_quantity": alloc.get("remaining_quantity", 0),
            "remaining_amount": alloc.get("remaining_amount", 0),
            "allocation_status": alloc.get("allocation_status", "INFO"),
            "manual_approved": manual_approved,
            "decision_source": "MANUAL" if manual_approved else "ERP",
            "manual_approved_by": override.get("approved_by") or "",
            "manual_approved_at": str(override.get("approved_at") or ""),
            "quantity_status": quantity_status,
            "amount_status": amount_status,
            "issue_text": (
                f"人工确认无误，忽略ERP差异（{override.get('approved_by') or '人工'}）"
                if manual_approved else
                (alloc.get("issue_text") or "ERP未找到对应采购到库明细")
            ),
        })
    if shared_erp_conn:
        shared_erp_conn.close()
    return jsonify({
        "summary": summary,
        "statement_no": summary["statement_no"],
        "supplier": summary["supplier"],
        "statement_total": summary["statement_total"],
        "erp_purchase_total": summary["erp_purchase_total"],
        "recognition_metadata": recognition_metadata,
        "files": [],
        "records": [
            {
                **dict(row),
                "download_url": f"/api/history/{stmt_id}/records/{row['id']}/download" if row['file_name'] else "",
            }
            for row in records
        ],
        "line_checks": _merge_compare_line_checks(line_checks),
        "statement_lines": statement_lines,
    })


@app.route(
    "/api/statements/<int:stmt_id>/items/<int:item_id>/allocation-history",
    methods=["GET"],
)
def statement_allocation_history(stmt_id, item_id):
    """Return every statement line contributing to the clicked cumulative value."""
    with get_db() as conn:
        current = conn.execute("""
            SELECT a.line_key, a.material_code, a.unit_price, s.supplier_name
            FROM stm_statement_allocation a
            JOIN stm_statement s ON s.id=a.statement_id
            WHERE a.statement_id=? AND a.statement_item_id=?
            LIMIT 1
        """, (stmt_id, item_id)).fetchone()
        if not current:
            return jsonify({"error": "未找到该明细的占用记录"}), 404
        rows = conn.execute("""
            SELECT
                a.statement_id,
                a.statement_item_id,
                s.statement_no,
                s.statement_period,
                s.original_filename,
                s.created_at,
                a.purchase_order_id,
                a.material_code,
                a.delivery_date,
                a.unit_price,
                a.current_quantity,
                a.current_amount,
                a.cumulative_quantity,
                a.cumulative_amount
            FROM stm_statement_allocation a
            JOIN stm_statement s ON s.id=a.statement_id
            WHERE a.line_key=?
            ORDER BY s.created_at, s.id, a.id
        """, (current["line_key"],)).fetchall()
    return jsonify({
        "supplier": current["supplier_name"],
        "material_code": current["material_code"],
        "unit_price": current["unit_price"],
        "rows": [
            {**dict(row), "is_current": row["statement_id"] == stmt_id}
            for row in rows
        ],
    })


@app.route("/api/statements/<int:stmt_id>/items/<int:item_id>", methods=["PUT"])
def update_statement_item(stmt_id, item_id):
    """人工修正识别明细；保存后重算汇总、指纹和占用记录。"""
    payload = request.get_json(silent=True) or {}
    text_fields = (
        "customer_order_no", "customer_material_code", "delivery_no",
        "delivery_date", "product_name", "specification", "unit",
    )
    values = {field: str(payload.get(field) or "").strip() for field in text_fields}
    try:
        values["quantity"] = float(payload.get("quantity") or 0)
        values["unit_price_incl_tax"] = float(payload.get("unit_price_incl_tax") or 0)
        values["amount_incl_tax"] = float(payload.get("amount_incl_tax") or 0)
    except (TypeError, ValueError):
        return jsonify({"error": "数量、单价和金额必须是数字"}), 400
    issues = _recognition_row_issues(values)

    with get_db() as conn:
        statement = conn.execute(
            "SELECT * FROM stm_statement WHERE id=?", (stmt_id,)
        ).fetchone()
        item = conn.execute(
            "SELECT id FROM stm_statement_item WHERE id=? AND statement_id=?",
            (item_id, stmt_id),
        ).fetchone()
        if not statement or not item:
            return jsonify({"error": "对账单明细不存在"}), 404
        conn.execute("""
            UPDATE stm_statement_item
            SET customer_order_no=?, customer_material_code=?, delivery_no=?,
                delivery_date=?, product_name=?, specification=?, quantity=?,
                unit=?, unit_price_incl_tax=?, amount_incl_tax=?
            WHERE id=? AND statement_id=?
        """, (
            values["customer_order_no"], values["customer_material_code"],
            values["delivery_no"], values["delivery_date"],
            values["product_name"], values["specification"], values["quantity"],
            values["unit"], values["unit_price_incl_tax"],
            values["amount_incl_tax"], item_id, stmt_id,
        ))
        rows = [
            dict(row) for row in conn.execute(
                "SELECT * FROM stm_statement_item WHERE statement_id=? ORDER BY seq",
                (stmt_id,),
            ).fetchall()
        ]
        total_qty = sum(float(row.get("quantity") or 0) for row in rows)
        total_amount = round(
            sum(float(row.get("amount_incl_tax") or 0) for row in rows), 2
        )
        all_issues = [
            issue for row in rows for issue in _recognition_row_issues(row)
        ]
        fingerprint = _statement_fingerprint(
            statement.get("supplier_code") or "",
            statement.get("supplier_name") or "",
            rows,
        )
        conn.execute("""
            UPDATE stm_statement
            SET total_quantity=?, total_invoice_amount=?, closing_balance=?,
                statement_fingerprint=?, overall_status=?, updated_at=CURRENT_TIMESTAMP
            WHERE id=?
        """, (
            total_qty, total_amount, total_amount, fingerprint,
            "ERP_FAILED" if all_issues else "ERP_PENDING", stmt_id,
        ))
        _sync_statement_allocations(conn, stmt_id)
        audit_log(
            conn, session.get('user_id'), session.get('username') or 'system',
            "UPDATE", "statement_item", item_id,
            new_values=values, ip=_get_client_ip(),
        )
    return jsonify({
        "ok": True,
        "row_status": "PASS" if not issues else "FAIL",
        "row_issues": issues,
        "total_quantity": total_qty,
        "total_amount": total_amount,
    })


@app.route("/api/history/<int:stmt_id>/usage", methods=["PATCH"])
def history_update_usage(stmt_id):
    """单独更新用途备注，不触发重新识别或重新比对。"""
    payload = request.get_json(silent=True) or {}
    usage_remark = str(payload.get('usage_remark') or '').strip()
    if len(usage_remark) > 200:
        return jsonify({"error": "用途不能超过 200 个字符"}), 400
    with get_db() as conn:
        row = conn.execute(
            "SELECT usage_remark FROM stm_statement WHERE id=?", (stmt_id,)
        ).fetchone()
        if not row:
            return jsonify({"error": "对账单不存在"}), 404
        old_value = dict(row).get('usage_remark') or ''
        conn.execute(
            "UPDATE stm_statement SET usage_remark=? WHERE id=?",
            (usage_remark, stmt_id),
        )
        audit_log(
            conn, session.get('user_id'), session.get('username') or 'system',
            'UPDATE', 'statement', stmt_id,
            old_values={"usage_remark": old_value},
            new_values={"usage_remark": usage_remark},
            ip=_get_client_ip(),
        )
    return jsonify({"id": stmt_id, "usage_remark": usage_remark})


@app.route("/api/history/<int:stmt_id>", methods=["DELETE"])
def history_delete(stmt_id):
    with get_db() as conn:
        row = conn.execute("SELECT id FROM stm_statement WHERE id=?", (stmt_id,)).fetchone()
        if not row:
            return jsonify({"error": "对账记录不存在"}), 404
        conn.execute("DELETE FROM stm_statement_item WHERE statement_id=?", (stmt_id,))
        conn.execute("DELETE FROM stm_statement WHERE id=?", (stmt_id,))
        audit_log(conn, None, 'system', 'DELETE', 'statement', stmt_id, ip=_get_client_ip())
    return jsonify({"ok": True})


@app.route("/api/history/supplier-summary", methods=["GET"])
def supplier_summary():
    with get_db() as conn:
        rows = conn.execute("SELECT * FROM stm_statement ORDER BY created_at DESC, id DESC").fetchall()
    history_rows = _enrich_history_rows(rows)
    grouped = {}
    for row in history_rows:
        key = row.get('supplier') or '-'
        item = grouped.setdefault(key, {
            "supplier": key,
            "supplier_code": "",
            "statement_count": 0,
            "statement_amount": 0.0,
            "payment_amount": 0.0,
            "invoice_amount": 0.0,
            "unpaid_count": 0,
            "partial_paid_count": 0,
            "paid_count": 0,
            "over_paid_count": 0,
            "uninvoiced_count": 0,
            "partial_invoice_count": 0,
            "invoiced_count": 0,
            "over_invoice_count": 0,
        })
        item["supplier_code"] = item["supplier_code"] or row.get("supplier_code", "") or _resolve_supplier_code(row.get("supplier"))
        item["statement_count"] += 1
        item["statement_amount"] += float(row.get("statement_total") or 0)
        item["payment_amount"] += float(row.get("payment_amount") or 0)
        item["invoice_amount"] += float(row.get("invoice_amount") or 0)
        if row.get("payment_status") == "PAID":
            item["paid_count"] += 1
        elif row.get("payment_status") == "PARTIAL_PAID":
            item["partial_paid_count"] += 1
        elif row.get("payment_status") == "OVER_PAID":
            item["over_paid_count"] += 1
        else:
            item["unpaid_count"] += 1
        if row.get("invoice_amount_status") == "INVOICED":
            item["invoiced_count"] += 1
        elif row.get("invoice_amount_status") == "UNDER_INVOICED":
            item["partial_invoice_count"] += 1
        elif row.get("invoice_amount_status") == "OVER_INVOICED":
            item["over_invoice_count"] += 1
        else:
            item["uninvoiced_count"] += 1
    data = []
    for item in grouped.values():
        for key in ("statement_amount", "payment_amount", "invoice_amount"):
            item[key] = f"{item[key]:.2f}"
        data.append(item)
    data.sort(key=lambda x: float(x["statement_amount"]), reverse=True)
    return jsonify({"rows": data, "total": len(data)})


@app.route("/api/history/supplier-summary/details", methods=["GET"])
def supplier_summary_details():
    supplier = request.args.get("supplier", "")
    supplier_code = request.args.get("supplier_code", "")
    with get_db() as conn:
        rows = conn.execute("SELECT * FROM stm_statement ORDER BY created_at DESC, id DESC").fetchall()
    history_rows = _enrich_history_rows(rows)
    details = []
    for row in history_rows:
        if not row.get("supplier_code"):
            row["supplier_code"] = _resolve_supplier_code(row.get("supplier"))
        row_supplier = row.get("supplier") or "-"
        row_supplier_code = row.get("supplier_code") or "-"
        if supplier and row_supplier != supplier:
            continue
        if supplier_code and row_supplier_code != supplier_code:
            continue
        details.append(row)
    return jsonify({"rows": details, "total": len(details), "summary": _history_summary(details)})


@app.route("/api/history/<int:stmt_id>/runs", methods=["GET"])
def history_runs(stmt_id):
    with get_db() as conn:
        row = conn.execute("SELECT * FROM stm_statement WHERE id=?", (stmt_id,)).fetchone()
    return jsonify({"rows": [_history_row(row)] if row else []})


@app.route("/api/history/<int:stmt_id>/records", methods=["GET", "POST"])
def statement_records(stmt_id):
    with get_db() as conn:
        statement = conn.execute("SELECT id FROM stm_statement WHERE id=?", (stmt_id,)).fetchone()
        if not statement:
            return jsonify({"error": "对账记录不存在"}), 404
        if request.method == "POST":
            title = request.form.get('title', '').strip()
            record_date = request.form.get('record_date', '').strip()
            text_content = request.form.get('text_content', '').strip()
            amount = float(request.form.get('amount') or 0)
            file = request.files.get('file')
            if not title:
                return jsonify({"error": "标题必填"}), 400
            file_path = ''
            file_name = ''
            record_type = 'text'
            if file and file.filename:
                file_name = _display_filename(file.filename)
                suffix = Path(file_name).suffix
                stored = f"record_{stmt_id}_{datetime.now().strftime('%Y%m%d%H%M%S%f')}{suffix}"
                target = STATEMENT_RECORD_DIR / stored
                file.save(target)
                file_path = str(target)
                record_type = 'image' if suffix.lower() in ('.png', '.jpg', '.jpeg', '.gif', '.webp', '.bmp') else 'file'
            cursor = conn.execute("""
                INSERT INTO stm_statement_record (
                    statement_id, title, record_type, record_date, amount, text_content,
                    file_path, file_name, created_by
                ) VALUES (?,?,?,?,?,?,?,?,?)
            """, (
                stmt_id, title, record_type, record_date, amount, text_content,
                file_path, file_name, session.get('username', ''),
            ))
            audit_log(conn, None, session.get('username', 'system'), 'CREATE', 'statement_record', cursor.lastrowid,
                      new_values={"statement_id": stmt_id, "title": title}, ip=_get_client_ip())
            if record_date and ('付款' in title or '支付' in title):
                conn.execute("""
                    UPDATE stm_statement
                    SET payment_date=?
                    WHERE id=?
                """, (record_date, stmt_id))
        rows = conn.execute("""
            SELECT id, statement_id, title, record_type, record_date, text_content,
                   amount, file_name, created_by, created_at
            FROM stm_statement_record
            WHERE statement_id=?
            ORDER BY created_at DESC, id DESC
        """, (stmt_id,)).fetchall()
    data = []
    for row in rows:
        item = dict(row)
        item['download_url'] = f"/api/history/{stmt_id}/records/{item['id']}/download" if item.get('file_name') else ''
        data.append(item)
    return jsonify({"rows": data})


@app.route("/api/history/<int:stmt_id>/records/<int:record_id>/download", methods=["GET"])
def statement_record_download(stmt_id, record_id):
    with get_db() as conn:
        row = conn.execute("""
            SELECT file_path, file_name
            FROM stm_statement_record
            WHERE id=? AND statement_id=?
        """, (record_id, stmt_id)).fetchone()
    if not row or not row['file_path'] or not Path(row['file_path']).exists():
        return jsonify({"error": "附件不存在"}), 404
    path = Path(row['file_path'])
    return send_file(path, as_attachment=True, download_name=row['file_name'] or path.name)


@app.route("/api/history/<int:stmt_id>/lines/<int:line_index>/approve", methods=["POST"])
def history_line_approve(stmt_id, line_index):
    return jsonify({"ok": True, "id": stmt_id, "line": line_index})


@app.route(
    "/api/statements/<int:stmt_id>/items/<int:item_id>/manual-approve",
    methods=["POST"],
)
def manual_approve_statement_line(stmt_id, item_id):
    payload = request.get_json(silent=True) or {}
    reason = str(payload.get("reason") or "人工确认订单无问题，忽略ERP差异").strip()
    approved_by = str(session.get("username") or "人工").strip()
    with get_db() as conn:
        item = conn.execute("""
            SELECT id FROM stm_statement_item
            WHERE id=? AND statement_id=?
        """, (item_id, stmt_id)).fetchone()
        if not item:
            return jsonify({"error": "对账单明细不存在"}), 404
        conn.execute("""
            INSERT INTO stm_statement_line_override (
                statement_id, statement_item_id, override_type,
                reason, approved_by, approved_at
            ) VALUES (?,?,'MANUAL_PASS',?,?,CURRENT_TIMESTAMP)
            ON DUPLICATE KEY UPDATE
                override_type='MANUAL_PASS', reason=VALUES(reason),
                approved_by=VALUES(approved_by), approved_at=CURRENT_TIMESTAMP
        """, (stmt_id, item_id, reason, approved_by))
        audit_log(
            conn, session.get("user_id"), approved_by, "MANUAL_APPROVE",
            "statement_item", item_id,
            new_values={"statement_id": stmt_id, "reason": reason},
            ip=_get_client_ip(),
        )
    return jsonify({"ok": True, "message": "已人工确认该订单无问题"})


@app.route("/api/statements/<int:stmt_id>/approve-statement", methods=["POST"])
def approve_statement_107(stmt_id):
    with get_db() as conn:
        conn.execute("""
            UPDATE stm_statement
            SET status='confirmed', overall_status='COMPLETED'
            WHERE id=?
        """, (stmt_id,))
    return jsonify({"ok": True})


@app.route("/api/progress/<task_id>", methods=["GET"])
def progress(task_id):
    data = PROGRESS.get(task_id)
    if data is None:
        return jsonify({
            "status": "error", "percent": 100, "step": 8, "total": 8,
            "message": "任务不存在或已过期，请重新提交",
        }), 404
    return jsonify(data)


@app.route("/api/statements/export", methods=["GET"])
def export_supplier_107():
    with get_db() as conn:
        rows = conn.execute("SELECT * FROM stm_statement ORDER BY id DESC").fetchall()
    output = io.StringIO()
    output.write("statement_no,reconciliation_key,supplier,statement_total,overall_status,created_at\n")
    for row in rows:
        h = _history_row(row)
        output.write(f"{h['statement_no']},{h['reconciliation_key']},{h['supplier']},{h['statement_total']},{h['overall_status']},{h['created_at']}\n")
    data = io.BytesIO(output.getvalue().encode('utf-8-sig'))
    return send_file(data, mimetype='text/csv', as_attachment=True, download_name='statements.csv')


@app.route("/api/statements/template", methods=["GET"])
def download_statement_template():
    template_path = Path(__file__).resolve().parent.parent / 'templates' / '对账单统一模板_v1.xlsx'
    if not template_path.exists():
        return jsonify({"error": "对账单模板不存在"}), 404
    return send_file(
        template_path,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='对账单统一模板_v1.xlsx',
    )


@app.route("/api/statements/<int:stmt_id>/invoice/preview", methods=["POST"])
def invoice_preview_107(stmt_id):
    task_id = request.form.get('task_id') or ''
    if task_id:
        PROGRESS[task_id] = {"status": "running", "percent": 30, "step": 2, "total": 4, "message": "正在识别发票"}
    file = request.files.get('invoice_pdf')
    if not file:
        return jsonify({"error": "请选择发票 PDF"}), 400
    filename = secure_filename(file.filename)
    filepath = str(UPLOAD_DIR / f"inv_{datetime.now().strftime('%Y%m%d%H%M%S')}_{filename}")
    file.save(filepath)
    try:
        invoice = parse_invoice_pdf(filepath)
    except Exception as exc:
        return jsonify({"error": str(exc)}), 500
    with get_db() as conn:
        stmt = conn.execute("SELECT * FROM stm_statement WHERE id=?", (stmt_id,)).fetchone()
    total = float((dict(stmt).get('total_invoice_amount') if stmt else 0) or 0)
    result = {
        "invoice": {
            "invoice_number": invoice.get('invoice_number') or '',
            "invoice_date": invoice.get('invoice_date') or '',
            "invoice_total": invoice.get('total_amount_incl') or invoice.get('invoice_total') or 0,
            "raw_text": invoice.get('raw_text') or '',
        },
        "statement_total": total,
    }
    if task_id:
        PROGRESS[task_id] = {"status": "done", "percent": 100, "step": 4, "total": 4, "message": "识别完成"}
    return jsonify(result)


@app.route("/api/statements/<int:stmt_id>/invoice/confirm", methods=["POST"])
def invoice_confirm_107(stmt_id):
    data = request.get_json() or {}
    invoice_total = float(data.get('invoice_total') or 0)
    with get_db() as conn:
        stmt = conn.execute("SELECT * FROM stm_statement WHERE id=?", (stmt_id,)).fetchone()
        if not stmt:
            return jsonify({"error": "对账记录不存在"}), 404
        statement_total = float(dict(stmt).get('total_invoice_amount') or 0)
        invoice_date = str(data.get('invoice_date') or '').strip()
        invoice_number = str(data.get('invoice_number') or '').strip()
        record_text = f"发票号码：{invoice_number or '-'}；发票金额：{invoice_total:.2f}"
        cursor = conn.execute("""
            INSERT INTO stm_statement_record (
                statement_id, title, record_type, record_date, amount, text_content,
                file_path, file_name, created_by
            ) VALUES (?,?,?,?,?,?,?,?,?)
        """, (
            stmt_id, '开票记录', 'text', invoice_date, invoice_total, record_text,
            '', '', session.get('username', ''),
        ))
        audit_log(conn, None, session.get('username', 'system'), 'CREATE', 'statement_record', cursor.lastrowid,
                  new_values={"statement_id": stmt_id, "title": "开票记录"}, ip=_get_client_ip())
        invoice_sum = conn.execute("""
            SELECT COALESCE(SUM(amount), 0) AS amount
            FROM stm_statement_record
            WHERE statement_id=? AND (title LIKE ? OR title LIKE ?)
        """, (stmt_id, '%开票%', '%发票%')).fetchone()['amount'] or 0
        invoice_state = _money_status(statement_total, invoice_sum, 'invoice')
        passed = invoice_state == 'INVOICED'
        conn.execute("""
            UPDATE stm_statement
            SET invoice_number=?, invoice_date=?, invoice_total=?, invoice_raw_text=?,
                invoice_status=?, overall_status=?
            WHERE id=?
        """, (
            data.get('invoice_number', ''),
            invoice_date,
            invoice_sum,
            data.get('raw_text', ''),
            'PASS' if passed else 'FAIL',
            'COMPLETED' if passed else 'INVOICE_FAILED',
            stmt_id,
        ))
        row = conn.execute("SELECT * FROM stm_statement WHERE id=?", (stmt_id,)).fetchone()
    return jsonify({"summary": _enrich_history_rows([row])[0]})


@app.route("/api/delivery/reconcile", methods=["POST"])
def delivery_reconcile():
    """运行快递对账，返回结果 Excel。"""
    if not _can_use_delivery():
        return jsonify({"error": "无权限访问快递对账"}), 403
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    uploaded = request.files.get('statement')
    if uploaded and uploaded.filename:
        with tempfile.TemporaryDirectory(prefix='lwfp_delivery_single_') as tmp:
            work_dir = Path(tmp)
            original_filename = _display_filename(uploaded.filename)
            statement_path = work_dir / _storage_filename(uploaded.filename)
            uploaded.save(statement_path)
            file_hash = _file_sha256(statement_path)
            try:
                batches = _read_delivery_template(statement_path)
            except ValueError as exc:
                return jsonify({"error": str(exc)}), 400
            if not batches:
                return jsonify({"error": "未识别到快递单号，请按快递对账模板填写"}), 400
            delivery_rows = _load_table_rows('delivery')
            return_rows = _load_table_rows('returnandexchangestop')
            summary = [
                ['指标', '数值'],
                ['上传文件', original_filename],
                ['Sheet批次数', len(batches)],
            ]
            batch_details = []
            history_before_total = 0
            months = []
            result_sheets = []
            output_path = DELIVERY_OUTPUT_DIR / f"快递对账结果_{timestamp}.xlsx"
            for meta, statement_records in batches:
                month = meta.get('对账月份') or _previous_month()
                unique_key = meta.get('唯一标识') or f"{meta.get('快递公司') or '快递'}-{month}-{meta.get('批次名称') or '月度账单'}"
                meta['对账月份'] = month
                meta['唯一标识'] = unique_key
                history_before = _delivery_history_count(unique_key=unique_key)
                history_before_total += history_before
                months.append(month)
                batch_summary, matched, only_statement, only_system = _delivery_result_rows(
                    statement_records,
                    delivery_rows,
                    return_rows,
                    month,
                    'transaction_time',
                )
                company = meta.get('快递公司') or '快递'
                company_path = DELIVERY_OUTPUT_DIR / f"快递对账结果_{company}_{month}_{timestamp}.xlsx"
                delivery_compare.write_xlsx(company_path, [
                    ('汇总', batch_summary),
                    ('匹配成功', delivery_compare.rows_to_table(matched)),
                    ('仅快递账单', delivery_compare.rows_to_table(only_statement)),
                    ('仅系统', delivery_compare.rows_to_table(only_system)),
                ])
                result_sheets.extend([
                    (_excel_sheet_name(f"{company}_汇总"), batch_summary),
                    (_excel_sheet_name(f"{company}_匹配成功"), delivery_compare.rows_to_table(matched)),
                    (_excel_sheet_name(f"{company}_仅快递账单"), delivery_compare.rows_to_table(only_statement)),
                    (_excel_sheet_name(f"{company}_仅系统"), delivery_compare.rows_to_table(only_system)),
                ])
                batch_details.append({
                    'company': company,
                    'month': month,
                    'unique_key': unique_key,
                    'history_before': history_before,
                    'statement_count': len(statement_records),
                    'matched_count': len(matched),
                    'only_statement_count': len(only_statement),
                    'only_system_count': len(only_system),
                    'result_path': str(company_path),
                })
                summary.extend([
                    ['', ''],
                    ['唯一标识', unique_key],
                    ['快递公司', meta.get('快递公司', '')],
                    ['对账月份', month],
                    ['确认时间', meta.get('填写日期', '')],
                    ['历史对账次数', history_before],
                    ['快递账单唯一单号数', len({r['_tracking'] for r in statement_records})],
                    ['匹配成功记录数', len(matched)],
                    ['仅快递账单记录数', len(only_statement)],
                    ['仅系统记录数', len(only_system)],
                ])
            delivery_compare.write_xlsx(output_path, [
                ('汇总', summary),
                *result_sheets,
            ])
            manifest_path = output_path.with_suffix('.json')
            manifest_path.write_text(json.dumps({'batches': batch_details}, ensure_ascii=False, indent=2), encoding='utf-8')
            aggregate_meta = {
                '唯一标识': f"快递对账-{timestamp}",
                '快递公司': '/'.join(item['company'] for item in batch_details),
                '对账月份': ','.join(sorted(set(months))) if months else _previous_month(),
                '填写日期': datetime.now().strftime('%Y-%m-%d'),
            }
            run_id = _record_delivery_run(aggregate_meta, original_filename, file_hash, {
                'statement_count': sum(item['statement_count'] for item in batch_details),
                'matched_count': sum(item['matched_count'] for item in batch_details),
                'only_statement_count': sum(item['only_statement_count'] for item in batch_details),
                'only_system_count': sum(item['only_system_count'] for item in batch_details),
            }, output_path)
        response = send_file(
            output_path,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=output_path.name,
        )
        response.headers['X-Delivery-History-Before'] = str(history_before_total)
        response.headers['X-Delivery-History-After'] = str(history_before_total + 1)
        response.headers['X-Delivery-Run-Id'] = str(run_id)
        return response

    # Backward-compatible full comparison path. The page no longer uses this.
    with tempfile.TemporaryDirectory(prefix='lwfp_delivery_') as tmp:
        work_dir = Path(tmp)
        delivery_csv = work_dir / 'delivery.csv'
        returns_csv = work_dir / 'returnandexchangestop.csv'
        _export_table_csv('delivery', delivery_csv)
        _export_table_csv('returnandexchangestop', returns_csv)

        _save_delivery_upload(work_dir, 'jitu', '26极兔.xlsx')
        _save_delivery_upload(work_dir, 'zhongtong', '26淘品中通.xlsx')
        _save_delivery_upload(work_dir, 'shunfeng', '26顺丰.xlsx')
        _save_delivery_upload(work_dir, 'supplement_feb', '2月发货在线汇总表.xlsx')
        _save_delivery_upload(work_dir, 'supplement_may', '5月骊威发货汇总表.xlsx')

        output_path = DELIVERY_OUTPUT_DIR / f"快递对账结果_{timestamp}.xlsx"
        args = [
            '--base-dir', str(work_dir),
            '--delivery', str(delivery_csv),
            '--returns', str(returns_csv),
            '--jitu', str(work_dir / '26极兔.xlsx'),
            '--zhongtong', str(work_dir / '26淘品中通.xlsx'),
            '--shunfeng', str(work_dir / '26顺丰.xlsx'),
            '--supplement-feb', str(work_dir / '2月发货在线汇总表.xlsx'),
            '--supplement-may', str(work_dir / '5月骊威发货汇总表.xlsx'),
            '--system-date-column', request.form.get('system_date_column') or 'created_at',
            '--output', str(output_path),
        ]
        code = delivery_compare.main(args)
        if code != 0 or not output_path.exists():
            return jsonify({"error": "快递对账失败"}), 500
    return send_file(
        output_path,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=output_path.name,
    )


@app.route("/api/delivery/template", methods=["GET"])
def delivery_template():
    if not _can_use_delivery():
        return jsonify({"error": "无权限访问快递对账"}), 403
    from openpyxl import Workbook
    wb = Workbook()
    default_month = _previous_month()
    examples = [
        ('顺丰干配', 'SF0000000000000'),
        ('极兔快递', 'JT0000000000000'),
        ('中通', '79000000000000'),
    ]
    for idx, (sheet_name, tracking_no) in enumerate(examples):
        ws = wb.active if idx == 0 else wb.create_sheet()
        ws.title = sheet_name
        ws.append(['快递单号*', '账单日期*', '运费', '重量', '备注'])
        ws.append([tracking_no, f'{default_month}-01', '0.00', '', '示例行，填写时删除'])
        for width, col in [(20, 'A'), (14, 'B'), (14, 'C'), (14, 'D'), (48, 'E')]:
            ws.column_dimensions[col].width = width
        for cell in ws[1]:
            cell.font = cell.font.copy(bold=True)
    data = io.BytesIO()
    wb.save(data)
    data.seek(0)
    return send_file(
        data,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='快递对账模板.xlsx',
    )


@app.route("/api/delivery/history-count", methods=["GET"])
def delivery_history_count():
    if not _can_use_delivery():
        return jsonify({"error": "无权限访问快递对账"}), 403
    unique_key = request.args.get('unique_key', '').strip()
    month = request.args.get('month', '')
    filename = _display_filename(request.args.get('filename', ''))
    return jsonify({"count": _delivery_history_count(unique_key=unique_key, month=month, filename=filename)})


@app.route("/api/delivery/history", methods=["GET"])
def delivery_history():
    if not _can_use_delivery():
        return jsonify({"error": "无权限访问快递对账"}), 403
    with get_db() as conn:
        rows = conn.execute("""
            SELECT id, unique_key, courier_company, fill_date,
                   statement_month, original_filename, statement_count,
                   matched_count, only_statement_count, only_system_count,
                   created_by, created_at
            FROM delivery_reconciliation_run
            ORDER BY id DESC
            LIMIT 50
        """).fetchall()
    data = []
    for row in rows:
        item = dict(row)
        item['download_url'] = f"/api/delivery/history/{item['id']}/download"
        data.append(item)
    return jsonify({"rows": data})


@app.route("/api/delivery/history/<int:run_id>/download", methods=["GET"])
def delivery_history_download(run_id):
    if not _can_use_delivery():
        return jsonify({"error": "无权限访问快递对账"}), 403
    with get_db() as conn:
        row = conn.execute(
            "SELECT result_path FROM delivery_reconciliation_run WHERE id=?", (run_id,)
        ).fetchone()
    if not row or not row['result_path'] or not Path(row['result_path']).exists():
        return jsonify({"error": "结果文件不存在"}), 404
    company = request.args.get('company', '').strip()
    if company:
        manifest_path = Path(row['result_path']).with_suffix('.json')
        if not manifest_path.exists():
            return jsonify({"error": "详情文件不存在"}), 404
        manifest = json.loads(manifest_path.read_text(encoding='utf-8'))
        detail = next((item for item in manifest.get('batches', []) if item.get('company') == company), None)
        if not detail or not detail.get('result_path') or not Path(detail['result_path']).exists():
            return jsonify({"error": "公司结果文件不存在"}), 404
        path = Path(detail['result_path'])
        return send_file(
            path,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=path.name,
        )
    path = Path(row['result_path'])
    return send_file(
        path,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=path.name,
    )


@app.route("/api/delivery/history/<int:run_id>/detail", methods=["GET"])
def delivery_history_detail(run_id):
    if not _can_use_delivery():
        return jsonify({"error": "无权限访问快递对账"}), 403
    with get_db() as conn:
        row = conn.execute(
            "SELECT id, result_path FROM delivery_reconciliation_run WHERE id=?", (run_id,)
        ).fetchone()
    if not row or not row['result_path']:
        return jsonify({"error": "历史记录不存在"}), 404
    manifest_path = Path(row['result_path']).with_suffix('.json')
    if not manifest_path.exists():
        return jsonify({"rows": []})
    manifest = json.loads(manifest_path.read_text(encoding='utf-8'))
    rows = []
    for item in manifest.get('batches', []):
        data = dict(item)
        data.pop('result_path', None)
        data['download_url'] = f"/api/delivery/history/{run_id}/download?company={data.get('company', '')}"
        rows.append(data)
    return jsonify({"rows": rows})


# ================================================================
#  对账核销/匹配 API
# ================================================================

@app.route("/api/match", methods=["POST"])
def run_match():
    """
    执行发票-对账单自动匹配
    REQ-043: 调用 matching_engine.match_invoice_statement
    """
    data = request.get_json()
    if not data:
        return jsonify({"code": 400, "message": "请提供invoice_id和statement_id"}), 400

    invoice_id = data.get('invoice_id')
    statement_id = data.get('statement_id')

    if not invoice_id or not statement_id:
        return jsonify({"code": 400, "message": "invoice_id 和 statement_id 必填"}), 400

    try:
        with get_db() as conn:
            result = match_invoice_statement(conn, invoice_id, statement_id)

        return jsonify({"code": 0, "message": "匹配完成", "data": result})

    except ValueError as e:
        return jsonify({"code": 404, "message": str(e)}), 404
    except Exception as e:
        logger.exception("匹配执行异常")
        return jsonify({"code": 500, "message": f"匹配异常: {str(e)}"}), 500


@app.route("/api/match/batch", methods=["POST"])
def run_batch_match():
    """批量匹配：对所有未匹配的发票和对账单执行匹配"""
    try:
        with get_db() as conn:
            invoices = conn.execute(
                "SELECT id FROM inv_invoice WHERE status='normal'"
            ).fetchall()
            statements = conn.execute(
                "SELECT id FROM stm_statement WHERE status IN ('draft','confirmed')"
            ).fetchall()

            results = []
            for inv in invoices:
                for stmt in statements:
                    try:
                        r = match_invoice_statement(conn, inv['id'], stmt['id'])
                        results.append(r)
                    except ValueError:
                        continue

        total_auto = sum(r['auto_matched'] for r in results)
        total_suggest = sum(r['suggested'] for r in results)
        total_unmatched = sum(r['unmatched'] for r in results)

        return jsonify({
            "code": 0,
            "message": f"批量匹配完成: {len(results)}组",
            "data": {
                "total_pairs": len(results),
                "auto_matched": total_auto,
                "suggested": total_suggest,
                "unmatched": total_unmatched,
                "details": results
            }
        })

    except Exception as e:
        logger.exception("批量匹配异常")
        return jsonify({"code": 500, "message": f"批量匹配异常: {str(e)}"}), 500


@app.route("/api/match/results", methods=["GET"])
def get_match_results():
    """查询匹配结果"""
    page = request.args.get('page', 1, type=int)
    size = request.args.get('size', 20, type=int)
    level = request.args.get('level', '')

    where_clauses = []
    params = []

    if level:
        where_clauses.append("r.match_level = ?")
        params.append(level)

    where_sql = f"WHERE {' AND '.join(where_clauses)}" if where_clauses else ""

    with get_db() as conn:
        data, total, page = _paginate_query(
            conn,
            f"""SELECT r.*, 
                       i.invoice_number, i.invoice_date,
                       s.statement_period, s.customer_name
                FROM rcn_reconciliation r
                LEFT JOIN inv_invoice i ON r.invoice_id = i.id
                LEFT JOIN stm_statement s ON r.statement_id = s.id
                {where_sql}
                ORDER BY r.match_score DESC""",
            f"SELECT count(*) FROM rcn_reconciliation r {where_sql}",
            page, size, params
        )

    return jsonify({"code": 0, "data": data, "total": total, "page": page})


@app.route("/api/match/<int:match_id>/confirm", methods=["POST"])
def confirm_match(match_id):
    """人工确认匹配结果"""
    with get_db() as conn:
        row = conn.execute(
            "SELECT * FROM rcn_reconciliation WHERE id=?", (match_id,)
        ).fetchone()
        if not row:
            return jsonify({"code": 404, "message": "匹配记录不存在"}), 404

        conn.execute(
            "UPDATE rcn_reconciliation SET is_confirmed=1, confirmed_by=? WHERE id=?",
            ('manual', match_id)
        )
        audit_log(conn, None, 'system', 'MATCH', 'reconciliation', match_id,
                  ip=_get_client_ip())

        # 记录反馈到反馈表（用于匹配引擎学习）
        if row['invoice_item_id'] and row['statement_item_id']:
            feedback_engine.record_feedback(
                invoice_item_id=row['invoice_item_id'],
                statement_item_id=row['statement_item_id'],
                feedback_type='confirm',
                original_score=row['match_score'],
                original_level=row['match_level'],
                feedback_reason='人工确认匹配',
                created_by=session.get('user_id')
            )

    return jsonify({"code": 0, "message": "确认成功"})


# ================================================================
#  回款 API
# ================================================================

@app.route("/api/payments", methods=["GET"])
def list_payments():
    """查询回款记录"""
    page = request.args.get('page', 1, type=int)
    size = request.args.get('size', 20, type=int)

    with get_db() as conn:
        data, total, page = _paginate_query(
            conn,
            """SELECT p.*, s.statement_period, s.customer_name
               FROM stm_payment p
               LEFT JOIN stm_statement s ON p.statement_id = s.id
               ORDER BY p.payment_date DESC""",
            "SELECT count(*) FROM stm_payment",
            page, size
        )

    return jsonify({"code": 0, "data": data, "total": total, "page": page})


@app.route("/api/payments", methods=["POST"])
def create_payment():
    """录入回款"""
    data = request.get_json()
    if not data:
        return jsonify({"code": 400, "message": "请求体为空"}), 400

    required = ['statement_id', 'payment_date', 'amount']
    for field in required:
        if not data.get(field):
            return jsonify({"code": 400, "message": f"缺少必填字段: {field}"}), 400

    with get_db() as conn:
        cursor = conn.execute("""
            INSERT INTO stm_payment (
                statement_id, invoice_id, payment_date, amount,
                payment_method, bill_number, bill_maturity, bank_ref_no, remark
            ) VALUES (?,?,?,?,?,?,?,?,?)
        """, (
            data['statement_id'],
            data.get('invoice_id'),
            data['payment_date'],
            data['amount'],
            data.get('payment_method', 'bank_transfer'),
            data.get('bill_number', ''),
            data.get('bill_maturity', ''),
            data.get('bank_ref_no', ''),
            data.get('remark', ''),
        ))

        audit_log(conn, None, 'system', 'CREATE', 'payment', cursor.lastrowid,
                  new_values=data, ip=_get_client_ip())

    return jsonify({"code": 0, "message": "回款录入成功", "data": {"id": cursor.lastrowid}})


# ================================================================
#  企业信息 API
# ================================================================

@app.route("/api/enterprises", methods=["GET"])
def list_enterprises():
    """查询企业列表"""
    page = request.args.get('page', 1, type=int)
    size = request.args.get('size', 20, type=int)
    etype = request.args.get('type', '')

    where_sql = ""
    params = []
    if etype:
        where_sql = "WHERE enterprise_type = ? OR enterprise_type = 'both'"
        params.append(etype)

    with get_db() as conn:
        data, total, page = _paginate_query(
            conn,
            f"SELECT * FROM sys_enterprise {where_sql} ORDER BY id DESC",
            f"SELECT count(*) FROM sys_enterprise {where_sql}",
            page, size, params
        )

    return jsonify({"code": 0, "data": data, "total": total, "page": page})


@app.route("/api/enterprises", methods=["POST"])
def create_enterprise():
    """新建企业"""
    data = request.get_json()
    if not data or not data.get('enterprise_name'):
        return jsonify({"code": 400, "message": "企业名称必填"}), 400

    with get_db() as conn:
        cursor = conn.execute("""
            INSERT INTO sys_enterprise (
                enterprise_name, tax_id, address, phone,
                bank_name, bank_account, seal_number,
                enterprise_type, contact_person
            ) VALUES (?,?,?,?,?,?,?,?,?)
        """, (
            data['enterprise_name'],
            data.get('tax_id', ''),
            data.get('address', ''),
            data.get('phone', ''),
            data.get('bank_name', ''),
            data.get('bank_account', ''),
            data.get('seal_number', ''),
            data.get('enterprise_type', 'both'),
            data.get('contact_person', ''),
        ))
    return jsonify({"code": 0, "message": "创建成功", "data": {"id": cursor.lastrowid}})


# ================================================================
#  物料管理 API
# ================================================================

@app.route("/api/materials", methods=["GET"])
def list_materials():
    """查询物料列表"""
    page = request.args.get('page', 1, type=int)
    size = request.args.get('size', 20, type=int)
    keyword = request.args.get('keyword', '')

    where_sql = ""
    params = []
    if keyword:
        where_sql = "WHERE material_name LIKE ? OR material_code LIKE ?"
        params = [f"%{keyword}%", f"%{keyword}%"]

    with get_db() as conn:
        data, total, page = _paginate_query(
            conn,
            f"SELECT * FROM sys_material {where_sql} ORDER BY id DESC",
            f"SELECT count(*) FROM sys_material {where_sql}",
            page, size, params
        )

    return jsonify({"code": 0, "data": data, "total": total, "page": page})


@app.route("/api/materials", methods=["POST"])
def create_material():
    """新建物料"""
    data = request.get_json()
    if not data or not data.get('material_code') or not data.get('material_name'):
        return jsonify({"code": 400, "message": "物料编码和名称必填"}), 400

    with get_db() as conn:
        try:
            cursor = conn.execute("""
                INSERT INTO sys_material (
                    material_code, material_name, category,
                    specification, unit, tax_rate
                ) VALUES (?,?,?,?,?,?)
            """, (
                data['material_code'],
                data['material_name'],
                data.get('category', ''),
                data.get('specification', ''),
                data.get('unit', 'PCS'),
                data.get('tax_rate', 13.0),
            ))
        except Exception as e:
            return jsonify({"code": 409, "message": f"物料编码已存在: {str(e)}"}), 409

    return jsonify({"code": 0, "message": "创建成功", "data": {"id": cursor.lastrowid}})


# ================================================================
#  料号映射 API
# ================================================================

@app.route("/api/material-mappings", methods=["GET"])
def list_material_mappings():
    """查询料号映射"""
    page = request.args.get('page', 1, type=int)
    size = request.args.get('size', 50, type=int)

    with get_db() as conn:
        data, total, page = _paginate_query(
            conn,
            "SELECT m.*, e.enterprise_name FROM sys_material_mapping m LEFT JOIN sys_enterprise e ON m.enterprise_id = e.id ORDER BY m.id DESC",
            "SELECT count(*) FROM sys_material_mapping",
            page, size
        )

    return jsonify({"code": 0, "data": data, "total": total, "page": page})


@app.route("/api/material-mappings", methods=["POST"])
def create_material_mapping():
    """新建料号映射"""
    data = request.get_json()
    if not data:
        return jsonify({"code": 400, "message": "请求体为空"}), 400

    with get_db() as conn:
        cursor = conn.execute("""
            INSERT INTO sys_material_mapping (
                enterprise_id, customer_material_code, supplier_material_code,
                customer_name, supplier_name, spec, unit, tax_rate
            ) VALUES (?,?,?,?,?,?,?,?)
        """, (
            data.get('enterprise_id'),
            data['customer_material_code'],
            data['supplier_material_code'],
            data.get('customer_name', ''),
            data.get('supplier_name', ''),
            data.get('spec', ''),
            data.get('unit', 'PCS'),
            data.get('tax_rate', '13%'),
        ))

    return jsonify({"code": 0, "message": "映射创建成功", "data": {"id": cursor.lastrowid}})


# ================================================================
#  Dashboard API
# ================================================================

@app.route("/api/dashboard", methods=["GET"])
def dashboard_lwfp():
    """LWFP 首页统计接口，兼容原轻量框架。"""
    with get_db() as conn:
        total = conn.execute("SELECT count(*) FROM stm_statement").fetchone()[0]
        total_amount = conn.execute(
            "SELECT COALESCE(SUM(current_payment),0) FROM stm_statement"
        ).fetchone()[0]
        status_rows = conn.execute("""
            SELECT COALESCE(NULLIF(overall_status, ''), status) AS status, count(*) AS cnt
            FROM stm_statement
            GROUP BY COALESCE(NULLIF(overall_status, ''), status)
        """).fetchall()
        anomaly_open = conn.execute(
            "SELECT count(*) FROM sys_anomaly WHERE status='open'"
        ).fetchone()[0]
        trend_rows = conn.execute("""
            SELECT substr(created_at, 1, 10) AS day,
                   count(*) AS cnt,
                   COALESCE(SUM(current_payment),0) AS amount
            FROM stm_statement
            GROUP BY substr(created_at, 1, 10)
            ORDER BY day
            LIMIT 12
        """).fetchall()

    counts = {
        "WAITING_INVOICE": 0,
        "ERP_FAILED": anomaly_open,
        "INVOICE_FAILED": 0,
        "COMPLETED": 0,
        "OTHER": 0,
    }
    for row in status_rows:
        status = row["status"]
        cnt = row["cnt"]
        if status == "ERP_FAILED":
            counts["ERP_FAILED"] += cnt
        elif status == "INVOICE_FAILED":
            counts["INVOICE_FAILED"] += cnt
        elif status in ("COMPLETED", "confirmed", "archived"):
            counts["COMPLETED"] += cnt
        elif status in ("WAITING_INVOICE", "draft", "pending_review", "pending_customer"):
            counts["WAITING_INVOICE"] += cnt
        else:
            counts["OTHER"] += cnt

    trend = [
        {
            "date": row["day"] or "",
            "count": row["cnt"],
            "amount": round(row["amount"] or 0, 2),
        }
        for row in trend_rows
    ]
    pending = counts["WAITING_INVOICE"] + counts["INVOICE_FAILED"]
    abnormal = counts["ERP_FAILED"] + counts["INVOICE_FAILED"]
    return jsonify({
        "total_amount": f"{round(total_amount or 0, 2):.2f}",
        "total": total,
        "pending": pending,
        "abnormal": abnormal,
        "status_counts": counts,
        "trend": trend,
    })


@app.route("/api/dashboard/summary", methods=["GET"])
def dashboard_summary():
    """首页仪表盘汇总数据"""
    with get_db() as conn:
        invoice_count = conn.execute("SELECT count(*) FROM inv_invoice").fetchone()[0]
        invoice_total = conn.execute(
            "SELECT COALESCE(SUM(total_amount_incl),0) FROM inv_invoice"
        ).fetchone()[0]

        stmt_count = conn.execute("SELECT count(*) FROM stm_statement").fetchone()[0]
        payment_total = conn.execute(
            "SELECT COALESCE(SUM(current_payment),0) FROM stm_statement"
        ).fetchone()[0]

        # 匹配统计
        match_full = conn.execute(
            "SELECT count(*) FROM rcn_reconciliation WHERE match_level='full'"
        ).fetchone()[0]
        match_partial = conn.execute(
            "SELECT count(*) FROM rcn_reconciliation WHERE match_level='partial'"
        ).fetchone()[0]
        match_unmatched = conn.execute(
            "SELECT count(*) FROM rcn_reconciliation WHERE match_level='unmatched'"
        ).fetchone()[0]

        # 异常统计
        anomaly_open = conn.execute(
            "SELECT count(*) FROM sys_anomaly WHERE status='open'"
        ).fetchone()[0]

        # 企业数
        enterprise_count = conn.execute("SELECT count(*) FROM sys_enterprise").fetchone()[0]

        # 物料数
        material_count = conn.execute("SELECT count(*) FROM sys_material WHERE is_active=1").fetchone()[0]

    return jsonify({"code": 0, "data": {
        "invoice_count": invoice_count,
        "invoice_total": round(invoice_total, 2),
        "statement_count": stmt_count,
        "payment_total": round(payment_total, 2),
        "unpaid": round(invoice_total - payment_total, 2),
        "match_stats": {
            "auto_matched": match_full,
            "suggested": match_partial,
            "unmatched": match_unmatched,
        },
        "anomaly_open": anomaly_open,
        "enterprise_count": enterprise_count,
        "material_count": material_count,
    }})


@app.route("/api/dashboard/anomalies", methods=["GET"])
def dashboard_anomalies():
    """查询未处理异常"""
    with get_db() as conn:
        rows = conn.execute(
            "SELECT * FROM sys_anomaly WHERE status='open' ORDER BY created_at DESC LIMIT 20"
        ).fetchall()
    return jsonify({"code": 0, "data": rows_to_list(rows)})


# ================================================================
#  导出 API
# ================================================================

@app.route("/api/export/invoices", methods=["GET"])
def export_invoices():
    """导出发票 — 支持 CSV 和 Excel"""
    fmt = request.args.get('format', 'xlsx')

    with get_db() as conn:
        rows = conn.execute("SELECT * FROM inv_invoice ORDER BY id DESC").fetchall()
    invoices = rows_to_list(rows)

    if fmt == 'csv':
        csv_str = export_invoices_csv(invoices)
        return send_file(
            io.BytesIO(csv_str.encode('utf-8-sig')),
            mimetype='text/csv',
            as_attachment=True,
            download_name='invoices.csv'
        )
    else:
        wb = export_invoices_excel(invoices)
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return send_file(
            buf,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='invoices.xlsx'
        )


@app.route("/api/export/statements", methods=["GET"])
def export_statements():
    """导出对账单"""
    fmt = request.args.get('format', 'xlsx')

    with get_db() as conn:
        rows = conn.execute("SELECT * FROM stm_statement ORDER BY id DESC").fetchall()
    statements = rows_to_list(rows)

    if fmt == 'csv':
        csv_str = export_statements_csv(statements)
        return send_file(
            io.BytesIO(csv_str.encode('utf-8-sig')),
            mimetype='text/csv',
            as_attachment=True,
            download_name='statements.csv'
        )
    else:
        wb = export_statements_excel(statements)
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return send_file(
            buf,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='statements.xlsx'
        )


@app.route("/api/export/match-results", methods=["GET"])
def export_match_results():
    """导出匹配结果"""
    fmt = request.args.get('format', 'csv')

    with get_db() as conn:
        rows = conn.execute("""
            SELECT r.*, i.invoice_number, i.total_amount_incl AS invoice_amount,
                   s.statement_period, s.closing_balance AS statement_amount
            FROM rcn_reconciliation r
            LEFT JOIN inv_invoice i ON r.invoice_id = i.id
            LEFT JOIN stm_statement s ON r.statement_id = s.id
            ORDER BY r.match_score DESC
        """).fetchall()
    results = rows_to_list(rows)

    csv_str = export_match_results_csv(results)
    return send_file(
        io.BytesIO(csv_str.encode('utf-8-sig')),
        mimetype='text/csv',
        as_attachment=True,
        download_name='match_results.csv'
    )


# ================================================================
#  系统配置 API
# ================================================================

@app.route("/api/config", methods=["GET"])
def get_config():
    """获取系统配置"""
    with get_db() as conn:
        rows = conn.execute("SELECT * FROM sys_config ORDER BY id").fetchall()
    return jsonify({"code": 0, "data": rows_to_list(rows)})


@app.route("/api/config/<key>", methods=["PUT"])
def update_config(key):
    """更新系统配置"""
    data = request.get_json()
    value = data.get('value', '')

    with get_db() as conn:
        existing = conn.execute(
            "SELECT id FROM sys_config WHERE config_key=?", (key,)
        ).fetchone()
        if existing:
            conn.execute(
                "UPDATE sys_config SET config_value=?, updated_at=datetime('now','localtime') WHERE config_key=?",
                (value, key)
            )
        else:
            conn.execute(
                "INSERT INTO sys_config (config_key, config_value) VALUES (?,?)",
                (key, value)
            )
        audit_log(conn, None, 'system', 'UPDATE', 'config', None,
                  new_values={key: value}, ip=_get_client_ip())

    return jsonify({"code": 0, "message": "配置更新成功"})


# ================================================================
#  审计日志 API
# ================================================================

@app.route("/api/audit-logs", methods=["GET"])
def get_audit_logs():
    """查询审计日志"""
    page = request.args.get('page', 1, type=int)
    size = request.args.get('size', 50, type=int)
    action = request.args.get('action', '')

    where_sql = ""
    params = []
    if action:
        where_sql = "WHERE action = ?"
        params.append(action)

    with get_db() as conn:
        data, total, page = _paginate_query(
            conn,
            f"SELECT * FROM sys_audit_log {where_sql} ORDER BY id DESC",
            f"SELECT count(*) FROM sys_audit_log {where_sql}",
            page, size, params
        )

    return jsonify({"code": 0, "data": data, "total": total, "page": page})


# ================================================================
#  匹配反馈闭环
# ================================================================

@app.route("/api/feedback", methods=["POST"])
def submit_feedback():
    """提交匹配反馈（确认/拒绝/手动关联）"""
    data = request.get_json()
    if not data:
        return jsonify({"code": 400, "message": "请提供反馈数据"}), 400

    invoice_item_id = data.get('invoice_item_id')
    statement_item_id = data.get('statement_item_id')
    reconciliation_id = data.get('reconciliation_id')
    feedback_type = data.get('feedback_type')  # confirm / reject / manual_link

    if feedback_type not in ('confirm', 'reject', 'manual_link'):
        return jsonify({"code": 400, "message": "feedback_type 必须是 confirm/reject/manual_link"}), 400

    original_score = data.get('original_score')
    original_level = data.get('original_level')
    if reconciliation_id and (not invoice_item_id or not statement_item_id):
        with get_db() as conn:
            match = conn.execute(
                "SELECT invoice_item_id, statement_item_id, match_score, match_level FROM rcn_reconciliation WHERE id=?",
                (reconciliation_id,)
            ).fetchone()
        if not match:
            return jsonify({"code": 404, "message": "匹配记录不存在"}), 404
        invoice_item_id = invoice_item_id or match['invoice_item_id']
        statement_item_id = statement_item_id or match['statement_item_id']
        original_score = original_score if original_score is not None else match['match_score']
        original_level = original_level or match['match_level']

    if feedback_type == 'manual_link' and (not invoice_item_id or not statement_item_id):
        return jsonify({"code": 400, "message": "手动关联需提供 invoice_item_id 和 statement_item_id"}), 400

    user_id = session.get('user_id')
    result = feedback_engine.record_feedback(
        invoice_item_id=invoice_item_id,
        statement_item_id=statement_item_id,
        feedback_type=feedback_type,
        original_score=original_score,
        original_level=original_level,
        feedback_reason=data.get('feedback_reason'),
        created_by=user_id
    )

    with get_db() as conn:
        audit_log(conn, user_id, session.get('username'), 'CREATE', 'match_feedback',
                  result.get('feedback_id'), new_values=data, ip=_get_client_ip())

    return jsonify({"code": 0, "message": result['message'], "data": result})


@app.route("/api/feedback", methods=["GET"])
def get_feedback():
    """查询反馈历史"""
    supplier_code = request.args.get('supplier_code', '')
    limit = request.args.get('limit', 50, type=int)

    data = feedback_engine.get_feedback_history(
        supplier_code=supplier_code if supplier_code else None,
        limit=min(limit, 200)
    )
    return jsonify({"code": 0, "data": data, "total": len(data)})


@app.route("/api/feedback/stats", methods=["GET"])
def get_feedback_stats():
    """查询反馈统计"""
    supplier_code = request.args.get('supplier_code', '')
    stats = feedback_engine.get_feedback_statistics(
        supplier_code=supplier_code if supplier_code else None
    )
    return jsonify({"code": 0, "data": stats})


def _days_since(value):
    """Return whole days elapsed from a DB timestamp/date string to now."""
    if not value:
        return 0
    text = str(value).strip()
    for fmt, size in (("%Y-%m-%d %H:%M:%S", 19), ("%Y-%m-%dT%H:%M:%S", 19), ("%Y-%m-%d", 10)):
        try:
            return max((datetime.now() - datetime.strptime(text[:size], fmt)).days, 0)
        except ValueError:
            continue
    return 0


# ================================================================
#  超时预警
# ================================================================

@app.route("/api/warnings", methods=["GET"])
def get_warnings():
    """
    获取超时预警列表

    检测以下情况:
    1. 对账单上传超过 N 天未匹配
    2. 匹配结果超过 N 天未确认（partial/unmatched）
    3. 回款记录超过 N 天未核销
    """
    days_threshold = request.args.get('days', 7, type=int)
    severity = request.args.get('severity', '')  # info/warning/error/critical

    warnings = []

    try:
        with get_db() as conn:
            # 1. 对账单上传超过 N 天未匹配
            overdue_statements = conn.execute(
                """SELECT s.id, s.customer_name, s.supplier_name, s.supplier_code,
                          s.statement_period, s.total_invoice_amount, s.created_at,
                          (SELECT COUNT(*) FROM rcn_reconciliation r
                           WHERE r.statement_id = s.id AND r.match_level != 'unmatched') as matched_count,
                          (SELECT COUNT(*) FROM rcn_reconciliation r
                           WHERE r.statement_id = s.id) as total_match_count
                   FROM stm_statement s
                   WHERE s.status IN ('draft', 'pending_review')
                   ORDER BY s.created_at ASC""",
            ).fetchall()

            for row in overdue_statements:
                days = _days_since(row['created_at'])
                if days <= days_threshold:
                    continue
                sev = 'critical' if days > 30 else 'error' if days > 14 else 'warning'
                if severity and sev != severity:
                    continue
                # 优先使用 supplier_name，为空时使用 customer_name
                display_name = row['supplier_name'] or row['customer_name'] or '未知供应商'
                description = f"供应商 {display_name} 的对账单（{row['statement_period']}）已上传 {days} 天，仍有未匹配项"
                warnings.append({
                    "type": "statement_unmatched",
                    "warning_type": "overdue_statement",
                    "severity": sev,
                    "title": f"对账单超时未匹配: {display_name}",
                    "message": description,
                    "description": description,
                    "ref_type": "statement",
                    "ref_id": row['id'],
                    "related_id": row['id'],
                    "days": days,
                    "amount": row['total_invoice_amount'],
                    "created_at": row['created_at']
                })

            # 2. 匹配结果超过 N 天未确认（partial 级别）
            unconfirmed_matches = conn.execute(
                """SELECT r.id, r.match_score, r.match_level, r.difference_amount,
                          r.created_at,
                          i.invoice_number, s.customer_name, s.supplier_name, s.supplier_code
                   FROM rcn_reconciliation r
                   LEFT JOIN inv_invoice i ON r.invoice_id = i.id
                   LEFT JOIN stm_statement s ON r.statement_id = s.id
                   WHERE r.is_confirmed = 0
                     AND r.match_level IN ('partial', 'unmatched')
                   ORDER BY r.created_at ASC""",
            ).fetchall()

            for row in unconfirmed_matches:
                days = _days_since(row['created_at'])
                if days <= days_threshold:
                    continue
                sev = 'error' if days > 14 else 'warning'
                if severity and sev != severity:
                    continue
                description = f"发票 {row['invoice_number'] or '-'} 的匹配结果（得分 {row['match_score']}）已等待 {days} 天未确认"
                warnings.append({
                    "type": "match_unconfirmed",
                    "warning_type": "unconfirmed_match",
                    "severity": sev,
                    "title": f"匹配结果待确认: {row['supplier_name'] or '未知'}",
                    "message": description,
                    "description": description,
                    "ref_type": "match",
                    "ref_id": row['id'],
                    "related_id": row['id'],
                    "days": days,
                    "match_score": row['match_score'],
                    "created_at": row['created_at']
                })

            # 3. 已确认对账单但超过 N 天无回款
            unpaid_confirmed = conn.execute(
                """SELECT s.id, s.customer_name, s.supplier_name, s.supplier_code,
                          s.statement_period, s.total_invoice_amount, s.confirmed_at,
                          COALESCE((SELECT SUM(p.amount) FROM stm_payment p
                           WHERE p.statement_id = s.id), 0) as paid_amount
                   FROM stm_statement s
                   WHERE s.status = 'confirmed'
                     AND COALESCE((SELECT SUM(p.amount) FROM stm_payment p
                           WHERE p.statement_id = s.id), 0) < s.total_invoice_amount
                   ORDER BY s.confirmed_at ASC""",
            ).fetchall()

            for row in unpaid_confirmed:
                days = _days_since(row['confirmed_at'])
                if days <= days_threshold:
                    continue
                remaining = row['total_invoice_amount'] - row['paid_amount']
                sev = 'critical' if days > 60 else 'error' if days > 30 else 'warning'
                if severity and sev != severity:
                    continue
                description = f"供应商 {row['supplier_name']}（{row['statement_period']}）已确认 {days} 天，仍有 ¥{remaining:,.2f} 未回款"
                warnings.append({
                    "type": "reconciliation_unpaid",
                    "warning_type": "overdue_payment",
                    "severity": sev,
                    "title": f"回款超时: {row['supplier_name']}",
                    "message": description,
                    "description": description,
                    "ref_type": "statement",
                    "ref_id": row['id'],
                    "related_id": row['id'],
                    "days": days,
                    "total_amount": row['total_invoice_amount'],
                    "paid_amount": row['paid_amount'],
                    "remaining_amount": remaining,
                    "confirmed_at": row['confirmed_at']
                })

        # 按严重程度排序
        severity_order = {'critical': 0, 'error': 1, 'warning': 2, 'info': 3}
        warnings.sort(key=lambda w: (severity_order.get(w['severity'], 9), -w.get('days', 0)))

        return jsonify({
            "code": 0,
            "warnings": warnings,
            "data": warnings,
            "total": len(warnings),
            "summary": {
                "critical": sum(1 for w in warnings if w['severity'] == 'critical'),
                "error": sum(1 for w in warnings if w['severity'] == 'error'),
                "warning": sum(1 for w in warnings if w['severity'] == 'warning'),
            }
        })

    except Exception as e:
        logger.exception("获取预警信息异常")
        return jsonify({"code": 500, "message": f"获取预警异常: {str(e)}"}), 500


# ================================================================
#  供应商对账进度看板
# ================================================================

@app.route("/api/supplier-progress", methods=["GET"])
def get_supplier_progress():
    """
    按供应商维度展示对账进度

    返回每个供应商的:
    - 对账单数、已匹配数、待审核数、未匹配数
    - 完成率、平均耗时
    - 未回款金额
    """
    period = request.args.get('period', '')  # 可选：按月份筛选

    try:
        with get_db() as conn:
            where_sql = "WHERE statement_period = ?" if period else ""
            params = [period] if period else []
            statement_rows = conn.execute(
                f"""SELECT id, supplier_code, supplier_name, total_invoice_amount, created_at
                    FROM stm_statement {where_sql}""",
                params
            ).fetchall()
            statement_ids = [row['id'] for row in statement_rows]

            match_map = {}
            payment_map = {}
            if statement_ids:
                placeholders = ",".join(["?"] * len(statement_ids))
                match_rows = conn.execute(
                    f"""SELECT s.supplier_code, s.supplier_name,
                               COALESCE(SUM(CASE WHEN rc.match_level = 'full' THEN 1 ELSE 0 END), 0) as auto_matched,
                               COALESCE(SUM(CASE WHEN rc.match_level = 'partial' THEN 1 ELSE 0 END), 0) as suggested,
                               COALESCE(SUM(CASE WHEN rc.match_level = 'unmatched' THEN 1 ELSE 0 END), 0) as unmatched,
                               COALESCE(SUM(CASE WHEN rc.is_confirmed = 1 THEN 1 ELSE 0 END), 0) as confirmed,
                               COUNT(rc.id) as total_match_items
                        FROM stm_statement s
                        LEFT JOIN rcn_reconciliation rc ON rc.statement_id = s.id
                        WHERE s.id IN ({placeholders})
                        GROUP BY s.supplier_code, s.supplier_name""",
                    statement_ids
                ).fetchall()
                for row in match_rows:
                    key = row['supplier_code'] or row['supplier_name'] or ''
                    match_map[key] = dict(row)

                payment_rows = conn.execute(
                    f"""SELECT s.supplier_code, s.supplier_name, COALESCE(SUM(p.amount), 0) as paid_amount
                        FROM stm_statement s
                        LEFT JOIN stm_payment p ON p.statement_id = s.id
                        WHERE s.id IN ({placeholders})
                        GROUP BY s.supplier_code, s.supplier_name""",
                    statement_ids
                ).fetchall()
                for row in payment_rows:
                    key = row['supplier_code'] or row['supplier_name'] or ''
                    payment_map[key] = float(row['paid_amount'] or 0)

            grouped = {}
            for row in statement_rows:
                key = row['supplier_code'] or row['supplier_name'] or ''
                item = grouped.setdefault(key, {
                    "supplier_code": row['supplier_code'],
                    "supplier_name": row['supplier_name'],
                    "statement_count": 0,
                    "total_amount": 0.0,
                    "days_total": 0.0,
                })
                item["statement_count"] += 1
                item["total_amount"] += float(row['total_invoice_amount'] or 0)
                item["days_total"] += _days_since(row['created_at'])

            result = []
            for key, row in grouped.items():
                match = match_map.get(key, {})
                total_items = match.get('total_match_items') or 0
                matched_items = (match.get('auto_matched') or 0) + (match.get('confirmed') or 0)
                completion_rate = round(matched_items / total_items * 100, 1) if total_items > 0 else 0
                paid_amount = payment_map.get(key, 0)
                total_amount = row['total_amount']

                result.append({
                    "supplier_code": row['supplier_code'],
                    "supplier_name": row['supplier_name'],
                    "supplier": row['supplier_name'],
                    "statement_count": row['statement_count'],
                    "total_amount": total_amount,
                    "auto_matched": match.get('auto_matched') or 0,
                    "suggested": match.get('suggested') or 0,
                    "unmatched": match.get('unmatched') or 0,
                    "confirmed": match.get('confirmed') or 0,
                    "total_match_items": total_items,
                    "completion_rate": completion_rate,
                    "avg_days": round(row['days_total'] / row['statement_count'], 1) if row['statement_count'] else 0,
                    "paid_amount": paid_amount,
                    "remaining_amount": total_amount - paid_amount,
                })

            result.sort(key=lambda item: item["total_amount"], reverse=True)
            return jsonify({"code": 0, "data": result, "total": len(result)})

    except Exception as e:
        logger.exception("获取供应商进度异常")
        return jsonify({"code": 500, "message": f"获取进度异常: {str(e)}"}), 500


# ================================================================
#  批量操作
# ================================================================

@app.route("/api/match/batch-confirm", methods=["POST"])
def batch_confirm_matches():
    """批量确认匹配结果"""
    data = request.get_json()
    if not data:
        return jsonify({"code": 400, "message": "请提供匹配ID列表"}), 400

    match_ids = data.get('match_ids') or []
    statement_ids = data.get('statement_ids') or data.get('ids') or []
    if not match_ids and not statement_ids:
        return jsonify({"code": 400, "message": "match_ids 或 statement_ids 不能为空"}), 400

    user_id = session.get('user_id')
    username = session.get('username')

    try:
        with get_db() as conn:
            confirmed = 0
            if match_ids:
                for mid in match_ids:
                    cursor = conn.execute(
                        """UPDATE rcn_reconciliation
                           SET is_confirmed = 1, confirmed_by = ?
                           WHERE id = ? AND is_confirmed = 0""",
                        (username, mid)
                    )
                    confirmed += cursor.rowcount
            if statement_ids:
                placeholders = ",".join(["?"] * len(statement_ids))
                cursor = conn.execute(
                    f"""UPDATE rcn_reconciliation
                        SET is_confirmed = 1, confirmed_by = ?
                        WHERE statement_id IN ({placeholders}) AND is_confirmed = 0""",
                    [username] + statement_ids
                )
                confirmed += cursor.rowcount

            audit_log(conn, user_id, username, 'UPDATE', 'match', None,
                      new_values={
                          "batch_confirmed": confirmed,
                          "match_ids": match_ids,
                          "statement_ids": statement_ids,
                      },
                      ip=_get_client_ip())

        return jsonify({
            "code": 0,
            "message": f"已确认 {confirmed} 条匹配记录",
            "confirmed_count": confirmed,
            "data": {"confirmed_count": confirmed}
        })

    except Exception as e:
        logger.exception("批量确认异常")
        return jsonify({"code": 500, "message": f"批量确认异常: {str(e)}"}), 500


# ================================================================
#  健康检查
# ================================================================

@app.route("/api/health", methods=["GET"])
def health_check():
    """健康检查端点"""
    try:
        with get_db() as conn:
            conn.execute("SELECT 1")
        return jsonify({"code": 0, "status": "healthy", "timestamp": datetime.now().isoformat()})
    except Exception as e:
        return jsonify({"code": 500, "status": "unhealthy", "error": str(e)}), 500


# ================================================================
#  启动 — REQ-039: 端口统一 8090
# ================================================================

if __name__ == "__main__":
    port = int(os.getenv('PORT', 8090))
    logger.info("FP进销存财务系统启动 → port=%d", port)
    debug = os.getenv('FLASK_DEBUG', 'true').lower() == 'true'
    app.run(host="0.0.0.0", port=port, debug=debug)
