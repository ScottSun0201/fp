#!/usr/bin/env python3
"""
对账单PDF解析引擎
REQ-031: 解析月度对账单PDF -> 提取客户信息+明细行+四项资金
"""
import re
import tempfile
from pathlib import Path
from openpyxl import load_workbook
from config import AMOUNT_TOLERANCE
from qwen_ocr import recognize_image as recognize_image_with_qwen
from qwen_ocr import recognize_pdf as recognize_pdf_with_qwen


def _empty_statement_result() -> dict:
    return {
        'customer_name': '', 'customer_tax_id': '', 'supplier_name': '',
        'supplier_code': '', 'supplier_tax_id': '', 'statement_month': '',
        'statement_date': '', 'usage_remark': '', 'settlement_days': 30,
        'opening_balance': 0.0, 'current_payment': 0.0,
        'closing_balance': 0.0, 'delivered_unpaid': 0.0,
        'total_invoice_amount': 0.0, 'total_quantity': 0, 'items': [],
        'balance_check': True, 'raw_text': '', 'column_mapping': {}, 'errors': []
    }


def parse_statement_image(image_path: str) -> dict:
    """从对账单图片中提取全部字段。"""
    result = _empty_statement_result()
    try:
        data = recognize_image_with_qwen(image_path, 'statement')
        if not data:
            result['errors'].append('千问未返回结果，已切换本地OCR')
            return _merge_local_scan_result(result, _ocr_image_text(image_path))
        _merge_qwen_statement(result, data)
        _validate_item_roles(result)
        result['raw_text'] = data.get('raw_text', '')
        if not result['items']:
            result['errors'].append('未能解析对账单明细行')
            return result
        result['total_quantity'] = sum(float(item.get('quantity') or 0) for item in result['items'])
        result['total_invoice_amount'] = round(
            sum(float(item.get('amount_incl_tax') or 0) for item in result['items']), 2
        )
    except Exception as exc:
        result['errors'].append(f'千问图片识别失败，已切换本地OCR: {exc}')
        _merge_local_scan_result(result, _ocr_image_text(image_path))
    return result


def _merge_local_scan_result(result: dict, text: str) -> dict:
    result['raw_text'] = text
    result['supplier_name'] = result.get('supplier_name') or _supplier_from_local_text(text)
    month = re.search(r'(\d{4})[./年-](\d{1,2})(?:[./月-]\d{1,2})?', text)
    if month:
        result['statement_month'] = f"{month.group(1)}-{int(month.group(2)):02d}"
    result['items'] = _parse_ocr_column_items(text) or _parse_fixed_scan_items(text)
    result['total_quantity'] = sum(float(item.get('quantity') or 0) for item in result['items'])
    result['total_invoice_amount'] = round(
        sum(float(item.get('amount_incl_tax') or 0) for item in result['items']), 2
    )
    _validate_item_roles(result)
    if not result['items']:
        result['errors'].append('本地OCR未能解析对账单明细行')
    return result


def parse_statement_pdf(pdf_path: str, progress_callback=None) -> dict:
    """从对账单PDF中提取全部字段"""
    result = _empty_statement_result()

    try:
        import pdfplumber
    except Exception as e:
        result['errors'].append(f'PDF识别依赖不可用: {str(e)}')
        return result

    try:
        ocr_used = False
        qwen_data = {}
        with pdfplumber.open(pdf_path) as pdf:
            all_text = ''
            table_rows = []
            for page in pdf.pages:
                t = page.extract_text()
                if t:
                    all_text += t + '\n'
                for table in page.extract_tables() or []:
                    for row in table:
                        if row and any(cell not in (None, '') for cell in row):
                            table_rows.append(tuple(row))
            result['raw_text'] = all_text

            if not all_text.strip() and not table_rows:
                ocr_used = True
                try:
                    qwen_data = recognize_pdf_with_qwen(
                        pdf_path, 'statement', progress_callback=progress_callback
                    )
                except Exception as exc:
                    # 云端识别不可用时不能终止整个流程，本地 Vision 仍可提供
                    # 固定版式字段和可人工修订的明细。
                    result['errors'].append(f'千问识别失败，已切换本地OCR: {exc}')
                    qwen_data = {}
                if qwen_data:
                    _merge_qwen_statement(result, qwen_data)
                    all_text = qwen_data.get('raw_text', '')
                else:
                    all_text = _ocr_pdf_text(pdf_path)
                    result['raw_text'] = all_text
                if not all_text.strip():
                    result['errors'].append('PDF无可提取文本（可能是扫描件，需OCR）')
                    return result

            if table_rows:
                _merge_template_fields(result, table_rows)
                result['items'] = _parse_xlsx_items(table_rows)

            # 客户名称
            m = re.search(r'(?:客户名称|客户)[：:\s]*([^\n]+?(?:公司|集团)[^\n]*)', all_text)
            if m:
                result['customer_name'] = result['customer_name'] or _clean_party_name(m.group(1))

            # 供应商名称
            m = re.search(r'(?:供应商名称|供应商)[：:\s]*([^\n]+?(?:公司|集团)[^\n]*)', all_text)
            if m:
                result['supplier_name'] = result['supplier_name'] or _clean_party_name(m.group(1))
            if not result['supplier_name']:
                title_party = re.search(
                    r'与\s*([\u4e00-\u9fffA-Za-z0-9（）()·]{2,}?(?:有限责任公司|有限公司|公司))',
                    all_text,
                )
                if title_party and '骊威' not in title_party.group(1):
                    result['supplier_name'] = _clean_party_name(title_party.group(1))
            if not result['supplier_name']:
                companies = re.findall(
                    r'[\u4e00-\u9fffA-Za-z0-9（）()·]{2,}(?:有限责任公司|有限公司|公司)',
                    all_text,
                )
                result['supplier_name'] = next(
                    (_clean_party_name(name) for name in companies if '骊威' not in name),
                    '',
                )

            # 对账月份
            m = re.search(r'(?:对账月份|账期|月份)[：:\s]*(\d{4}[-./年]\s*\d{1,2})', all_text)
            if m:
                result['statement_month'] = result['statement_month'] or _format_month(m.group(1))
            m = re.search(r'(\d{4})\s*年\s*(\d{1,2})\s*月', all_text)
            if m:
                result['statement_month'] = result['statement_month'] or f"{m.group(1)}-{int(m.group(2)):02d}"
            m = re.search(r'(\d{4})[./-](\d{1,2})[./-]\d{1,2}', all_text)
            if m:
                result['statement_month'] = result['statement_month'] or f"{m.group(1)}-{int(m.group(2)):02d}"

            # 月结天数
            m = re.search(r'月结\s*(\d+)\s*天', all_text)
            if m:
                result['settlement_days'] = int(m.group(1))

            # 四项资金
            result['opening_balance'] = _extract_amount(all_text, r'期初[^\d]*?([\d,]+\.\d{2})')
            result['current_payment'] = _extract_amount(all_text, r'本期(?:付款|回款)[^\d]*?[¥￥]?\s*([\d,]+\.?\d*)')
            result['total_invoice_amount'] = _extract_amount(all_text, r'(?:本次|本期)(?:对帐|对账)?开票[^\d]*?[¥￥]?\s*([\d,]+\.?\d*)')
            result['delivered_unpaid'] = _extract_amount(all_text, r'已交货未回款[：:\s]*[¥￥]?\s*([\d,]+\.?\d*)')
            result['closing_balance'] = _extract_amount(all_text, r'期末[^\d]*?([\d,]+\.\d{2})')

            if result['closing_balance'] == 0:
                result['closing_balance'] = result['opening_balance'] + result['total_invoice_amount'] - result['current_payment']

            # 平衡校验
            expected = result['opening_balance'] + result['total_invoice_amount'] - result['current_payment']
            if abs(expected - result['closing_balance']) > AMOUNT_TOLERANCE and result['closing_balance'] > 0:
                result['balance_check'] = False
                diff = round(expected - result['closing_balance'], 2)
                result['errors'].append(f'四项资金不平: 差异={diff:.2f}')

            if ocr_used and not result['items']:
                result['items'] = _parse_ocr_column_items(all_text)
            if ocr_used and not result['items']:
                result['items'] = _parse_fixed_scan_items(all_text)
            if not ocr_used and not result['items']:
                result['items'] = _parse_template_text_items(all_text)
            if not ocr_used and not result['items']:
                result['items'] = _parse_ocr_column_items(all_text)
            if not ocr_used and not result['items']:
                result['items'] = _parse_items(all_text)
            if result['items']:
                _repair_lw_pdf_rows(result['items'], all_text)
            # 有些 PDF 能提取零散文字，但表格结构已经丢失。传统解析得不到
            # 明细时仍应交给千问，而不是保存一条空对账记录。
            if (
                not qwen_data
                and (
                    not result['items']
                    or _items_are_suspicious(result['items'])
                )
            ):
                try:
                    qwen_data = recognize_pdf_with_qwen(
                        pdf_path, 'statement', progress_callback=progress_callback
                    )
                except Exception as exc:
                    result['errors'].append(f'千问识别失败，保留本地解析结果: {exc}')
                    qwen_data = {}
                if qwen_data:
                    ocr_used = True
                    _merge_qwen_statement(result, qwen_data)
            if not result['items']:
                result['errors'].append('未能解析对账单明细行')

            if result['items']:
                calc_total = sum(it['amount_incl_tax'] for it in result['items'])
                calc_qty = sum(it['quantity'] for it in result['items'])
                result['total_quantity'] = int(calc_qty)
                if result['total_invoice_amount'] == 0 or ocr_used:
                    result['total_invoice_amount'] = round(calc_total, 2)
            _validate_item_roles(result)

        return result

    except Exception as e:
        result['errors'].append(f'PDF解析异常: {str(e)}')
        return result


def _merge_qwen_statement(result: dict, data: dict) -> None:
    text_fields = (
        'customer_name', 'customer_tax_id', 'supplier_name', 'supplier_code',
        'supplier_tax_id', 'statement_month', 'statement_date', 'usage_remark',
    )
    number_fields = (
        'settlement_days', 'opening_balance', 'current_payment',
        'closing_balance', 'delivered_unpaid', 'total_invoice_amount',
    )
    for key in text_fields:
        if data.get(key) not in (None, ''):
            result[key] = str(data[key]).strip()
    for key in number_fields:
        if data.get(key) not in (None, ''):
            result[key] = _to_float(data[key])
    items = []
    for index, source in enumerate(data.get('items') or [], start=1):
        qty = _to_float(source.get('quantity'))
        amount = _to_float(source.get('amount_incl_tax'))
        if qty == 0 and amount == 0:
            continue
        unit_price = _to_float(source.get('unit_price_incl_tax'))
        items.append({
            'seq': int(_to_float(source.get('seq')) or index),
            'customer_order_no': str(source.get('customer_order_no') or ''),
            'customer_material_code': _normalize_customer_material_code(
                source.get('customer_material_code')
            ),
            'delivery_no': str(source.get('delivery_no') or ''),
            'product_name': str(source.get('product_name') or ''),
            'specification': str(source.get('specification') or ''),
            'unit': str(source.get('unit') or ''),
            'quantity': qty,
            'unit_price_incl_tax': unit_price,
            'amount_incl_tax': amount,
            'delivery_date': _format_date(source.get('delivery_date')),
        })
    result['items'] = items
    result['statement_month'] = _format_month(result.get('statement_month'))
    if items:
        result['total_quantity'] = sum(item['quantity'] for item in items)
        result['total_invoice_amount'] = round(
            sum(item['amount_incl_tax'] for item in items), 2
        )
    result['raw_text'] = data.get('raw_text', '')
    result['column_mapping'] = data.get('column_mapping') or {}


def _ocr_pdf_text(pdf_path: str, max_pages=None, render_scale=3) -> str:
    """Use macOS Vision OCR for scanned PDFs."""
    try:
        import Foundation
        import Vision
        import pypdfium2 as pdfium
    except Exception:
        return ''

    texts = []
    try:
        pdf = pdfium.PdfDocument(pdf_path)
        for idx, page in enumerate(pdf):
            if max_pages is not None and idx >= max_pages:
                break
            with tempfile.NamedTemporaryFile(suffix='.png') as temp_image:
                page.render(scale=render_scale).to_pil().save(temp_image.name)
                page_text = _ocr_image_text(temp_image.name)
                if page_text:
                    texts.append(page_text)
    except Exception:
        return ''
    return '\n'.join(texts)


def _ocr_image_text(image_path: str) -> str:
    """Use macOS Vision for one local image without any cloud dependency."""
    try:
        import Foundation
        import Vision
        req = Vision.VNRecognizeTextRequest.alloc().init()
        req.setRecognitionLevel_(Vision.VNRequestTextRecognitionLevelAccurate)
        req.setRecognitionLanguages_(['zh-Hans', 'en-US'])
        req.setUsesLanguageCorrection_(True)
        url = Foundation.NSURL.fileURLWithPath_(str(Path(image_path).resolve()))
        handler = Vision.VNImageRequestHandler.alloc().initWithURL_options_(url, {})
        ok, _ = handler.performRequests_error_([req], None)
        if not ok:
            return ''
        texts = []
        for obs in req.results() or []:
            candidates = obs.topCandidates_(1)
            if candidates:
                texts.append(str(candidates[0].string()))
        return '\n'.join(texts)
    except Exception:
        return ''


def recognize_supplier_locally(file_path: str) -> str:
    """Read only the counterparty name locally for the confirmation step."""
    path = Path(file_path)
    text = (
        _ocr_pdf_text(str(path), max_pages=1, render_scale=1.7)
        if path.suffix.lower() == '.pdf'
        else _ocr_image_text(str(path))
    )
    return _supplier_from_local_text(text)


def recognize_excel_supplier_locally(file_path: str) -> str:
    """OCR company names embedded as logos in xlsx instead of guessing from filename."""
    try:
        workbook = load_workbook(file_path, data_only=True, read_only=False)
        for sheet in workbook.worksheets:
            for image in getattr(sheet, '_images', []):
                suffix = '.' + str(getattr(image, 'format', '') or 'png').lower()
                with tempfile.NamedTemporaryFile(suffix=suffix) as temp_image:
                    temp_image.write(image._data())
                    temp_image.flush()
                    supplier = _supplier_from_local_text(
                        _ocr_image_text(temp_image.name)
                    )
                    if supplier:
                        return supplier
    except Exception:
        return ''
    return ''


def _supplier_from_local_text(text: str) -> str:
    lines = [
        re.sub(r'\s+', '', line).strip()
        for line in str(text or '').splitlines()
        if str(line or '').strip()
    ]

    # Apple Vision may split a large rotated company title into two observations.
    # A frequent split is "东莞市创慧" + "子有限公司", where the small "电" is
    # missed at the boundary. Rebuild the complete "电子有限公司" suffix before
    # evaluating standalone company-name matches.
    for index, line in enumerate(lines[:-1]):
        following = lines[index + 1]
        if (
            following == '子有限公司'
            and len(re.findall(r'[\u4e00-\u9fff]', line)) >= 4
            and '骊威' not in line
        ):
            candidate = _clean_party_name(f'{line}电子有限公司')
            if _is_plausible_supplier_name(candidate):
                return candidate
        if (
            re.fullmatch(
                r'[\u4e00-\u9fff]{1,8}(?:有限责任公司|有限公司|公司)',
                following,
            )
            and 4 <= len(re.findall(r'[\u4e00-\u9fff]', line)) <= 20
            and not re.search(r'[：:，,。；;]', line)
            and '骊威' not in line
        ):
            candidate = _clean_party_name(f'{line}{following}')
            if _is_plausible_supplier_name(candidate):
                return candidate

    title_party = re.search(
        r'与\s*([\u4e00-\u9fffA-Za-z0-9（）()·]{2,}?(?:有限责任公司|有限公司|公司))',
        text,
    )
    if title_party and '骊威' not in title_party.group(1):
        candidate = _clean_party_name(title_party.group(1))
        if _is_plausible_supplier_name(candidate):
            return candidate
    companies = re.findall(
        r'[\u4e00-\u9fffA-Za-z0-9（）()·]{2,}?(?:有限责任公司|有限公司|公司)',
        text,
    )
    return next((
        candidate
        for name in companies
        if '骊威' not in name
        for candidate in [_clean_party_name(name)]
        if _is_plausible_supplier_name(candidate)
    ), '')


def _is_plausible_supplier_name(name: str) -> bool:
    """Reject OCR suffix fragments such as '子有限公司' and bare company types."""
    cleaned = _clean_party_name(str(name or ''))
    core = re.sub(r'(?:有限责任公司|有限公司|股份有限公司|公司)$', '', cleaned)
    chinese_core = ''.join(re.findall(r'[\u4e00-\u9fff]', core))
    return len(chinese_core) >= 4 and len(set(chinese_core)) >= 2


def parse_statement_xlsx(xlsx_path: str) -> dict:
    """从对账单 Excel 中提取客户信息、账期、汇总和明细。"""
    result = {
        'customer_name': '',
        'customer_tax_id': '',
        'supplier_name': '',
        'supplier_tax_id': '',
        'statement_month': '',
        'statement_date': '',
        'settlement_days': 30,
        'opening_balance': 0.0,
        'current_payment': 0.0,
        'closing_balance': 0.0,
        'delivered_unpaid': 0.0,
        'total_invoice_amount': 0.0,
        'total_quantity': 0,
        'items': [],
        'balance_check': True,
        'raw_text': '',
        'errors': []
    }

    try:
        wb = load_workbook(xlsx_path, data_only=True, read_only=True)
        sheet_rows = []
        text_parts = []
        for ws in wb.worksheets:
            text_parts.append(ws.title)
            rows = []
            for row in ws.iter_rows(values_only=True):
                values = [v for v in row if v is not None]
                if not values:
                    continue
                rows.append(row)
                text_parts.extend(str(v) for v in values)
            if rows:
                sheet_rows.append(rows)
        rows = max(sheet_rows, key=lambda value: len(_parse_xlsx_items(value)), default=[])
        return _populate_excel_result(result, rows, text_parts)
    except Exception as e:
        result['errors'].append(f'Excel解析异常: {str(e)}')
        return result


def parse_statement_xls(xls_path: str) -> dict:
    """读取旧版 Excel 97-2003 对账单。"""
    result = _empty_statement_result()
    try:
        import xlrd
        workbook = xlrd.open_workbook(xls_path)
        candidates = []
        text_parts = []
        for sheet in workbook.sheets():
            text_parts.append(sheet.name)
            rows = []
            for row_index in range(sheet.nrows):
                values = []
                for cell in sheet.row(row_index):
                    if cell.ctype == xlrd.XL_CELL_DATE:
                        values.append(xlrd.xldate_as_datetime(cell.value, workbook.datemode))
                    else:
                        values.append(cell.value)
                rows.append(tuple(values))
            candidates.append(rows)
            text_parts.extend(
                str(value) for row in rows for value in row if value not in (None, '')
            )
        rows = max(candidates, key=lambda value: len(_parse_xlsx_items(value)), default=[])
        return _populate_excel_result(result, rows, text_parts)
    except Exception as exc:
        result['errors'].append(f'旧版Excel解析异常: {exc}')
        return result


def _populate_excel_result(result: dict, rows: list, text_parts: list) -> dict:
    all_text = '\n'.join(text_parts)
    result['raw_text'] = all_text
    result['supplier_code'] = _find_adjacent_value(rows, ['供应商编码'])
    result['customer_name'] = _find_adjacent_value(rows, ['客户名称', '客户'])
    result['supplier_name'] = _find_adjacent_value(rows, ['供应商名称', '供应商'])
    result['supplier_tax_id'] = _find_adjacent_value(
        rows, ['供应商税号', '供应商纳税人识别号', '税号']
    )
    result['statement_month'] = _format_month(
        _find_adjacent_value(rows, ['对账月份', '账期', '月份'])
    )
    result['statement_date'] = _format_date(
        _find_adjacent_value(rows, ['制表日期', '对账日期'])
    )
    result['usage_remark'] = _find_adjacent_value(rows, ['用途备注', '用途', '付款用途'])
    settlement_days = _to_float(_find_adjacent_value(rows, ['结算天数', '月结天数']))
    if settlement_days:
        result['settlement_days'] = int(settlement_days)
    customer_match = re.search(r'客户[：:]\s*([^\n]+)', all_text)
    if customer_match:
        result['customer_name'] = result['customer_name'] or customer_match.group(1).strip()
    supplier_match = re.search(
        r'(?:供应商|供货商|供货单位|协力厂商)[：:]\s*([^\n|]+)', all_text
    )
    if supplier_match:
        result['supplier_name'] = result['supplier_name'] or supplier_match.group(1).strip()
    if not result['supplier_name']:
        company_candidates = re.findall(
            r'[\u4e00-\u9fffA-Za-z0-9（）()·]{2,}(?:有限责任公司|有限公司|公司)',
            all_text,
        )
        result['supplier_name'] = next(
            (
                _clean_party_name(name) for name in company_candidates
                if '骊威' not in name and '采购方' not in name and '客户' not in name
            ),
            '',
        )
        result['supplier_name'] = re.sub(
            r'^\d{4}\s*年\s*\d{1,2}\s*月', '', result['supplier_name']
        ).strip()
    if not result['supplier_name']:
        title_supplier = re.search(r'与([^流水\n]{2,40}?)(?=\d{4}\s*年)', all_text)
        if title_supplier:
            result['supplier_name'] = title_supplier.group(1).strip(' -—')
    month_match = re.search(r'(\d{4})\s*(?:年|[./-])\s*(\d{1,2})\s*月份?', all_text)
    if month_match:
        result['statement_month'] = result['statement_month'] or (
            f"{month_match.group(1)}-{int(month_match.group(2)):02d}"
        )
    result['items'] = _parse_xlsx_items(rows)
    if not result['items']:
        result['errors'].append('未能解析对账单明细行')
    else:
        result['total_quantity'] = sum(item['quantity'] for item in result['items'])
        result['total_invoice_amount'] = round(
            sum(item['amount_incl_tax'] for item in result['items']), 2
        )
        result['closing_balance'] = result['total_invoice_amount']
        if not result.get('statement_month'):
            item_months = sorted({
                str(item.get('delivery_date') or '')[:7]
                for item in result['items']
                if re.fullmatch(r'\d{4}-\d{1,2}-\d{1,2}', str(item.get('delivery_date') or ''))
            })
            if item_months:
                year, month = item_months[-1].split('-')
                result['statement_month'] = f"{year}-{int(month):02d}"
    if result.get('supplier_code') and result.get('statement_month'):
        result['statement_key'] = f"{result['supplier_code']}_{result['statement_month']}"
    _validate_item_roles(result)
    return result


def _parse_xlsx_items(rows: list) -> list:
    items = []
    header = None
    header_idx = -1

    for idx, row in enumerate(rows):
        values = [str(v).strip() if v is not None else '' for v in row]
        normalized_values = [_normalize_header(v) for v in values]
        header_hits = sum(
            name in normalized_values
            for name in ('数量', '金额', '单价', '物料编码', '订单号', '送货单号', '日期')
        )
        if '数量' in normalized_values and '金额' in normalized_values and header_hits >= 3:
            header = values
            header_idx = idx
            break
    if header is None:
        return items

    index = {}
    raw_headers = [str(value or '').strip() for value in header]
    # Some supplier statements use both “名称” and “品名”: in that layout
    # 名称 is the product name while 品名 contains the model/specification.
    # Normalizing both to 商品名称 silently overwrote the former column.
    split_name_and_spec = '名称' in raw_headers and '品名' in raw_headers
    for i, name in enumerate(header):
        raw_name = str(name or '').strip()
        if split_name_and_spec and raw_name == '名称':
            index['商品名称'] = i
            continue
        if split_name_and_spec and raw_name == '品名':
            index['规格型号'] = i
            continue
        normalized = _normalize_header(name)
        if normalized:
            index[normalized] = i
    required = ['数量', '金额']
    if any(name not in index for name in required):
        return items

    for row in rows[header_idx + 1:]:
        seq = _cell(row, index.get('序号'))
        row_text = ' '.join(str(value or '') for value in row)
        if re.search(r'(本月合计|总计|合计|应付款合计|以下空白)', row_text):
            break
        qty = _to_float(_cell(row, index.get('数量')))
        amount = _to_float(_cell(row, index.get('金额')))
        if qty <= 0 or amount <= 0:
            continue
        if isinstance(seq, (int, float)) and int(seq) > 0:
            seq_no = int(seq)
        else:
            seq_no = len(items) + 1

        row_values = [str(value or '') for value in row]
        customer_material_code = _normalize_customer_material_code(
            _cell(row, index.get('客户物料编码'))
        )
        if not customer_material_code:
            # 非标准表头时只按完整单元格的LW物料码补充，禁止从AHLW订单号
            # 或描述文本中截取一段内容冒充物料编码。
            customer_material_code = next(
                (
                    code for code in (
                        _normalize_customer_material_code(value)
                        for value in row_values
                    )
                    if code
                ),
                '',
            )

        item = {
            'seq': seq_no,
            'customer_order_no': str(_cell(row, index.get('客户订单号')) or '').strip(),
            'customer_material_code': customer_material_code,
            'delivery_no': str(_cell(row, index.get('送货单号')) or '').strip(),
            'delivery_date': _format_date(_cell(row, index.get('交货日期'))),
            'product_name': str(_cell(row, index.get('商品名称')) or '').strip(),
            'specification': str(_cell(row, index.get('规格型号')) or '').strip(),
            'quantity': qty,
            'unit': str(_cell(row, index.get('单位')) or 'PCS').strip(),
            'unit_price_incl_tax': round(_to_float(_cell(row, index.get('单价'))), 6),
            'amount_incl_tax': round(amount, 2),
        }
        if not item['product_name']:
            item['product_name'] = str(_cell(row, index.get('商品型号')) or '').strip()
        items.append(item)

    return items


def _repair_lw_pdf_rows(items: list, text: str) -> None:
    """Repair cells dropped by a PDF table extractor from its embedded text."""
    pattern = re.compile(
        r'(?P<order>20\d{2}\.\d{1,2}\.\d{1,2}\s+QQ群下单)\s+'
        r'(?P<delivery>\d{6,8})\s+'
        r'(?P<date>20\d{2}\.\d{1,2}\.\d{1,2})\s+'
        r'(?P<material>LW[A-Z0-9]{6,})\s+'
        r'(?P<description>.+?)\s+PCS\s+'
        r'(?P<qty>\d+(?:\.\d+)?)\s+'
        r'(?P<price>\d+(?:\.\d+)?)\s+'
        r'(?P<amount>\d+(?:\.\d+)?)',
        re.I | re.S,
    )
    extracted = []
    for match in pattern.finditer(text or ''):
        description = re.sub(r'\s+', ' ', match.group('description')).strip()
        name, _, specification = description.partition(' ')
        extracted.append({
            'customer_order_no': match.group('order'),
            'delivery_no': match.group('delivery'),
            'delivery_date': _format_date(match.group('date')),
            'customer_material_code': match.group('material').upper(),
            'product_name': name,
            'specification': specification,
            'quantity': _to_float(match.group('qty')),
            'unit_price_incl_tax': _to_float(match.group('price')),
            'amount_incl_tax': _to_float(match.group('amount')),
        })
    for item in items:
        if (
            item.get('customer_material_code')
            and item.get('product_name')
            and item.get('specification')
            and _to_float(item.get('unit_price_incl_tax')) > 0
        ):
            continue
        candidate = next((
            row for row in extracted
            if row['customer_order_no'] == item.get('customer_order_no')
            and row['delivery_no'] == item.get('delivery_no')
            and row['delivery_date'] == item.get('delivery_date')
            and abs(row['quantity'] - _to_float(item.get('quantity'))) < 1e-9
            and abs(row['amount_incl_tax'] - _to_float(item.get('amount_incl_tax'))) < 0.01
        ), None)
        if not candidate:
            continue
        for key in (
            'customer_material_code', 'product_name', 'specification',
            'unit_price_incl_tax',
        ):
            if item.get(key) in (None, '', 0, 0.0):
                item[key] = candidate[key]


def _merge_template_fields(result: dict, rows: list):
    result['supplier_code'] = result.get('supplier_code') or _find_adjacent_value(rows, ['供应商编码'])
    result['customer_name'] = result.get('customer_name') or _find_adjacent_value(rows, ['客户名称', '客户'])
    result['supplier_name'] = result.get('supplier_name') or _find_adjacent_value(rows, ['供应商名称', '供应商'])
    result['supplier_tax_id'] = result.get('supplier_tax_id') or _find_adjacent_value(rows, ['供应商税号', '供应商纳税人识别号', '税号'])
    result['statement_month'] = result.get('statement_month') or _format_month(_find_adjacent_value(rows, ['对账月份', '账期', '月份']))
    result['statement_date'] = result.get('statement_date') or _format_date(_find_adjacent_value(rows, ['制表日期', '对账日期']))
    settlement_days = _to_float(_find_adjacent_value(rows, ['结算天数', '月结天数']))
    if settlement_days:
        result['settlement_days'] = int(settlement_days)
    if result.get('supplier_code') and result.get('statement_month'):
        result['statement_key'] = f"{result['supplier_code']}_{result['statement_month']}"


def _parse_template_text_items(text: str) -> list:
    """Parse text PDFs generated from the unified Excel template."""
    items = []
    lines = [line.strip() for line in text.splitlines() if line.strip()]
    header_seen = False

    for line in lines:
        normalized_line = _normalize_header(line.replace(' ', ''))
        if '序号' in normalized_line and '数量' in normalized_line and '金额' in normalized_line:
            header_seen = True
            continue
        if not header_seen:
            continue
        if re.match(r'^(合计|填写说明|必填项)', line):
            break

        item = _parse_template_text_line(line)
        if item:
            item['seq'] = len(items) + 1
            items.append(item)

    return items


def _parse_template_text_line(line: str) -> dict:
    parts = [part for part in re.split(r'\s+', line.strip()) if part]
    if len(parts) < 8:
        return None
    if not re.match(r'^\d+$', parts[0]):
        return None

    date_idx = None
    for idx, part in enumerate(parts):
        if re.match(r'^\d{4}[-./]\d{1,2}[-./]\d{1,2}$', part):
            date_idx = idx
            break
    if date_idx is None or date_idx < 2:
        return None

    numeric_positions = []
    for idx, part in enumerate(parts):
        if _to_float(part) > 0:
            numeric_positions.append(idx)
    if len(numeric_positions) < 3:
        return None

    amount_idx = numeric_positions[-1]
    price_idx = numeric_positions[-2]
    qty_idx = numeric_positions[-3]
    if qty_idx <= date_idx:
        return None

    unit_idx = qty_idx - 1
    product_parts = parts[date_idx + 1:unit_idx]
    product_name = product_parts[0] if product_parts else ''
    specification = ' '.join(product_parts[1:]) if len(product_parts) > 1 else ''

    customer_material_code = _extract_customer_material_code(' '.join(parts))
    if not customer_material_code:
        return None

    return {
        'customer_order_no': parts[1] if len(parts) > 1 else '',
        'customer_material_code': customer_material_code,
        'supplier_material_code': parts[3] if len(parts) > 3 and date_idx > 4 else '',
        'delivery_no': parts[date_idx - 1] if date_idx >= 4 else '',
        'delivery_date': _format_date(parts[date_idx]),
        'product_name': product_name or specification or line[:30],
        'specification': specification,
        'quantity': _to_float(parts[qty_idx]),
        'unit': parts[unit_idx] if unit_idx >= 0 else 'PCS',
        'unit_price_incl_tax': round(_to_float(parts[price_idx]), 6),
        'amount_incl_tax': round(_to_float(parts[amount_idx]), 2),
    }


def _parse_ocr_column_items(text: str) -> list:
    """Parse scanned statement OCR output where columns are emitted top-to-bottom."""
    lines = [line.strip() for line in str(text or '').splitlines() if line.strip()]
    if not lines:
        return []

    material_codes = [_extract_customer_material_code(line) for line in lines]
    material_codes = [code for code in material_codes if code]
    delivery_nos = []
    for line in lines:
        normalized = line.replace('$', 'S').replace('＄', 'S').upper()
        m = re.search(r'\bS\d{6,8}\b', normalized)
        if m:
            delivery_nos.append(m.group(0))
    dates = [_format_date(m.group(1)) for line in lines for m in [re.search(r'(\d{4}[-/]\d{1,2}[-/]\d{1,2})', line)] if m]
    order_nos = []
    for line in lines:
        normalized = line.upper().replace('™', 'W').replace('LR', 'LW')
        m = re.search(r'\b(\d{6}-L[WIV]{1,2})\b', normalized)
        if m:
            order_nos.append(m.group(1).replace('LIV', 'LW').replace('LV', 'LW'))

    product_names = [
        line for line in _section_lines(lines, ['产品名称'], ['款项信息', '商品名称'])
        if re.search(r'[\u4e00-\u9fff]', line)
    ]
    specs = _section_lines(lines, ['商品名称'], ['本次对帐开票', '本次对账开票', '出库数量'])
    quantities = [_to_float(v) for v in _numeric_section(lines, ['出库数量'], ['单位'])]
    units = [line for line in _section_lines(lines, ['单位'], ['对帐日期', '对账日期', '含型单价', '含税单价']) if re.match(r'^[A-Za-z]+$', line)]
    prices = [_to_float(v) for v in _numeric_section(lines, ['含型单价', '含税单价', '含税单价'], ['Total', '合计'])]
    amounts = [_to_float(v) for v in _numeric_section(lines, ['Total', '合计'], ['每注', '备注', '路合同', '合同'])]

    if amounts and len(amounts) > len(material_codes):
        amounts = amounts[:len(material_codes)]
    count = max(len(material_codes), len(delivery_nos), len(dates), len(quantities), len(amounts))
    items = []
    for idx in range(count):
        code = _at(material_codes, idx)
        qty = _at(quantities, idx, 0)
        price = _at(prices, idx, 0)
        amount = _at(amounts, idx, 0)
        if not code or qty <= 0 or amount <= 0:
            continue
        items.append({
            'customer_order_no': _at(order_nos, idx, _at(order_nos, 0, '')),
            'customer_material_code': code,
            'delivery_no': _at(delivery_nos, idx, ''),
            'delivery_date': _at(dates, idx, ''),
            'product_name': _at(product_names, idx, _at(specs, idx, code)),
            'specification': _at(specs, idx, ''),
            'quantity': qty,
            'unit': _at(units, idx, 'PCS').upper(),
            'unit_price_incl_tax': round(price, 6),
            'amount_incl_tax': round(amount, 2),
        })
    return items


def _parse_fixed_scan_items(text: str) -> list:
    """Fallback for fixed vertical-column scans; keep uncertain values empty."""
    lines = [line.strip() for line in str(text or '').splitlines() if line.strip()]
    detail_block = _section_lines(lines, ['规格'], ['单位'])
    dates, names, specs = [], [], []
    index = 0
    while index < len(detail_block):
        match = re.search(r'(\d{4}[./-]\d{1,2}[./-]\d{1,2})', detail_block[index])
        if not match:
            index += 1
            continue
        dates.append(_format_date(match.group(1)))
        names.append(detail_block[index + 1] if index + 1 < len(detail_block) else '')
        specs.append(detail_block[index + 2] if index + 2 < len(detail_block) else '')
        index += 3
    amounts = [
        _to_float(value)
        for value in _numeric_section(
            lines, ['到货合计', '金额合计', '含税金额'], ['供方', '需方', '合同', '盖章']
        )
    ]
    count = min(len(dates), len(names), len(specs), len(amounts))
    if count <= 0:
        return []
    return [{
        'customer_order_no': '',
        'customer_material_code': '',
        'delivery_no': '',
        'delivery_date': dates[index],
        'product_name': names[index],
        'specification': specs[index],
        'quantity': 0,
        'unit': '',
        'unit_price_incl_tax': 0,
        'amount_incl_tax': round(amounts[index], 2),
    } for index in range(count)]


def _section_lines(lines, start_markers, end_markers):
    start = None
    for idx, line in enumerate(lines):
        if any(marker in line for marker in start_markers):
            start = idx + 1
            break
    if start is None:
        return []
    end = len(lines)
    for idx in range(start, len(lines)):
        if any(marker in lines[idx] for marker in end_markers):
            end = idx
            break
    return [line for line in lines[start:end] if line]


def _numeric_section(lines, start_markers, end_markers):
    values = []
    for line in _section_lines(lines, start_markers, end_markers):
        cleaned = line.replace('：', '').replace(':', '').replace(' ', '')
        if re.match(r'^\d+(?:\.\d+)?$', cleaned):
            values.append(cleaned)
    return values


def _at(values, index, default=''):
    return values[index] if index < len(values) else default


def _normalize_header(value) -> str:
    text = str(value or '').strip()
    text = re.sub(r'[\s*＊（）()]', '', text)
    if '数量' in text:
        return '数量'
    if '单价' in text or '价格' in text:
        return '单价'
    if '金额' in text and '不含税' not in text:
        return '金额'
    if '客户订单' in text or text in ('订单号', '订单编号', '订单号码', '订单PO#'):
        return '客户订单号'
    if '客户物料' in text or '客户货号' in text or text in ('物料编码', '物料编号', '编号'):
        return '客户物料编码'
    if text in ('日期', '送货时期', '送货时间', '送货日期', '出货日期', '交货日期', '出库日期', '发货日期', '单据日期'):
        return '交货日期'
    aliases = {
        '含税单价': '单价',
        '含税金额': '金额',
        '本次金额': '金额',
        '销售金额': '金额',
        '销售数量': '数量',
        '客户料号': '客户物料编码',
        '物料编码': '客户物料编码',
        '物料编号': '客户物料编码',
        '编号': '客户物料编码',
        '订单PO#': '客户订单号',
        '订单编号': '客户订单号',
        '订单号码': '客户订单号',
        '订单号': '客户订单号',
        '采购单号': '客户订单号',
        '单号': '送货单号',
        '规格': '规格型号',
        '品名规格': '规格型号',
        '型号': '规格型号',
        '品名': '商品名称',
        '名称': '商品名称',
        '产品名称': '商品名称',
        '出库单号': '送货单号',
        '发货单号': '送货单号',
        '日期': '交货日期',
        '送货时间': '交货日期',
        '送货日期': '交货日期',
        '出库日期': '交货日期',
        '发货日期': '交货日期',
        '供应商纳税人识别号': '供应商税号',
    }
    return aliases.get(text, text)


def _items_are_suspicious(items: list) -> bool:
    """Reject column-shifted rows instead of presenting invented financial data."""
    checked = 0
    invalid = 0
    for item in items:
        quantity = _to_float(item.get('quantity'))
        price = _to_float(item.get('unit_price_incl_tax'))
        amount = _to_float(item.get('amount_incl_tax'))
        if quantity <= 0 or amount <= 0:
            invalid += 1
            continue
        checked += 1
        expected = quantity * price
        tolerance = max(0.05, abs(amount) * 0.005)
        if price <= 0 or abs(expected - amount) > tolerance:
            invalid += 1
    return not checked or invalid > max(1, len(items) // 3)


def _validate_item_roles(result: dict) -> None:
    """Detect cross-column assignments; never repair one field with another field."""
    issues = []
    for index, item in enumerate(result.get('items') or [], start=1):
        order_no = str(item.get('customer_order_no') or '').strip()
        material_code = str(item.get('customer_material_code') or '').strip().upper()
        delivery_no = str(item.get('delivery_no') or '').strip()
        delivery_date = str(item.get('delivery_date') or '').strip()

        if order_no and re.fullmatch(r'LW[A-Z0-9]{6,}', order_no, re.I):
            issues.append(f'第{index}行采购单号疑似物料编码')
            item['customer_order_no'] = ''
        if material_code and re.match(r'^(?:AHLW[-_/]|\d{8}[-_.])', material_code, re.I):
            issues.append(f'第{index}行物料编码疑似采购单号')
            item['customer_material_code'] = ''
        if delivery_no and re.fullmatch(r'\d{4}[-/.]\d{1,2}[-/.]\d{1,2}', delivery_no):
            issues.append(f'第{index}行送货单号疑似日期')
            item['delivery_no'] = ''
        if delivery_date and not re.fullmatch(
            r'\d{4}-\d{1,2}(?:-\d{1,2})?', delivery_date
        ):
            issues.append(f'第{index}行日期格式异常')
            item['delivery_date'] = ''
        if order_no and delivery_no and order_no == delivery_no:
            issues.append(f'第{index}行采购单号与送货单号相同，请人工确认')

        quantity = _to_float(item.get('quantity'))
        price = _to_float(item.get('unit_price_incl_tax'))
        amount = _to_float(item.get('amount_incl_tax'))
        if quantity <= 0:
            issues.append(f'第{index}行数量缺失或无效')
        if price <= 0:
            issues.append(f'第{index}行含税单价缺失或无效')
        if amount <= 0:
            issues.append(f'第{index}行含税金额缺失或无效')
        if quantity > 0 and price > 0 and amount > 0:
            tolerance = max(0.05, abs(amount) * 0.005)
            if abs(quantity * price - amount) > tolerance:
                issues.append(f'第{index}行数量×单价与金额不一致')
    for issue in issues:
        if issue not in result['errors']:
            result['errors'].append(issue)


def _clean_party_name(value: str) -> str:
    text = str(value or '').strip()
    text = re.sub(r'\s+', ' ', text)
    for marker in ('对账月份', '客户名称', '供应商名称', '制表日期', '结算天数'):
        if marker in text:
            text = text.split(marker, 1)[0].strip()
    return text


def _normalize_customer_material_code(value) -> str:
    text = str(value or '').strip().upper()
    if not text:
        return ''
    text = text.replace('Ｌ', 'L').replace('Ｗ', 'W')
    text = re.sub(r'[^A-Z0-9]', '', text)
    replacements = (
        ('1W', 'LW'),
        ('IW', 'LW'),
        ('|W', 'LW'),
        ('LIV', 'LW'),
        ('LVV', 'LW'),
        ('LNW', 'LW'),
        ('LR', 'LW'),
    )
    for wrong, right in replacements:
        if text.startswith(wrong):
            text = right + text[len(wrong):]
            break
    match = re.fullmatch(r'LW(\d{5,12})', text)
    if match:
        return f"LW{match.group(1)}"
    return ''


def _extract_customer_material_code(text: str) -> str:
    direct = _normalize_customer_material_code(text)
    if direct:
        return direct
    compact = re.sub(r'\s+', '', str(text or '').upper())
    for match in re.finditer(r'(?:L|1|I|\|)\s*W\s*[\d\s]{5,16}', compact):
        code = _normalize_customer_material_code(match.group(0))
        if code:
            return code
    return ''


def _find_adjacent_value(rows, labels):
    normalized_labels = {_normalize_header(label) for label in labels}
    for row in rows:
        values = [str(v).strip() if v is not None else '' for v in row]
        for idx, value in enumerate(values):
            if not value:
                continue
            inline = re.match(r'^(.+?)[：:]\s*(.+)$', value)
            if inline and _normalize_header(inline.group(1)) in normalized_labels:
                return inline.group(2).strip()
            normalized = _normalize_header(value.rstrip('：:'))
            if normalized in normalized_labels:
                for next_idx in range(idx + 1, len(values)):
                    if values[next_idx]:
                        return values[next_idx].strip()
    return ''


def _cell(row, idx):
    if idx is None or idx >= len(row):
        return None
    return row[idx]


def _to_float(value) -> float:
    if value is None or value == '':
        return 0.0
    try:
        return float(str(value).replace(',', ''))
    except ValueError:
        return 0.0


def _format_date(value) -> str:
    if value is None:
        return ''
    if hasattr(value, 'strftime'):
        return value.strftime('%Y-%m-%d')
    return str(value).replace('/', '-').replace('.', '-')


def _format_month(value) -> str:
    if value is None or value == '':
        return ''
    if hasattr(value, 'strftime'):
        return value.strftime('%Y-%m')
    text = str(value).strip()
    m = re.search(r'(\d{4})[-./年]\s*(\d{1,2})', text)
    if m:
        return f"{m.group(1)}-{int(m.group(2)):02d}"
    return text


def _extract_amount(text: str, pattern: str) -> float:
    m = re.search(pattern, text)
    if m:
        try:
            return float(m.group(1).replace(',', ''))
        except (ValueError, IndexError):
            pass
    return 0.0


def _parse_items(text: str) -> list:
    items = []
    lines = text.split('\n')
    in_items = False

    for line in lines:
        stripped = line.strip()
        if re.search(r'序号.*?(?:出库日期|发货日期)', stripped) or \
           re.search(r'(?:物料编码|客户料号).*?(?:数量|金额)', stripped):
            in_items = True
            continue
        if not in_items:
            continue
        if re.match(r'^\s*(?:合\s*计|小\s*计|未税|本期|期初|期末|备注)', stripped):
            break
        if re.search(r'(?:未税总额|税额|含税总额|本次.*开票)', stripped):
            break
        if not stripped:
            continue
        item = _parse_item_line(stripped)
        if item:
            items.append(item)

    return items


def _parse_item_line(line: str) -> dict:
    customer_code = _extract_customer_material_code(line)

    dm = re.search(r'(S\d{7})', line)
    delivery_no = dm.group(1) if dm else ''

    om = re.search(r'(\d{6}-[A-Z]{2,4})', line)
    order_no = om.group(1) if om else ''

    dtm = re.search(r'(\d{4}[-./]\d{1,2}[-./]\d{1,2})', line)
    delivery_date = dtm.group(1).replace('/', '-').replace('.', '-') if dtm else ''

    qty_matches = re.findall(r'\b(\d{3,6})\b', line)
    quantity = 0
    for q in qty_matches:
        v = int(q)
        if 100 <= v <= 999999:
            quantity = v
            break

    price_matches = re.findall(r'\b(\d+\.\d{2,6})\b', line)
    unit_price = 0.0
    amount = 0.0
    if len(price_matches) >= 2:
        unit_price = float(price_matches[-2])
        amount = float(price_matches[-1])
    elif len(price_matches) == 1:
        amount = float(price_matches[0])

    name_parts = re.findall(r'[\u4e00-\u9fff]+', line)
    product_name = ' '.join(name_parts) if name_parts else line[:30]

    if customer_code and quantity > 0 and (unit_price > 0 or amount > 0):
        if amount == 0 and unit_price > 0:
            amount = round(quantity * unit_price, 2)
        return {
            'customer_order_no': order_no,
            'customer_material_code': customer_code,
            'delivery_no': delivery_no,
            'delivery_date': delivery_date,
            'product_name': product_name,
            'quantity': quantity,
            'unit': 'PCS',
            'unit_price_incl_tax': round(unit_price, 6),
            'amount_incl_tax': round(amount, 2),
        }
    return None
