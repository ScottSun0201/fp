"""
REQ-030: CSV/Excel 导出工具模块
提供发票、对账单、匹配结果的 CSV 和 Excel 导出功能。
"""

import csv
import io

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

# UTF-8 BOM，确保 Excel 正确识别中文编码
CSV_BOM = "\ufeff"

# ---------- 表头定义 ----------

INVOICE_HEADERS = [
    "发票号码", "开票日期", "购买方", "销售方",
    "不含税金额", "税额", "价税合计", "状态",
]

STATEMENT_HEADERS = [
    "对账月份", "客户名称", "供应商", "期初未回款",
    "本期开票", "本期付款", "期末未回款", "状态",
]

MATCH_HEADERS = [
    "发票号码", "发票金额", "对账单月份", "对账单金额",
    "匹配分数", "匹配等级", "差异金额",
]

# ---------- 表头样式（Excel 用） ----------

_HEADER_FONT = Font(bold=True, size=11)
_HEADER_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
_HEADER_ALIGN = Alignment(horizontal="center")


# ============================================================
# 内部工具函数
# ============================================================

def _rows_to_csv(headers, rows):
    """将表头和数据行写入带 BOM 的 CSV 字符串。"""
    buf = io.StringIO()
    buf.write(CSV_BOM)
    writer = csv.writer(buf)
    writer.writerow(headers)
    writer.writerows(rows)
    return buf.getvalue()


def _rows_to_excel(headers, rows, sheet_title="Sheet1"):
    """将表头和数据行写入 openpyxl Workbook 并返回。"""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title

    # 写入表头并设置样式
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN

    # 写入数据行
    for row_idx, row in enumerate(rows, start=2):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # 自动调整列宽
    for col_idx, header in enumerate(headers, start=1):
        max_len = len(str(header))
        for row_idx in range(2, len(rows) + 2):
            cell_value = ws.cell(row=row_idx, column=col_idx).value
            if cell_value is not None:
                max_len = max(max_len, len(str(cell_value)))
        # 中文字符大约占 2 个字符宽度，这里做粗略补偿
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = max_len + 4

    return wb


def _get(obj, key, default=""):
    """
    从 dict 或对象中安全取值。
    支持字典键访问和对象属性访问两种方式。
    """
    if isinstance(obj, dict):
        return obj.get(key, default)
    return getattr(obj, key, default)


# ============================================================
# 发票导出
# ============================================================

def _invoice_to_row(inv):
    """将单条发票数据转为行列表。"""
    return [
        _get(inv, "invoice_number", _get(inv, "发票号码")),
        _get(inv, "invoice_date", _get(inv, "开票日期")),
        _get(inv, "buyer", _get(inv, "购买方")),
        _get(inv, "seller", _get(inv, "销售方")),
        _get(inv, "amount_excl_tax", _get(inv, "不含税金额")),
        _get(inv, "tax_amount", _get(inv, "税额")),
        _get(inv, "total_amount", _get(inv, "价税合计")),
        _get(inv, "status", _get(inv, "状态")),
    ]


def export_invoices_csv(invoices):
    """
    导出发票列表为 CSV 字符串（含 BOM）。
    参数:
        invoices: 发票数据列表，每项可以是 dict 或对象。
    返回:
        带 UTF-8 BOM 的 CSV 字符串。
    """
    rows = [_invoice_to_row(inv) for inv in invoices]
    return _rows_to_csv(INVOICE_HEADERS, rows)


def export_invoices_excel(invoices):
    """
    导出发票列表为 openpyxl Workbook。
    参数:
        invoices: 发票数据列表，每项可以是 dict 或对象。
    返回:
        openpyxl.Workbook 实例。
    """
    rows = [_invoice_to_row(inv) for inv in invoices]
    return _rows_to_excel(INVOICE_HEADERS, rows, sheet_title="发票列表")


# ============================================================
# 对账单导出
# ============================================================

def _statement_to_row(stmt):
    """将单条对账单数据转为行列表。"""
    return [
        _get(stmt, "statement_month", _get(stmt, "对账月份")),
        _get(stmt, "customer_name", _get(stmt, "客户名称")),
        _get(stmt, "supplier", _get(stmt, "供应商")),
        _get(stmt, "opening_balance", _get(stmt, "期初未回款")),
        _get(stmt, "current_invoiced", _get(stmt, "本期开票")),
        _get(stmt, "current_paid", _get(stmt, "本期付款")),
        _get(stmt, "closing_balance", _get(stmt, "期末未回款")),
        _get(stmt, "status", _get(stmt, "状态")),
    ]


def export_statements_csv(statements):
    """
    导出对账单列表为 CSV 字符串（含 BOM）。
    参数:
        statements: 对账单数据列表，每项可以是 dict 或对象。
    返回:
        带 UTF-8 BOM 的 CSV 字符串。
    """
    rows = [_statement_to_row(stmt) for stmt in statements]
    return _rows_to_csv(STATEMENT_HEADERS, rows)


def export_statements_excel(statements):
    """
    导出对账单列表为 openpyxl Workbook。
    参数:
        statements: 对账单数据列表，每项可以是 dict 或对象。
    返回:
        openpyxl.Workbook 实例。
    """
    rows = [_statement_to_row(stmt) for stmt in statements]
    return _rows_to_excel(STATEMENT_HEADERS, rows, sheet_title="对账单")


# ============================================================
# 匹配结果导出
# ============================================================

def _match_to_row(match):
    """将单条匹配结果转为行列表。"""
    return [
        _get(match, "invoice_number", _get(match, "发票号码")),
        _get(match, "invoice_amount", _get(match, "发票金额")),
        _get(match, "statement_month", _get(match, "对账单月份")),
        _get(match, "statement_amount", _get(match, "对账单金额")),
        _get(match, "match_score", _get(match, "匹配分数")),
        _get(match, "match_level", _get(match, "匹配等级")),
        _get(match, "difference", _get(match, "差异金额")),
    ]


def export_match_results_csv(results):
    """
    导出匹配结果为 CSV 字符串（含 BOM）。
    参数:
        results: 匹配结果列表，每项可以是 dict 或对象。
    返回:
        带 UTF-8 BOM 的 CSV 字符串。
    """
    rows = [_match_to_row(m) for m in results]
    return _rows_to_csv(MATCH_HEADERS, rows)
